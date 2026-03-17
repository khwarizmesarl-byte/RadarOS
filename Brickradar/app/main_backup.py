import os
import re
import json
import sqlite3
from dataclasses import dataclass
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional, Tuple
from concurrent.futures import ThreadPoolExecutor, as_completed

import time
import random
import httpx
from bs4 import BeautifulSoup

# Load .env file manually (no dotenv dependency needed)
def _load_env_file():
    env_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env")
    if not os.path.exists(env_path):
        print(f"[ENV] No .env file found at {env_path}")
        return
    try:
        with open(env_path, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("#") or "=" not in line:
                    continue
                key, _, val = line.partition("=")
                key = key.strip()
                val = val.strip().strip('"').strip("'")
                if key and key not in os.environ:
                    os.environ[key] = val
        print(f"[ENV] Loaded .env from {env_path}")
    except Exception as e:
        print(f"[ENV] Error reading .env: {e}")

_load_env_file()

ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY", "")
GROQ_API_KEY = os.getenv("GROQ_API_KEY", "")

from fastapi import FastAPI, Request
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from fastapi.staticfiles import StaticFiles


# ----------------------------
# Config
# ----------------------------

APP_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(APP_DIR, "lego_tracker.db")
TEMPLATES_DIR = os.path.join(APP_DIR, "templates")

TEMPLATE_FILE = os.getenv("LEGO_TEMPLATE", "dashboard.html")

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                  "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0 Safari/537.36",
    "Accept": "text/html,application/json;q=0.9,*/*;q=0.8",
}

SHOPIFY_PAGE_LIMIT = 250
BRICKING_VAT_MULTIPLIER = 1.11  # 11% VAT
DEFAULT_TIMEOUT = 30.0
SHOPIFY_429_RETRIES = 4          # max retries on rate-limit
SHOPIFY_429_BACKOFF = [5, 10, 20, 40]  # seconds to wait between retries — total ~75s max


def _shopify_get(client: httpx.Client, url: str, store_name: str):
    """GET with automatic retry on HTTP 429 (rate limited)."""
    for attempt, wait in enumerate([0] + SHOPIFY_429_BACKOFF):
        if wait:
            jitter = random.uniform(0, wait * 0.3)
            print(f"[{store_name}] 429 — waiting {wait:.0f}s before retry {attempt}/{SHOPIFY_429_RETRIES}")
            time.sleep(wait + jitter)
        try:
            r = client.get(url)
        except Exception as e:
            print(f"[{store_name}] fetch error: {e}")
            return None
        if r.status_code == 429:
            if attempt >= SHOPIFY_429_RETRIES:
                print(f"[{store_name}] HTTP 429 — giving up after {attempt} retries")
                return r
            continue  # retry
        return r
    return None

# Preferred store order in the grid
STORE_ORDER = ["Brickmania", "BRICKSHOP", "Bricking", "PlayOne"]


def order_stores(stores: List[str]) -> List[str]:
    return sorted([s for s in stores if s], key=lambda s: s.lower())


# Shopify stores
# Collections that contain new/latest products per store
NEW_ARRIVAL_COLLECTIONS: Dict[str, str] = {
    "Brickmania": "new-arrivals",
    "Bricking":   "2026-releases",
}

SHOPIFY_STORES = {
    "Brickmania": {
        "base_url": "https://thebrickmania.com",
        "vat_multiplier": 1.00,
    },
    # BRICKSHOP removed from Shopify — it's WooCommerce, scraped separately
    "Bricking": {
        "base_url": "https://bricking.com",
        "vat_multiplier": BRICKING_VAT_MULTIPLIER,
    },
    "KLAPTAP": {
        "base_url": "https://klaptap.com",
        "vat_multiplier": 1.00,
        "lego_only": True,  # Filter by title since vendor is "Klaptap" not "Lego"
    },
}

# PlayOne is NOT Shopify. We scrape HTML.
PLAYONE_LISTING_URLS = [
    "https://playone.com.lb/brands/lego/",
]

# BRICKSHOP is WooCommerce — scrape HTML listing pages
BRICKSHOP_LISTING_URL = "https://brickshop.me/shop/"

# BigCommerce stores — scraped via generic HTML scraper
BIGCOMMERCE_STORES = {
    "Ayoub Computers": {
        "base_url": "https://ayoubcomputers.com",
        "category_slug": "lego",
        "lego_only": True,
        "vat_multiplier": 1.0,
    },
}

# ── Store Logo Fetching ────────────────────────────────────────────────────────

def fetch_store_logo(base_url: str, store_name: str) -> Optional[str]:
    """Fetch store logo/favicon and save to static/logos/{store_name}.png
    Returns the relative path if saved, None if failed."""
    import urllib.parse, base64 as b64mod

    static_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "static", "logos")
    os.makedirs(static_dir, exist_ok=True)

    safe_name = re.sub(r"[^a-zA-Z0-9_-]", "_", store_name)
    save_path = os.path.join(static_dir, f"{safe_name}.png")

    # Already fetched
    if os.path.exists(save_path):
        return f"/static/logos/{safe_name}.png"

    try:
        with httpx.Client(timeout=10, follow_redirects=True, headers=HEADERS) as client:
            r = client.get(base_url.rstrip("/") + "/")
            if r.status_code != 200:
                return None
            soup = BeautifulSoup(r.text, "lxml")

            logo_url = None

            # 1. og:image
            og = soup.select_one('meta[property="og:image"]')
            if og and og.get("content"):
                logo_url = og["content"]

            # 2. apple-touch-icon
            if not logo_url:
                atl = soup.select_one('link[rel*="apple-touch-icon"]')
                if atl and atl.get("href"):
                    logo_url = atl["href"]

            # 3. shortcut icon / favicon
            if not logo_url:
                fav = soup.select_one('link[rel*="icon"]')
                if fav and fav.get("href"):
                    logo_url = fav["href"]

            # 4. /favicon.ico fallback
            if not logo_url:
                logo_url = base_url.rstrip("/") + "/favicon.ico"

            # Make absolute
            if logo_url and logo_url.startswith("//"):
                logo_url = "https:" + logo_url
            elif logo_url and not logo_url.startswith("http"):
                logo_url = urllib.parse.urljoin(base_url, logo_url)

            # Download logo
            lr = client.get(logo_url)
            if lr.status_code == 200 and len(lr.content) > 100:
                # Convert to PNG if needed using basic save
                content_type = lr.headers.get("content-type", "")
                # Save raw bytes (browser handles most image types)
                ext = ".png"
                if "svg" in content_type:
                    ext = ".svg"
                elif "ico" in content_type or logo_url.endswith(".ico"):
                    ext = ".ico"
                elif "jpeg" in content_type or "jpg" in content_type:
                    ext = ".jpg"
                elif "webp" in content_type:
                    ext = ".webp"

                save_path_ext = os.path.join(static_dir, f"{safe_name}{ext}")
                with open(save_path_ext, "wb") as f:
                    f.write(lr.content)
                return f"/static/logos/{safe_name}{ext}"
    except Exception as e:
        print(f"[Logo] Failed to fetch logo for {store_name}: {e}")
    return None


def get_all_store_logos() -> dict:
    """Return dict of store_name -> logo_url for all enabled stores."""
    logos = {}
    conn = db_connect()
    cur = conn.cursor()
    cur.execute("SELECT name, base_url FROM stores WHERE enabled=1")
    db_stores = cur.fetchall()
    conn.close()

    # Hardcoded stores (built-in, not in DB)
    hardcoded_urls = {
        "Brickmania":    "https://thebrickmania.com",
        "Bricking":      "https://bricking.com",
        "KLAPTAP":       "https://klaptap.com",
        "BRICKSHOP":     "https://brickshop.me",
        "PlayOne":       "https://playone.com.lb",
    }
    for bc_name, bc_cfg in BIGCOMMERCE_STORES.items():
        hardcoded_urls[bc_name] = bc_cfg["base_url"]

    all_stores = {r["name"]: r["base_url"] for r in db_stores}
    all_stores.update(hardcoded_urls)

    static_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "static", "logos")
    os.makedirs(static_dir, exist_ok=True)

    for name, url in all_stores.items():
        safe_name = re.sub(r"[^a-zA-Z0-9_-]", "_", name)
        # Check for any extension
        for ext in [".png", ".jpg", ".svg", ".ico", ".webp", ".avif"]:
            p = os.path.join(static_dir, f"{safe_name}{ext}")
            if os.path.exists(p):
                logos[name] = f"/static/logos/{safe_name}{ext}"
                break
    return logos




# ----------------------------
# Helpers
# ----------------------------

def utc_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def safe_float(x: Any) -> Optional[float]:
    try:
        if x is None:
            return None
        if isinstance(x, (int, float)):
            return float(x)
        s = str(x).strip().replace(",", "")
        if not s:
            return None
        return float(s)
    except Exception:
        return None


def extract_item_number(text: str) -> Optional[str]:
    if not text:
        return None

    t = str(text)

    m = re.search(r"Item\s*#\s*([0-9]{4,7})", t, flags=re.IGNORECASE)
    if m:
        return m.group(1)

    # SKU like "10345-LEGO" or "10345-1" — number before a dash followed by letters or digits
    m = re.search(r"\b([0-9]{4,7})-[A-Za-z0-9]+\b", t)
    if m:
        return m.group(1)

    m = re.search(r"\b([0-9]{4,7})\b", t)
    if m:
        return m.group(1)

    return None


def normalize_theme_category_from_shopify(prod: Dict[str, Any]) -> Tuple[str, str]:
    """
    Fill theme/category so filters actually work.
    - theme: product_type (usually meaningful)
    - category: derived from product_type (remove "LEGO "), else vendor left, else "LEGO" if title says LEGO
    """
    title = (prod.get("title") or "").strip()
    vendor = (prod.get("vendor") or "").strip()
    product_type = (prod.get("product_type") or "").strip()

    # vendor often like "Lego / Icons"
    vparts = [p.strip() for p in vendor.split("/") if p.strip()]
    vendor_left = vparts[0] if len(vparts) >= 1 else ""
    vendor_right = vparts[1] if len(vparts) >= 2 else ""

    # theme
    theme = product_type or vendor_right or ""

    # category
    category = ""
    if product_type:
        pt = re.sub(r"(?i)^lego\s+", "", product_type).strip()
        if pt:
            category = pt.split()[0].strip()

    if not category:
        category = vendor_left

    if not category and "LEGO" in title.upper():
        category = "LEGO"

    return (theme or "").strip(), (category or "").strip()


def normalize_theme_category_from_playone(title: str) -> Tuple[str, str]:
    category = "LEGO"
    m = re.match(r"^\s*([A-Z][A-Z0-9&\-\s]{2,20})\b", title.strip())
    theme = ""
    if m:
        candidate = m.group(1).strip()
        if candidate.upper() != "LEGO":
            theme = candidate.title()
    return theme, category


@dataclass
class StoreOffer:
    price: Optional[float]
    availability: str
    link: str
    discount_pct: Optional[int] = None
    stock_qty: Optional[int] = None


def compute_discount_pct(price: Optional[float], compare_at: Optional[float]) -> Optional[int]:
    if price is None or compare_at is None:
        return None
    if compare_at <= 0:
        return None
    if price >= compare_at:
        return None
    pct = int(round((1.0 - (price / compare_at)) * 100))
    return pct if pct > 0 else None


# Known brand names that appear as the first word(s) in product titles
KNOWN_BRANDS = [
    "LEGO", "Mould King", "Blokees", "Nifeliz", "CaDA", "Reobrix",
    "Lumibricks", "JAKI", "Jaki", "LOZ", "GULY", "Guly", "Brickmania",
]

def normalize_brand_from_vendor_title(vendor: str, title: str) -> str:
    """
    Extract brand from title first (most reliable), then fall back to vendor.
    Brand is always the first word(s) of the product title.
    """
    # If vendor explicitly says LEGO, trust it immediately (case-insensitive, strip ® and spaces)
    v = (vendor or "").strip()
    if re.sub(r"[®\s/]", "", v).upper() == "LEGO":
        return "LEGO"
    # Also handle "Lego / Icons", "lego/city", etc. — first segment is LEGO
    v_first = v.split("/")[0].strip()
    if re.sub(r"[®\s]", "", v_first).upper() == "LEGO":
        return "LEGO"

    t = (title or "").strip()

    # Check known multi-word brands first (e.g. "Mould King")
    for brand in KNOWN_BRANDS:
        if t.lower().startswith(brand.lower()):
            return brand.upper()

    # Extract first CamelCase/ALL-CAPS word from title as brand
    m = re.match(r'^([A-Z][A-Za-z0-9&-]+)', t)
    if m:
        candidate = m.group(1)
        # Skip if it looks like a set number or generic word
        if not candidate.isdigit() and candidate.upper() not in ("THE", "A", "AN"):
            return candidate.upper()

    # Fall back to vendor field
    if v:
        return v.split("/")[0].strip().upper()

    return "UNKNOWN"


# ----------------------------
# DB
# ----------------------------

def db_connect() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def db_init() -> None:
    conn = db_connect()
    cur = conn.cursor()

    cur.execute("""
    CREATE TABLE IF NOT EXISTS snapshots (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        captured_at TEXT NOT NULL,
        store TEXT NOT NULL,
        item_number TEXT NOT NULL,
        title TEXT,
        theme TEXT,
        category TEXT,
        brand TEXT,
        price REAL,
        availability TEXT,
        link TEXT,
        image_url TEXT,
        images_json TEXT,
        compare_at REAL,
        stock_qty INTEGER
    )
    """)

    # Migrate existing DB
    try:
        cur.execute("ALTER TABLE snapshots ADD COLUMN brand TEXT")
    except Exception:
        pass
    try:
        cur.execute("ALTER TABLE snapshots ADD COLUMN images_json TEXT")
    except Exception:
        pass
    try:
        cur.execute("ALTER TABLE snapshots ADD COLUMN stock_qty INTEGER")
    except Exception:
        pass

    cur.execute("""
    CREATE TABLE IF NOT EXISTS alerts (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        created_at TEXT NOT NULL,
        item_number TEXT NOT NULL,
        store TEXT NOT NULL,
        title TEXT,
        old_price REAL,
        new_price REAL,
        alert_type TEXT NOT NULL DEFAULT 'price_change',
        unread INTEGER NOT NULL DEFAULT 1
    )
    """)

    # Migrate: add alert_type if missing
    try:
        cur.execute("ALTER TABLE alerts ADD COLUMN alert_type TEXT NOT NULL DEFAULT 'price_change'")
    except Exception:
        pass

    cur.execute("""
    CREATE TABLE IF NOT EXISTS meta (
        k TEXT PRIMARY KEY,
        v TEXT
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS stores (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT UNIQUE NOT NULL,
        base_url TEXT NOT NULL,
        platform TEXT NOT NULL DEFAULT 'shopify',
        vat_multiplier REAL NOT NULL DEFAULT 1.0,
        new_arrivals_collection TEXT,
        collection_slug TEXT,
        enabled INTEGER NOT NULL DEFAULT 1,
        product_count INTEGER DEFAULT 0,
        last_scraped TEXT,
        added_at TEXT NOT NULL DEFAULT (datetime('now'))
    )
    """)

    # Migrate: add new_arrivals_collection if missing
    try:
        cur.execute("ALTER TABLE stores ADD COLUMN new_arrivals_collection TEXT")
    except Exception:
        pass
    try:
        cur.execute("ALTER TABLE stores ADD COLUMN collection_slug TEXT")
    except Exception:
        pass
    try:
        cur.execute("ALTER TABLE stores ADD COLUMN lego_only INTEGER NOT NULL DEFAULT 0")
    except Exception:
        pass

    cur.execute("""
    CREATE TABLE IF NOT EXISTS radarlist (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        item_number TEXT NOT NULL UNIQUE,
        title TEXT,
        brand TEXT,
        theme TEXT,
        added_at TEXT NOT NULL DEFAULT (datetime('now')),
        added_price REAL,
        added_store TEXT
    )
    """)

    conn.commit()
    conn.close()


def meta_set(key: str, value: str) -> None:
    conn = db_connect()
    cur = conn.cursor()
    cur.execute("INSERT INTO meta(k,v) VALUES(?,?) ON CONFLICT(k) DO UPDATE SET v=excluded.v", (key, value))
    conn.commit()
    conn.close()


def meta_get(key: str) -> Optional[str]:
    conn = db_connect()
    cur = conn.cursor()
    cur.execute("SELECT v FROM meta WHERE k=?", (key,))
    row = cur.fetchone()
    conn.close()
    return row["v"] if row else None


def alerts_unread_count() -> int:
    conn = db_connect()
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) AS c FROM alerts WHERE unread=1")
    row = cur.fetchone()
    conn.close()
    return int(row["c"]) if row else 0


def alerts_mark_read() -> None:
    conn = db_connect()
    cur = conn.cursor()
    cur.execute("UPDATE alerts SET unread=0 WHERE unread=1")
    conn.commit()
    conn.close()


# ----------------------------
# Fetchers
# ----------------------------

def fetch_new_arrival_items(store_name: str, base_url: str) -> set:
    """
    Fetch item numbers from a store new arrivals / latest releases collection.
    Returns a set of item_numbers that are currently in that collection.
    """
    collection_slug = NEW_ARRIVAL_COLLECTIONS.get(store_name)
    if not collection_slug:
        return set()

    new_items: set = set()
    with httpx.Client(timeout=DEFAULT_TIMEOUT, follow_redirects=True, headers=HEADERS) as client:
        page = 1
        while True:
            url = f"{base_url.rstrip('/')}/collections/{collection_slug}/products.json?limit={SHOPIFY_PAGE_LIMIT}&page={page}"
            r = _shopify_get(client, url, store_name)
            if r is None or r.status_code != 200:
                break
            products = r.json().get("products") or []
            if not products:
                break
            for prod in products:
                title = (prod.get("title") or "").strip()
                item_number = extract_item_number(title)
                if not item_number:
                    variants = prod.get("variants") or []
                    for vv in variants:
                        item_number = extract_item_number((vv.get("sku") or "").strip())
                        if item_number:
                            break
                if item_number:
                    new_items.add(item_number)
            if len(products) < SHOPIFY_PAGE_LIMIT:
                break
            page += 1

    print(f"[{store_name}/new-arrivals] found {len(new_items)} new items")
    return new_items

def fetch_shopify_store(store_name: str, base_url: str, vat_multiplier: float = 1.0, new_items: Optional[set] = None, collection_slug: Optional[str] = None, lego_only: bool = False) -> Dict[str, Dict[str, Any]]:
    """
    Returns dict keyed by item_number -> normalized product record with store offer.
    new_items: set of item_numbers from the store new arrivals collection.
    collection_slug: if set, only scrape products from this collection (e.g. "lego").
    """
    out: Dict[str, Dict[str, Any]] = {}

    with httpx.Client(timeout=DEFAULT_TIMEOUT, follow_redirects=True, headers=HEADERS) as client:
        page = 1
        while True:
            if collection_slug:
                url = f"{base_url.rstrip('/')}/collections/{collection_slug}/products.json?limit={SHOPIFY_PAGE_LIMIT}&page={page}"
            else:
                url = f"{base_url.rstrip('/')}/products.json?limit={SHOPIFY_PAGE_LIMIT}&page={page}"
            r = _shopify_get(client, url, store_name)
            if r is None or r.status_code != 200:
                if r is not None:
                    print(f"[{store_name}] HTTP {r.status_code} for {url}")
                break

            data = r.json()
            products = data.get("products") or []
            if not products:
                break

            for prod in products:
                title = (prod.get("title") or "").strip()
                handle = (prod.get("handle") or "").strip()
                link = f"{base_url.rstrip('/')}/products/{handle}" if handle else base_url

                item_number = extract_item_number(title)

                # fallback: SKU from variants
                if not item_number:
                    variants = prod.get("variants") or []
                    for vv in variants:
                        sku = (vv.get("sku") or "").strip()
                        item_number = extract_item_number(sku)
                        if item_number:
                            break

                # fallback: extract from handle (e.g. "lego-76784-wednesday-black-dahlia" → "76784")
                if not item_number:
                    item_number = extract_item_number(handle)

                if not item_number:
                    shopify_id = prod.get("id")
                    if shopify_id:
                        item_number = f"SID{shopify_id}"
                    else:
                        continue

                variants = prod.get("variants") or []
                v0 = variants[0] if variants else {}

                raw_price = safe_float(v0.get("price"))
                raw_compare = safe_float(v0.get("compare_at_price"))

                price = (raw_price * vat_multiplier) if raw_price is not None else None
                compare_at = (raw_compare * vat_multiplier) if raw_compare is not None else None

                available = bool(v0.get("available", True))
                availability = "In stock" if available else "Out of stock"
                # Sum inventory across all variants
                stock_qty = None
                if variants:
                    try:
                        stock_qty = max(sum(int(v.get("inventory_quantity") or 0) for v in variants), 0)
                    except Exception:
                        stock_qty = None

                images = prod.get("images") or []
                image_url = ""
                image_list = []
                if images and isinstance(images, list):
                    for img in images[:4]:
                        src = (img.get("src") or "").strip()
                        if src:
                            image_list.append(src)
                    image_url = image_list[0] if image_list else ""

                theme, category = normalize_theme_category_from_shopify(prod)

                discount_pct = compute_discount_pct(price, compare_at)

                vendor_string = (prod.get("vendor") or "").strip()
                brand = normalize_brand_from_vendor_title(vendor_string, title)

                # Skip non-LEGO products if store is configured as lego_only.
                # Only trust the vendor field — title-extracted brand is too noisy
                # on general toy/lifestyle stores (e.g. "BATTERY", "BAMBOO", "POLICE").
                if lego_only and brand != "LEGO":
                    v_check = (vendor_string or "").strip()
                    v_check_first = v_check.split("/")[0].strip()
                    vendor_is_lego = (
                        re.sub(r"[®\s/]", "", v_check).upper() == "LEGO"
                        or re.sub(r"[®\s]", "", v_check_first).upper() == "LEGO"
                    )
                    if vendor_is_lego:
                        brand = "LEGO"
                    else:
                        continue

                # Flag as new if item is in the store new arrivals collection
                is_new = bool(new_items and item_number in new_items)

                out[item_number] = {
                    "item_number": item_number,
                    "title": title,
                    "theme": theme,
                    "category": category,
                    "image_url": image_url,
                    "image_list": image_list,
                    "vendor": vendor_string,
                    "brand": brand,
                    "compare_at": compare_at,
                    "is_new": is_new,
                    "stores": {
                        store_name: StoreOffer(
                            price=price,
                            availability=availability,
                            link=link,
                            discount_pct=discount_pct,
                            stock_qty=stock_qty,
                        )
                    },
                }

            if len(products) < SHOPIFY_PAGE_LIMIT:
                break
            page += 1

    print(f"[{store_name}] scraped {len(out)} items (lego_only={lego_only})")
    return out



def fetch_brickshop() -> Dict[str, Dict[str, Any]]:
    """
    Scrape BRICKSHOP (WooCommerce) by iterating all theme category pages.
    Product cards are in <li> elements inside <ul class="products">.
    Real image is the second <img> in the card (first is lazy SVG placeholder).
    Item number extracted from product URL slug (most reliable).
    """
    out: Dict[str, Dict[str, Any]] = {}
    BASE = "https://brickshop.me"

    # All theme categories to scrape (from their nav menu)
    CATEGORIES = [
        ("architecture", "Architecture"),
        ("art", "Art"),
        ("brickheadz", "BrickHeadz"),
        ("city", "City"),
        ("classic", "Classic"),
        ("creator", "Creator"),
        ("dc-comics", "DC"),
        ("disney", "Disney"),
        ("dots", "DOTS"),
        ("dreamzzz", "DREAMZzz"),
        ("duplo", "DUPLO"),
        ("friends", "Friends"),
        ("gabbys-dollhouse", "Gabby's Dollhouse"),
        ("harry-potter", "Harry Potter"),
        ("icons", "Icons"),
        ("ideas", "Ideas"),
        ("indiana-jones", "Indiana Jones"),
        ("jurassic-world", "Jurassic World"),
        ("marvel", "Marvel"),
        ("minecraft", "Minecraft"),
        ("ninjago", "NINJAGO"),
        ("sonic-the-hedgehog", "Sonic"),
        ("speed-champions", "Speed Champions"),
        ("star-wars", "Star Wars"),
        ("super-mario", "Super Mario"),
        ("technic", "Technic"),
        ("botanical-collection", "Botanical"),
        ("modular-buildings", "Modular Buildings"),
        ("ultimate-collector-series", "UCS"),
        ("polybags", "Polybags"),
        ("minifigures", "Minifigures"),
        ("advent-calendar", "Advent Calendar"),
    ]

    out: Dict[str, Dict[str, Any]] = {}
    seen_links: set = set()
    import threading
    lock = threading.Lock()

    def _scrape_category(cat_slug: str, theme_name: str) -> None:
        local: Dict[str, Dict] = {}
        local_links: set = set()
        with httpx.Client(timeout=DEFAULT_TIMEOUT, follow_redirects=True, headers=HEADERS) as client:
            for page in range(1, 50):
                url = (f"{BASE}/product-category/{cat_slug}/"
                       if page == 1
                       else f"{BASE}/product-category/{cat_slug}/page/{page}/")
                try:
                    r = client.get(url)
                except Exception as e:
                    print(f"[BRICKSHOP/{cat_slug}] error: {e}")
                    break

                if r.status_code == 404:
                    break
                if r.status_code != 200:
                    print(f"[BRICKSHOP/{cat_slug}] HTTP {r.status_code}")
                    break

                soup = BeautifulSoup(r.text, "lxml")
                cards = soup.select("ul.products li.product")
                if not cards:
                    break

                found_new = False
                for card in cards:
                    a = card.select_one("a.woocommerce-LoopProduct-link, h2 a, a[href*='/product/']")
                    link = (a.get("href") or "").strip() if a else ""
                    if not link or link in local_links:
                        continue
                    local_links.add(link)
                    found_new = True

                    title_el = card.select_one(".woocommerce-loop-product__title, h2, h3")
                    title = (title_el.get_text(" ", strip=True) if title_el else "").strip()
                    if not title:
                        title = link.rstrip("/").split("/")[-1].replace("-", " ").title()

                    slug = link.rstrip("/").split("/")[-1]
                    item_number = extract_item_number(slug)
                    if not item_number:
                        item_number = extract_item_number(title)
                    if not item_number:
                        continue

                    price = None
                    compare_at = None
                    price_el = card.select_one(".price")
                    if price_el:
                        price_text = price_el.get_text(" ", strip=True)
                        amounts = re.findall(r"\$\s*([0-9]+(?:\.[0-9]+)?)", price_text)
                        if len(amounts) >= 2:
                            compare_at = safe_float(amounts[0])
                            price = safe_float(amounts[1])
                        elif len(amounts) == 1:
                            price = safe_float(amounts[0])

                    oos = card.select_one(".out-of-stock")
                    availability = "Out of stock" if oos else ("In stock" if price else "N/A")

                    image_url = ""
                    image_list = []
                    for img in card.select("img"):
                        src = (img.get("data-src") or img.get("src") or "").strip()
                        if src and not src.startswith("data:") and any(
                            ext in src.lower() for ext in [".jpg", ".jpeg", ".png", ".webp"]
                        ):
                            image_list.append(src)
                            if not image_url:
                                image_url = src

                    discount_pct = compute_discount_pct(price, compare_at)

                    # BRICKSHOP marks new products with a "new" badge/tag
                    is_new = bool(card.select_one(".badge-new, .onsale, [class*='new']"))

                    local[item_number] = {
                        "item_number": item_number,
                        "title": title,
                        "theme": theme_name,
                        "category": "LEGO",
                        "image_url": image_url,
                        "image_list": image_list[:4],
                        "vendor": "BRICKSHOP",
                        "brand": "LEGO",
                        "compare_at": compare_at,
                        "is_new": is_new,
                        "stores": {
                            "BRICKSHOP": StoreOffer(
                                price=price,
                                availability=availability,
                                link=link,
                                discount_pct=discount_pct,
                            )
                        },
                    }

                if not found_new:
                    break

        # Merge into shared dict, dedup by seen_links
        with lock:
            for item_number, rec in local.items():
                lnk = rec["stores"]["BRICKSHOP"].link
                if lnk not in seen_links:
                    seen_links.add(lnk)
                    out[item_number] = rec

    # Scrape all categories in parallel (8 workers)
    with ThreadPoolExecutor(max_workers=8) as pool:
        futs = [pool.submit(_scrape_category, slug, name) for slug, name in CATEGORIES]
        for f in as_completed(futs):
            try:
                f.result()
            except Exception as e:
                print(f"[BRICKSHOP] category error: {e}")

    print(f"[BRICKSHOP] scraped {len(out)} products")
    return out


def fetch_playone() -> Dict[str, Dict[str, Any]]:
    out: Dict[str, Dict[str, Any]] = {}

    # PlayOne is behind Cloudflare — use cloudscraper to bypass JS challenge
    try:
        import cloudscraper
        client = cloudscraper.create_scraper(
            browser={"browser": "chrome", "platform": "windows", "mobile": False}
        )
        def _get(url):
            return client.get(url, timeout=DEFAULT_TIMEOUT)
    except ImportError:
        print("[PlayOne] WARNING: cloudscraper not installed. Run: pip install cloudscraper")
        print("[PlayOne] Falling back to httpx (will likely fail Cloudflare check)")
        import httpx as _httpx
        _client = _httpx.Client(timeout=DEFAULT_TIMEOUT, follow_redirects=True, headers=HEADERS)
        def _get(url):
            return _client.get(url)

    for base in PLAYONE_LISTING_URLS:
        for page in range(1, 51):
            url = base.rstrip("/") + (f"/page-{page}/" if page > 1 else "/")
            try:
                r = _get(url)
            except Exception as e:
                print(f"[PlayOne] fetch error: {e}")
                break
            if r.status_code != 200:
                print(f"[PlayOne] HTTP {r.status_code} for {url}")
                break

            soup = BeautifulSoup(r.text, "lxml")

            # Each product is div.thumbnail.grid-thumbnail
            cards = soup.select("div.thumbnail.grid-thumbnail")
            if not cards:
                title_tag = soup.select_one("title")
                page_title = title_tag.get_text(strip=True) if title_tag else "N/A"
                print(f"[PlayOne] page {page}: 0 cards. Title: '{page_title}' len={len(r.text)}")
                break

            found_any = False
            for card in cards:
                a = card.select_one("a.product-title[href]")
                if not a:
                    continue
                href = a.get("href") or ""
                if href.startswith("/"):
                    href = "https://playone.com.lb" + href

                title = (a.get("title") or a.get_text(" ", strip=True) or "").strip()
                title = title.replace("\u00a0", " ").replace("&amp;", "&").strip()
                if not title:
                    continue

                item_number = extract_item_number(title)
                if not item_number:
                    item_number = extract_item_number(href)
                if not item_number:
                    continue

                # Price: <span id="sec_discounted_price_NNN" class="price-num">69<span>.95</span></span>
                price = None
                price_span = card.select_one("[id^='sec_discounted_price_']")
                if price_span:
                    price = safe_float(price_span.get_text(strip=True).replace(",", ""))

                # Compare-at: <span id="sec_list_price_NNN" class="list-price">91</span>
                compare_at = None
                list_span = card.select_one("[id^='sec_list_price_']")
                if list_span:
                    compare_at = safe_float(list_span.get_text(strip=True).replace(",", ""))

                discount_pct = None
                if compare_at and price and compare_at > price:
                    discount_pct = round((compare_at - price) / compare_at * 100, 1)

                image_url = ""
                img = card.select_one("div.grid-list-image img[src]")
                if img:
                    src = img.get("src") or img.get("data-src") or ""
                    if src and not src.startswith("data:"):
                        image_url = src if src.startswith("http") else "https://playone.com.lb" + src

                is_new = bool(card.select_one("div.new-label"))
                theme, category = normalize_theme_category_from_playone(title)
                found_any = True

                out[item_number] = {
                    "item_number": item_number,
                    "title": title,
                    "theme": theme,
                    "category": category,
                    "image_url": image_url,
                    "image_list": [image_url] if image_url else [],
                    "vendor": "PlayOne",
                    "brand": "LEGO",
                    "compare_at": compare_at,
                    "is_new": is_new,
                    "stores": {
                        "PlayOne": StoreOffer(
                            price=price,
                            availability="available" if price is not None else "unavailable",
                            link=href,
                            discount_pct=discount_pct,
                        )
                    },
                }

            print(f"[PlayOne] page {page}: {len(cards)} products")
            if not found_any:
                break

    print(f"[PlayOne] total scraped: {len(out)}")
    return out


# ── Generic BigCommerce scraper ───────────────────────────────────────────────

def fetch_bigcommerce_store(store_name: str, base_url: str, category_slug: str = "",
                             lego_only: bool = True, vat_multiplier: float = 1.0) -> Dict[str, Dict[str, Any]]:
    """
    Generic BigCommerce HTML scraper — works for any BigCommerce store.
    Uses cloudscraper to bypass Cloudflare where needed.
    """
    out: Dict[str, Dict[str, Any]] = {}
    base = base_url.rstrip("/")
    cat_path = ("/" + category_slug.strip("/") + "/") if category_slug else "/"

    try:
        import cloudscraper as _cs
        scraper = _cs.create_scraper(browser={"browser": "chrome", "platform": "windows", "mobile": False})
    except ImportError:
        print(f"[{store_name}] cloudscraper not installed — falling back to httpx")
        scraper = None

    def _get(url: str):
        try:
            if scraper:
                return scraper.get(url, timeout=20)
            return httpx.get(url, timeout=20, follow_redirects=True, headers=HEADERS)
        except Exception as e:
            print(f"[{store_name}] fetch error: {e}")
            return None

    page = 1
    while True:
        url = f"{base}{cat_path}" if page == 1 else f"{base}{cat_path}?page={page}"
        r = _get(url)
        if not r or r.status_code != 200:
            if r: print(f"[{store_name}] HTTP {r.status_code} on page {page}")
            break

        soup = BeautifulSoup(r.text, "lxml")
        cards = (soup.select("article.product") or
                 soup.select("li[data-product-id]") or
                 soup.select(".productGrid .product") or
                 soup.select("[class*='productCard']"))

        if not cards:
            print(f"[{store_name}] no cards on page {page} — stopping")
            break

        found_new = False
        for card in cards:
            title_el = (card.select_one("h4.card-title a") or card.select_one(".product-title a") or
                        card.select_one("h3 a") or card.select_one("a[data-product-id]") or card.select_one("a"))
            title = (title_el.get_text(" ", strip=True) if title_el else "").strip()
            link = title_el.get("href", "") if title_el else ""
            if link and not link.startswith("http"):
                link = base + link
            if not title:
                continue
            if lego_only and "lego" not in title.lower():
                continue

            item_number = extract_item_number(title)
            if not item_number and link:
                item_number = extract_item_number(link.rstrip("/").split("/")[-1])
            if not item_number:
                pid_el = card.select_one("[data-product-id]")
                pid = pid_el.get("data-product-id") if pid_el else card.get("data-product-id")
                if pid: item_number = f"BC{pid}"
            if not item_number or item_number in out:
                continue
            found_new = True

            price_el = (card.select_one(".price--withoutTax") or card.select_one(".price--withTax") or
                        card.select_one(".price") or card.select_one("[data-product-price]"))
            price = None
            if price_el:
                raw = safe_float(re.sub(r"[^\d.]", "", price_el.get_text(strip=True)))
                price = (raw * vat_multiplier) if raw is not None else None

            compare_el = (card.select_one(".price--rrp") or card.select_one(".price--non-sale") or card.select_one(".price-was"))
            compare_at = None
            if compare_el:
                raw_c = safe_float(re.sub(r"[^\d.]", "", compare_el.get_text(strip=True)))
                compare_at = (raw_c * vat_multiplier) if raw_c is not None else None

            img_el = (card.select_one("img.card-image") or card.select_one("img[data-src]") or card.select_one("img"))
            image_url = ""
            if img_el:
                image_url = (img_el.get("data-src") or img_el.get("src") or "").strip()
                if image_url.startswith("//"): image_url = "https:" + image_url

            availability = "In stock"
            avail_el = card.select_one(".stock-level, [data-in-stock]")
            if avail_el and "out" in avail_el.get_text(strip=True).lower():
                availability = "Out of stock"

            out[item_number] = {
                "item_number": item_number, "title": title,
                "theme": "", "category": "LEGO",
                "image_url": image_url, "image_list": [image_url] if image_url else [],
                "vendor": "LEGO", "brand": "LEGO",
                "compare_at": compare_at, "is_new": False,
                "stores": {
                    store_name: StoreOffer(
                        price=price, availability=availability,
                        link=link or f"{base}{cat_path}",
                        discount_pct=compute_discount_pct(price, compare_at),
                        stock_qty=None,
                    )
                },
            }

        print(f"[{store_name}] page {page}: {len(cards)} cards, {len(out)} total")
        if not found_new:
            break
        next_el = soup.select_one("a.pagination-item--next, .pagination a[rel='next'], a[aria-label='Next page']")
        if not next_el and len(cards) < 12:
            break
        page += 1
        time.sleep(random.uniform(2.0, 4.0))

    print(f"[{store_name}] total scraped: {len(out)}")
    return out



    """
    Collision-safe merge key: BRAND|item_number
    Prevents e.g. LEGO 1000 and JAKI 1000 from being merged together.
    """
    b = (brand or "UNKNOWN").strip().upper()
    return f"{b}|{item_number}"


def make_merge_key(brand: str, item_number: str) -> str:
    """Collision-safe merge key: BRAND|item_number"""
    b = (brand or "UNKNOWN").strip().upper()
    return f"{b}|{item_number}"


def merge_catalogs(catalogs: List[Dict[str, Dict[str, Any]]]) -> Dict[str, Dict[str, Any]]:
    """
    Merge by brand+item_number to avoid cross-brand collisions.
    The displayed item_number stays unchanged; brand is stored for filtering.
    """
    merged: Dict[str, Dict[str, Any]] = {}

    for cat in catalogs:
        for item_number, rec in cat.items():
            if not item_number:
                continue

            brand = (rec.get("brand") or "UNKNOWN").strip().upper()
            key = make_merge_key(brand, item_number)

            if key not in merged:
                merged[key] = {
                    "merge_key": key,
                    "item_number": item_number,
                    "title": rec.get("title") or "",
                    "theme": rec.get("theme") or "",
                    "category": rec.get("category") or "",
                    "image_url": rec.get("image_url") or "",
                    "image_list": rec.get("image_list") or [],
                    "vendor": rec.get("vendor") or "",
                    "brand": brand,
                    "compare_at": rec.get("compare_at"),
                    "stores": {},
                }

            # fill missing fields
            if not merged[key]["title"] and rec.get("title"):
                merged[key]["title"] = rec["title"]
            if not merged[key]["theme"] and rec.get("theme"):
                merged[key]["theme"] = rec["theme"]
            if not merged[key]["category"] and rec.get("category"):
                merged[key]["category"] = rec["category"]
            if not merged[key]["image_url"] and rec.get("image_url"):
                merged[key]["image_url"] = rec["image_url"]
            if not merged[key]["image_list"] and rec.get("image_list"):
                merged[key]["image_list"] = rec["image_list"]
            if not merged[key]["vendor"] and rec.get("vendor"):
                merged[key]["vendor"] = rec["vendor"]
            if merged[key].get("compare_at") is None and rec.get("compare_at") is not None:
                merged[key]["compare_at"] = rec.get("compare_at")

            # merge offers
            stores_obj = rec.get("stores") or {}
            for sname, offer in stores_obj.items():
                merged[key]["stores"][sname] = offer

    # compute lowest
    for rec in merged.values():
        lowest_price = None
        lowest_store = None
        for sname, offer in rec["stores"].items():
            p = getattr(offer, "price", None)
            if p is None:
                continue
            if lowest_price is None or p < lowest_price:
                lowest_price = p
                lowest_store = sname
        rec["lowest_price"] = lowest_price
        rec["lowest_store"] = lowest_store

    return merged


def persist_snapshot(captured_at: str, store: str, catalog: Dict[str, Dict[str, Any]]) -> None:
    conn = db_connect()
    cur = conn.cursor()

    for item, rec in catalog.items():
        offer = (rec.get("stores") or {}).get(store)
        if not offer:
            continue

        cur.execute("""
        INSERT INTO snapshots(
            captured_at, store, item_number, title, theme, category, brand,
            price, availability, link, image_url, images_json, compare_at, stock_qty
        ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            captured_at,
            store,
            item,
            rec.get("title") or "",
            rec.get("theme") or "",
            rec.get("category") or "",
            rec.get("brand") or "",
            offer.price,
            offer.availability,
            offer.link,
            rec.get("image_url") or "",
            json.dumps(rec.get("image_list") or []),
            rec.get("compare_at"),
            offer.stock_qty,
        ))

    conn.commit()
    conn.close()


def compute_alerts(captured_at: str, store: str, catalog: Dict[str, Dict[str, Any]]) -> None:
    conn = db_connect()
    cur = conn.cursor()

    for item, rec in catalog.items():
        offer = (rec.get("stores") or {}).get(store)
        if not offer or offer.price is None:
            continue

        title = rec.get("title") or ""

        # Count how many snapshots exist for this item+store
        # (persist_snapshot already ran, so current = latest)
        cur.execute("""
        SELECT COUNT(*) as cnt FROM snapshots
        WHERE store=? AND item_number=?
        """, (store, item))
        cnt = cur.fetchone()["cnt"]

        # Check if store marks this as new (in new arrivals collection)
        # Alert once per item+store regardless of snapshot count
        is_new_from_store = rec.get("is_new", False)
        if is_new_from_store:
            cur.execute("""
            SELECT 1 FROM alerts
            WHERE item_number=? AND store=? AND alert_type='new_arrival'
            LIMIT 1
            """, (item, store))
            if not cur.fetchone():
                cur.execute("""
                INSERT INTO alerts(created_at, item_number, store, title, old_price, new_price, alert_type, unread)
                VALUES(?,?,?,?,?,?,?,1)
                """, (captured_at, item, store, title, None, float(offer.price), "new_arrival"))

        if cnt <= 1:
            continue

        # Get the previous snapshot (second-to-last) to compare price
        cur.execute("""
        SELECT price FROM snapshots
        WHERE store=? AND item_number=?
        ORDER BY id DESC
        LIMIT 1 OFFSET 1
        """, (store, item))
        prev = cur.fetchone()

        if not prev or prev["price"] is None:
            continue

        old_price = float(prev["price"])
        new_price = float(offer.price)

        diff = new_price - old_price
        if abs(diff) >= 0.01:
            # Skip if new price equals compare_at — sale ended, not a real price increase
            compare_at = rec.get("compare_at")
            if diff > 0 and compare_at and abs(new_price - float(compare_at)) < 0.02:
                continue
            # Skip implausible >50% increases (data artifact)
            if diff > 0 and diff / old_price > 0.5:
                continue
            alert_type = "price_drop" if diff < 0 else "price_increase"
            cur.execute("""
            INSERT INTO alerts(created_at, item_number, store, title, old_price, new_price, alert_type, unread)
            VALUES(?,?,?,?,?,?,?,1)
            """, (captured_at, item, store, title, old_price, new_price, alert_type))

    conn.commit()
    conn.close()


def get_all_store_names() -> List[str]:
    names = list(SHOPIFY_STORES.keys())
    names.append("BRICKSHOP")
    names.append("PlayOne")
    for bc_name in BIGCOMMERCE_STORES:
        names.append(bc_name)
    # Also include enabled DB-managed stores
    try:
        conn = db_connect()
        cur = conn.cursor()
        cur.execute("SELECT name FROM stores WHERE enabled=1")
        for row in cur.fetchall():
            if row["name"] not in names:
                names.append(row["name"])
        conn.close()
    except Exception:
        pass
    return order_stores(names)


def refresh_all() -> Dict[str, Any]:
    captured_at = utc_now_iso()

    # Build list of (store_name, fetch_fn) tasks
    tasks = []
    for sname, cfg in SHOPIFY_STORES.items():
        tasks.append(("shopify", sname, cfg))

    # Also include DB-managed stores that are enabled
    _sconn = db_connect()
    _scur = _sconn.cursor()
    _scur.execute("SELECT name, base_url, platform, vat_multiplier, new_arrivals_collection, collection_slug, lego_only FROM stores WHERE enabled=1")
    db_stores = _scur.fetchall()
    _sconn.close()
    for ds in db_stores:
        sname = ds["name"]
        if sname in SHOPIFY_STORES:
            continue  # already included above
        if ds["platform"] == "shopify":
            cfg = {
                "base_url": ds["base_url"],
                "vat_multiplier": ds["vat_multiplier"],
                "new_arrivals_collection": ds["new_arrivals_collection"],
                "collection_slug": ds["collection_slug"],
                "lego_only": bool(ds["lego_only"]),
            }
            tasks.append(("shopify_db", sname, cfg))

    tasks.append(("brickshop", None, None))
    tasks.append(("playone", None, None))
    for bc_name, bc_cfg in BIGCOMMERCE_STORES.items():
        tasks.append(("bigcommerce", bc_name, bc_cfg))

    results: Dict[str, Dict[str, Dict[str, Any]]] = {}

    # Fetch all stores in parallel — I/O bound so threads work well
    def _run(task):
        kind, sname, cfg = task
        if kind in ("shopify", "shopify_db"):
            vat = float(cfg.get("vat_multiplier", 1.0))
            collection_slug = cfg.get("collection_slug")
            new_items = fetch_new_arrival_items(sname, cfg.get("base_url", ""))
            return sname, fetch_shopify_store(sname, cfg["base_url"], vat, new_items=new_items, collection_slug=collection_slug, lego_only=cfg.get("lego_only", False))
        elif kind == "brickshop":
            return "BRICKSHOP", fetch_brickshop()
        else:
            return "PlayOne", fetch_playone()

    with ThreadPoolExecutor(max_workers=len(tasks)) as pool:
        futures = {}
        for i, t in enumerate(tasks):
            if i > 0:
                time.sleep(3)
            futures[pool.submit(_run, t)] = t
        for future in as_completed(futures):
            try:
                store_name, cat = future.result()
                results[store_name] = cat
            except Exception as e:
                print(f"[refresh_all] fetch error: {e}")

    # Persist and compute alerts sequentially (SQLite is not thread-safe for writes)
    store_order = order_stores(list(results.keys()))
    catalogs = []
    for sname in store_order:
        cat = results.get(sname, {})
        catalogs.append(cat)
        persist_snapshot(captured_at, sname, cat)
        compute_alerts(captured_at, sname, cat)

    meta_set("last_updated", captured_at)

    merged = merge_catalogs(catalogs)
    return {
        "captured_at": captured_at,
        "total_items": len(merged),
        "stores": list(get_all_store_names()),
    }


# ----------------------------
# FastAPI App
# ----------------------------

db_init()

app = FastAPI()
templates = Jinja2Templates(directory=TEMPLATES_DIR)

# Register tojson filter (not included by default in FastAPI Jinja2)
import json as _json
templates.env.filters["tojson"] = lambda v: _json.dumps(v)

static_dir = os.path.join(APP_DIR, "static")
if os.path.isdir(static_dir):
    app.mount("/static", StaticFiles(directory=static_dir), name="static")


@app.get("/", response_class=HTMLResponse)
def dashboard(request: Request):
    store_names = get_all_store_names()

    conn = db_connect()
    cur = conn.cursor()

    latest_by_store_item: Dict[Tuple[str, str], sqlite3.Row] = {}

    cur.execute("""
    SELECT s.*
    FROM snapshots s
    INNER JOIN (
        SELECT store, item_number, MAX(id) AS max_id
        FROM snapshots
        GROUP BY store, item_number
    ) t
    ON s.id = t.max_id
    """)
    rows = cur.fetchall()
    for r in rows:
        item_key = r["item_number"] or f"SID_{r['store']}_{r['id']}"
        latest_by_store_item[(r["store"], item_key)] = r

    conn.close()

    catalogs: List[Dict[str, Dict[str, Any]]] = []
    for s in store_names:
        cat: Dict[str, Dict[str, Any]] = {}
        for (st, item), r in latest_by_store_item.items():
            if st != s:
                continue
            if not item:
                continue
            offer = StoreOffer(
                price=r["price"],
                availability=r["availability"] or "N/A",
                link=r["link"] or "",
                discount_pct=compute_discount_pct(r["price"], r["compare_at"]),
            )
            raw_images_json = r["images_json"] if r["images_json"] else "[]"
            try:
                img_list = json.loads(raw_images_json)
            except Exception:
                img_list = []
            if not img_list and r["image_url"]:
                img_list = [r["image_url"]]

            cat[item] = {
                "item_number": item,
                "title": r["title"] or "",
                "theme": r["theme"] or "",
                "category": r["category"] or "",
                "image_url": r["image_url"] or "",
                "image_list": img_list,
                "vendor": "",
                "brand": (r["brand"] if r["brand"] else "").strip().upper() or "LEGO",
                "compare_at": r["compare_at"],
                "stores": {s: offer},
            }
        catalogs.append(cat)

    merged = merge_catalogs(catalogs)

    qp = request.query_params
    selected_category = qp.get("category", "All")
    selected_theme = qp.get("theme", "All")
    selected_brand = qp.get("brand", "All")
    search_item = qp.get("search_item", "").strip()
    only_deals       = qp.get("only_deals", "0") == "1"
    only_alerts      = qp.get("only_alerts", "0") == "1"
    only_comparable  = qp.get("only_comparable", "0") == "1"
    only_instock     = qp.get("only_instock", "0") == "1"
    alert_type_filter = qp.get("alert_type", "all")

    sort = qp.get("sort", "item")
    order = qp.get("order", "asc")

    per_page = int(qp.get("per_page", "1000") or "1000")
    page = int(qp.get("page", "1") or "1")
    if page < 1:
        page = 1

    # stores to show — multi-checkbox: ?stores=Brickmania&stores=BRICKSHOP etc.
    raw_selected_stores = request.query_params.getlist("stores")
    if raw_selected_stores:
        stores = order_stores([s for s in raw_selected_stores if s in store_names])
        if not stores:
            stores = store_names
        selected_stores = stores  # only the ones the user picked
    else:
        stores = store_names
        selected_stores = []   # empty = "all" default — sort links won't add ?stores= params
    compare = "all"  # kept for backwards compat in template sort links

    all_categories = sorted({(rec.get("category") or "").strip() for rec in merged.values() if (rec.get("category") or "").strip()})
    all_themes = sorted({(rec.get("theme") or "").strip() for rec in merged.values() if (rec.get("theme") or "").strip()})
    all_brands = sorted({(rec.get("brand") or "").strip().upper() for rec in merged.values() if (rec.get("brand") or "").strip()})

    all_categories = ["All"] + all_categories
    all_themes = ["All"] + all_themes
    all_brands = ["All"] + all_brands

    # Build alert lookup BEFORE filter loop — store-aware
    # Structure: item_alerts[item_number][store] = {type, old_price, new_price}
    ALERT_PRIORITY = {"price_drop": 0, "new_arrival": 1, "new_in_store": 2, "price_increase": 3}
    _aconn = db_connect()
    _acur = _aconn.cursor()
    _acur.execute("""
        SELECT item_number, store, alert_type, old_price, new_price
        FROM alerts WHERE unread=1 ORDER BY id DESC
    """)
    # item_store_alerts[item][store] = best alert for that store
    item_store_alerts: Dict[str, Dict[str, Dict]] = {}
    for ar in _acur.fetchall():
        iid = ar["item_number"]
        sname = ar["store"]
        atype = ar["alert_type"] or "price_change"
        if iid not in item_store_alerts:
            item_store_alerts[iid] = {}
        cur_p = ALERT_PRIORITY.get(item_store_alerts[iid].get(sname, {}).get("type", "new_in_store"), 99)
        new_p = ALERT_PRIORITY.get(atype, 99)
        if sname not in item_store_alerts[iid] or new_p < cur_p:
            item_store_alerts[iid][sname] = {
                "type": atype,
                "store": sname,
                "old_price": ar["old_price"],
                "new_price": ar["new_price"],
            }
    _aconn.close()

    # Helper: get best alert for item across selected stores only
    def _get_item_alert(item_number: str, selected: List[str]) -> Optional[Dict]:
        store_map = item_store_alerts.get(item_number, {})
        best = None
        best_p = 99
        for s in selected:
            a = store_map.get(s)
            if not a:
                continue
            p = ALERT_PRIORITY.get(a["type"], 99)
            if p < best_p:
                best_p, best = p, a
        return best

    def _get_item_alert_types(item_number: str, selected: List[str]) -> set:
        store_map = item_store_alerts.get(item_number, {})
        types: set = set()
        for s in selected:
            a = store_map.get(s)
            if a:
                types.add(a["type"])
        return types

    filtered: List[Dict[str, Any]] = []
    for rec in merged.values():
        cat = (rec.get("category") or "").strip()
        th = (rec.get("theme") or "").strip()

        if selected_category != "All" and cat != selected_category:
            continue
        if selected_theme != "All" and th != selected_theme:
            continue

        br = (rec.get("brand") or "").strip().upper()
        if selected_brand != "All" and br != selected_brand.upper():
            continue

        if search_item:
            if search_item not in (rec.get("item_number") or ""):
                continue

        rec2 = dict(rec)
        rec2["stores"] = {s: rec["stores"].get(s) for s in stores if s in rec["stores"]}

        # Bug 2 fix: skip items with no offer in any selected store
        if not any(rec2["stores"].values()):
            continue

        # image_proxy: strip Shopify size params for a clean URL the browser can load
        raw_img = rec2.get("image_url") or ""
        rec2["image_proxy"] = re.sub(r"_[0-9]+x[0-9]*\.", ".", raw_img) if raw_img else ""
        # image_list_clean: all images with size params stripped, max 4
        rec2["image_list_clean"] = [
            re.sub(r"_[0-9]+x[0-9]*.", ".", u) for u in (rec2.get("image_list") or [])[:4]
        ]
        # Pre-serialise to JSON string for safe embedding in data attribute
        rec2["image_list_json"] = json.dumps(rec2["image_list_clean"])

        lp = None
        ls = None
        for s in stores:
            off = rec2["stores"].get(s)
            if not off or off.price is None:
                continue
            if lp is None or off.price < lp:
                lp = off.price
                ls = s
        rec2["lowest_price"] = lp
        rec2["lowest_store"] = ls

        if only_deals:
            has_disc = any((off.discount_pct or 0) > 0 for off in rec2["stores"].values() if off)
            if not has_disc:
                continue

        # Attach store-aware alert data
        # merge key is brand|item_number — extract plain item_number for alert lookup
        _plain_item = rec2["item_number"].split("|")[-1] if "|" in (rec2["item_number"] or "") else rec2["item_number"]
        rec2["alert"] = _get_item_alert(_plain_item, stores)
        rec2["alert_types"] = _get_item_alert_types(_plain_item, stores)

        if only_alerts:
            atypes = rec2["alert_types"]
            if not atypes:
                continue
            if alert_type_filter != "all" and alert_type_filter not in atypes:
                continue

        # Comparable: item must have a price in ALL selected stores
        if only_comparable:
            has_all = all(
                rec2["stores"].get(s) and rec2["stores"][s].price is not None
                for s in stores
            )
            if not has_all:
                continue

        # Availability: at least one selected store has "In stock"
        if only_instock:
            in_stock = any(
                rec2["stores"].get(s) and "in stock" in (rec2["stores"][s].availability or "").lower()
                for s in stores
            )
            if not in_stock:
                continue

        filtered.append(rec2)

    reverse = (order == "desc")

    def sort_key(r: Dict[str, Any]):
        if sort == "brand":
            return (r.get("brand") or "").upper()
        if sort == "item":
            return int(r["item_number"]) if str(r["item_number"]).isdigit() else 10**9
        if sort == "title":
            return (r.get("title") or "").lower()
        if sort == "theme":
            return (r.get("theme") or "").lower()
        if sort == "category":
            return (r.get("category") or "").lower()
        if sort == "price":
            return r.get("lowest_price") if r.get("lowest_price") is not None else 10**12
        if sort.startswith("store:"):
            sname = sort.split(":", 1)[1]
            off = (r.get("stores") or {}).get(sname)
            return off.price if (off and off.price is not None) else 10**12
        return (r.get("item_number") or "")

    # Fetch all unread alerts per item — store by type priority
    # Priority: price_drop > price_increase > new_arrival > new_in_store
    ALERT_PRIORITY = {"price_drop": 0, "new_arrival": 1, "new_in_store": 2, "price_increase": 3}
    conn2 = db_connect()
    cur2 = conn2.cursor()
    cur2.execute("""
        SELECT item_number, alert_type, old_price, new_price, created_at
        FROM alerts
        WHERE unread=1
        ORDER BY id DESC
    """)
    item_alerts: Dict[str, Dict] = {}       # best alert per item (for badge)
    item_alert_types: Dict[str, set] = {}   # all alert types per item (for filter)
    for ar in cur2.fetchall():
        iid = ar["item_number"]
        atype = ar["alert_type"] or "price_change"
        # track all types
        if iid not in item_alert_types:
            item_alert_types[iid] = set()
        item_alert_types[iid].add(atype)
        # keep highest-priority alert for badge
        cur_priority = ALERT_PRIORITY.get(item_alerts.get(iid, {}).get("type", "new_in_store"), 99)
        new_priority = ALERT_PRIORITY.get(atype, 99)
        if iid not in item_alerts or new_priority < cur_priority:
            item_alerts[iid] = {
                "type": atype,
                "old_price": ar["old_price"],
                "new_price": ar["new_price"],
                "created_at": ar["created_at"],
            }
    conn2.close()

    filtered.sort(key=sort_key, reverse=reverse)

    total = len(filtered)
    total_pages = max(1, (total + per_page - 1) // per_page)
    if page > total_pages:
        page = total_pages
    start = (page - 1) * per_page
    end = start + per_page
    page_rows = filtered[start:end]

    last_updated = meta_get("last_updated") or "never"
    alerts_unread = alerts_unread_count()

    return templates.TemplateResponse(
        TEMPLATE_FILE,
        {
            "request": request,
            "rows": page_rows,
            "total": total,
            "page": page,
            "per_page": per_page,
            "total_pages": total_pages,
            "sort": sort,
            "order": order,
            "stores": stores,
            "store_names": store_names,
            "all_categories": all_categories,
            "all_themes": all_themes,
            "all_brands": all_brands,
            "selected_category": selected_category,
            "selected_theme": selected_theme,
            "selected_brand": selected_brand,
            "search_item": search_item,
            "compare": compare,
            "selected_stores": selected_stores,
            "raw_selected_stores": raw_selected_stores,
            "only_deals": only_deals,
            "only_alerts": only_alerts,
            "only_comparable": only_comparable,
            "only_instock": only_instock,
            "alert_type_filter": alert_type_filter,
            "last_updated": last_updated,
            "alerts_unread": alerts_unread,
            "price_range": "all",
            "min_price": "",
            "max_price": "",
        },
    )


@app.post("/api/refresh")
def api_refresh():
    res = refresh_all()
    return JSONResponse(res)


@app.get("/api/refresh/stream")
def api_refresh_stream():
    """SSE endpoint — streams per-store progress then triggers full refresh logic."""
    import queue as _queue

    def event_stream():
        captured_at = utc_now_iso()
        tasks = []
        for sname, cfg in SHOPIFY_STORES.items():
            tasks.append(("shopify", sname, cfg))

        _sconn = db_connect()
        _scur = _sconn.cursor()
        _scur.execute("SELECT name, base_url, platform, vat_multiplier, new_arrivals_collection, collection_slug, lego_only FROM stores WHERE enabled=1")
        db_stores = _scur.fetchall()
        _sconn.close()
        for ds in db_stores:
            sname = ds["name"]
            if sname in SHOPIFY_STORES:
                continue
            if ds["platform"] == "shopify":
                cfg = {
                    "base_url": ds["base_url"],
                    "vat_multiplier": ds["vat_multiplier"],
                    "new_arrivals_collection": ds["new_arrivals_collection"],
                    "collection_slug": ds["collection_slug"],
                    "lego_only": bool(ds["lego_only"]),
                }
                tasks.append(("shopify_db", sname, cfg))

        tasks.append(("brickshop", None, None))
        tasks.append(("playone", None, None))
        for bc_name, bc_cfg in BIGCOMMERCE_STORES.items():
            tasks.append(("bigcommerce", bc_name, bc_cfg))

        total = len(tasks)
        print(f"[REFRESH] total tasks: {total} — {[t[1] or t[0] for t in tasks]}")
        q = _queue.Queue()

        def _run_and_report(task, stagger_idx: int = 0):
            if stagger_idx > 0:
                time.sleep(stagger_idx * 5)  # 5s per slot
            kind, sname, cfg = task
            label = sname or kind
            try:
                if kind in ("shopify", "shopify_db"):
                    vat = float(cfg.get("vat_multiplier", 1.0))
                    collection_slug = cfg.get("collection_slug")
                    new_items = fetch_new_arrival_items(label, cfg.get("base_url", ""))
                    result = fetch_shopify_store(label, cfg["base_url"], vat, new_items=new_items, collection_slug=collection_slug, lego_only=cfg.get("lego_only", False))
                    q.put(("ok", label, result))
                elif kind == "brickshop":
                    result = fetch_brickshop()
                    q.put(("ok", "BRICKSHOP", result))
                elif kind == "bigcommerce":
                    result = fetch_bigcommerce_store(label, cfg["base_url"], cfg.get("category_slug",""), cfg.get("lego_only", True), float(cfg.get("vat_multiplier", 1.0)))
                    q.put(("ok", label, result))
                else:
                    result = fetch_playone()
                    q.put(("ok", "PlayOne", result))
            except Exception as e:
                q.put(("err", label, str(e)))

        import json as _json2
        yield f"data: {_json2.dumps({'type':'start','total':total})}\n\n"

        # Keepalive: ping queue every 20s so SSE connection stays open during long retries
        import threading as _threading
        _stop_ping = _threading.Event()
        def _ping():
            while not _stop_ping.wait(20):
                q.put(("ping", None, None))
        _threading.Thread(target=_ping, daemon=True).start()

        # Background persister — drains queue and saves results regardless of SSE state
        import threading as _t2
        _persist_done = _t2.Event()
        _sse_q = _queue.Queue()  # forwards results to SSE generator

        def _background_persist():
            done_bg = 0
            while done_bg < total:
                try:
                    status, name, data = q.get(timeout=600)  # 10 min max wait
                except Exception:
                    print(f"[PERSIST] queue timeout after {done_bg}/{total} stores")
                    break
                if status == "ping":
                    _sse_q.put(("ping", None, None))
                    continue
                done_bg += 1
                if status == "ok":
                    count = len(data) if isinstance(data, dict) else 0
                    print(f"[PERSIST] {name}: {count} items ({done_bg}/{total})")
                    persist_snapshot(captured_at, name, data)
                    compute_alerts(captured_at, name, data)
                    _sse_q.put(("ok", name, count, done_bg))
                else:
                    print(f"[PERSIST ERROR] {name}: {data}")
                    _sse_q.put(("err", name, data, done_bg))
            meta_set("last_updated", captured_at)
            _sse_q.put(("done", None, None, total))
            _persist_done.set()

        _bg = _t2.Thread(target=_background_persist, daemon=True)
        _bg.start()

        # Submit all scrape tasks
        # PlayOne (cloudscraper, slow sequential) always starts immediately at idx=0
        # Shopify/WooCommerce stores get staggered — 5s apart to avoid simultaneous 429s
        pool = ThreadPoolExecutor(max_workers=len(tasks))
        shopify_idx = 0
        for t in tasks:
            kind = t[0]
            if kind in ("brickshop", "playone"):
                pool.submit(_run_and_report, t, 0)
            else:
                pool.submit(_run_and_report, t, shopify_idx)
                shopify_idx += 1
        pool.shutdown(wait=False)

        # Stream results to SSE client — if client disconnects, _background_persist keeps running
        while True:
            msg = _sse_q.get()
            kind = msg[0]
            if kind == "ping":
                yield ": keepalive\n\n"
            elif kind == "ok":
                _, name, count, done_n = msg
                yield f"data: {_json2.dumps({'type':'store','name':name,'count':count,'done':done_n,'total':total})}\n\n"
            elif kind == "err":
                _, name, err, done_n = msg
                yield f"data: {_json2.dumps({'type':'error','name':name,'error':err,'done':done_n,'total':total})}\n\n"
            elif kind == "done":
                break

        _stop_ping.set()
        yield f"data: {_json2.dumps({'type':'done'})}\n\n"

    return StreamingResponse(event_stream(), media_type="text/event-stream")


@app.post("/api/alerts/mark_read")
def api_mark_alerts_read():
    alerts_mark_read()
    return JSONResponse({"ok": True, "unread": alerts_unread_count()})


# ── Store Manager API ──────────────────────────────────────────────────────

@app.get("/stores", response_class=HTMLResponse)
def stores_page():
    stores_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "stores.html")
    with open(stores_file, "r", encoding="utf-8") as f:
        return f.read()

@app.get("/api/stores")
def api_get_stores():
    conn = db_connect()
    cur = conn.cursor()

    # Get product counts per store from snapshots
    cur.execute("""
        SELECT store, COUNT(DISTINCT item_number) as cnt, MAX(captured_at) as last_seen
        FROM snapshots GROUP BY store
    """)
    snap_counts = {r["store"]: {"count": r["cnt"], "last_seen": r["last_seen"]} for r in cur.fetchall()}

    # DB-managed stores
    cur.execute("SELECT * FROM stores ORDER BY name")
    db_rows = {r["name"]: dict(r) for r in cur.fetchall()}

    # Build full list: hardcoded + DB
    all_stores = []

    # Hardcoded stores
    hardcoded = []
    for sname, cfg in SHOPIFY_STORES.items():
        hardcoded.append({
            "id": None,
            "name": sname,
            "base_url": cfg["base_url"],
            "platform": "shopify",
            "vat_multiplier": cfg.get("vat_multiplier", 1.0),
            "collection_slug": cfg.get("collection_slug"),
            "new_arrivals_collection": NEW_ARRIVAL_COLLECTIONS.get(sname),
            "enabled": 1,
            "hardcoded": True,
            "lego_only": cfg.get("lego_only", False),
        })
    hardcoded.append({"id": None, "name": "BRICKSHOP", "base_url": "https://brickshop.me", "platform": "woocommerce", "vat_multiplier": 1.0, "collection_slug": None, "new_arrivals_collection": None, "enabled": 1, "hardcoded": True, "lego_only": False})
    hardcoded.append({"id": None, "name": "PlayOne", "base_url": "https://playone.com.lb", "platform": "html", "vat_multiplier": 1.0, "collection_slug": None, "new_arrivals_collection": None, "enabled": 1, "hardcoded": True, "lego_only": True})
    hardcoded.append({"id": None, "name": "Ayoub Computers", "base_url": "https://ayoubcomputers.com", "platform": "bigcommerce", "vat_multiplier": 1.0, "collection_slug": "lego", "new_arrivals_collection": None, "enabled": 1, "hardcoded": True, "lego_only": True})

    for s in hardcoded:
        snap = snap_counts.get(s["name"], {})
        s["product_count"] = snap.get("count", 0)
        s["last_scraped"] = snap.get("last_seen", "")
        all_stores.append(s)

    # DB-managed stores (not already in hardcoded)
    hardcoded_names = {s["name"] for s in hardcoded}
    for name, row in db_rows.items():
        if name in hardcoded_names:
            continue
        snap = snap_counts.get(name, {})
        row["product_count"] = snap.get("count", 0)
        row["last_scraped"] = snap.get("last_seen", row.get("last_scraped", ""))
        row["hardcoded"] = False
        all_stores.append(row)

    conn.close()
    return JSONResponse(all_stores)

@app.post("/api/stores/test")
async def api_test_store(request: Request):
    body = await request.json()
    url = (body.get("url") or "").strip().rstrip("/")
    collection_slug = (body.get("collection_slug") or "").strip() or None
    if not url.startswith("http"):
        url = "https://" + url

    result = {"url": url, "platform": None, "product_count": 0, "samples": [], "error": None, "collection_slug": collection_slug}

    # Build product list URL — use collection if specified
    def _products_url(base: str, page: int, slug: Optional[str] = None) -> str:
        if slug:
            return f"{base}/collections/{slug}/products.json?limit=250&page={page}"
        return f"{base}/products.json?limit=250&page={page}"

    try:
        import httpx as _httpx
        # Try Shopify
        test_url = _products_url(url, 1, collection_slug)
        r = _httpx.get(test_url, timeout=15, follow_redirects=True, headers=HEADERS)
        if r.status_code == 200:
            products = r.json().get("products") or []
            if products:
                result["platform"] = "shopify"
                # Count total products
                page = 1
                total = 0
                with _httpx.Client(timeout=15, follow_redirects=True, headers=HEADERS) as client:
                    while True:
                        pr = client.get(_products_url(url, page, collection_slug))
                        if pr.status_code != 200: break
                        prods = pr.json().get("products") or []
                        if not prods: break
                        total += len(prods)
                        if len(prods) < 250: break
                        page += 1
                        if page > 5: break  # cap at 5 pages for test
                result["product_count"] = total
                # Sample products with item numbers
                for p in products[:5]:
                    title = (p.get("title") or "").strip()
                    item_number = extract_item_number(title)
                    if not item_number:
                        for v in (p.get("variants") or []):
                            item_number = extract_item_number((v.get("sku") or ""))
                            if item_number: break
                    price = None
                    variants = p.get("variants") or []
                    if variants:
                        price = safe_float(variants[0].get("price"))
                    img = ""
                    images = p.get("images") or []
                    if images:
                        img = (images[0].get("src") or "")
                    result["samples"].append({
                        "title": title,
                        "item_number": item_number or "—",
                        "price": price,
                        "image": img,
                    })
            else:
                result["error"] = "Shopify endpoint returned no products"
        else:
            result["error"] = f"Not Shopify (HTTP {r.status_code}). WooCommerce support coming soon."
    except Exception as e:
        result["error"] = str(e)

    return JSONResponse(result)

@app.post("/api/stores/add")
async def api_add_store(request: Request):
    body = await request.json()
    name = (body.get("name") or "").strip()
    url  = (body.get("url") or "").strip().rstrip("/")
    platform = body.get("platform") or "shopify"
    vat  = float(body.get("vat_multiplier") or 1.0)
    new_arrivals = (body.get("new_arrivals_collection") or "").strip() or None
    collection_slug = (body.get("collection_slug") or "").strip() or None
    lego_only = int(bool(body.get("lego_only", False)))

    if not name or not url:
        return JSONResponse({"ok": False, "error": "Name and URL required"}, status_code=400)

    conn = db_connect()
    cur = conn.cursor()
    try:
        cur.execute("""
            INSERT INTO stores(name, base_url, platform, vat_multiplier, new_arrivals_collection, collection_slug, lego_only, enabled)
            VALUES(?,?,?,?,?,?,?,1)
        """, (name, url, platform, vat, new_arrivals, collection_slug, lego_only))
        conn.commit()
    except Exception as e:
        conn.close()
        return JSONResponse({"ok": False, "error": str(e)}, status_code=400)
    conn.close()
    # Fetch logo in background
    import threading as _lt
    _lt.Thread(target=fetch_store_logo, args=(url, name), daemon=True).start()
    return JSONResponse({"ok": True})

@app.post("/api/stores/toggle")
async def api_toggle_store(request: Request):
    body = await request.json()
    store_id = body.get("id")
    conn = db_connect()
    cur = conn.cursor()
    cur.execute("UPDATE stores SET enabled = 1 - enabled WHERE id=?", (store_id,))
    conn.commit()
    cur.execute("SELECT enabled FROM stores WHERE id=?", (store_id,))
    row = cur.fetchone()
    conn.close()
    return JSONResponse({"ok": True, "enabled": row["enabled"] if row else 0})

@app.post("/api/stores/delete")
async def api_delete_store(request: Request):
    body = await request.json()
    store_id = body.get("id")
    conn = db_connect()
    cur = conn.cursor()
    cur.execute("DELETE FROM stores WHERE id=?", (store_id,))
    conn.commit()
    conn.close()
    return JSONResponse({"ok": True})

@app.get("/favicon.ico")
def favicon():
    return JSONResponse({"ok": True})


# ── Analytics ──────────────────────────────────────────────────────────────

@app.get("/analytics", response_class=HTMLResponse)
def analytics_page():
    f = os.path.join(os.path.dirname(os.path.abspath(__file__)), "analytics.html")
    with open(f, "r", encoding="utf-8") as fh:
        return fh.read()


def _latest_snapshot_filter() -> str:
    """SQL snippet: restrict to the most recent captured_at per store."""
    return """
        AND captured_at IN (
            SELECT MAX(captured_at) FROM snapshots GROUP BY store
        )
    """


@app.get("/api/analytics/kpis")
def api_analytics_kpis():
    conn = db_connect()
    cur = conn.cursor()
    lsf = _latest_snapshot_filter()

    cur.execute(f"SELECT COUNT(DISTINCT item_number) FROM snapshots WHERE 1=1 {lsf}")
    total_products = cur.fetchone()[0]

    cur.execute(f"SELECT COUNT(DISTINCT store) FROM snapshots WHERE 1=1 {lsf}")
    total_stores = cur.fetchone()[0]

    cur.execute(f"SELECT COUNT(*) FROM snapshots WHERE compare_at IS NOT NULL AND compare_at > price {lsf}")
    discounted = cur.fetchone()[0]

    cur.execute("""
        SELECT COUNT(DISTINCT item_number) FROM alerts
        WHERE alert_type='new_arrival'
        AND created_at >= datetime('now','-7 days')
    """)
    new_arrivals_7d = cur.fetchone()[0]

    cur.execute(f"SELECT AVG(price) FROM snapshots WHERE price IS NOT NULL {lsf}")
    avg_price = cur.fetchone()[0]

    cur.execute(f"""
        SELECT COUNT(*) FROM snapshots
        WHERE availability NOT IN ('available','in_stock','')
        AND availability IS NOT NULL {lsf}
    """)
    out_of_stock = cur.fetchone()[0]

    conn.close()
    return JSONResponse({
        "total_products": total_products,
        "total_stores": total_stores,
        "discounted": discounted,
        "new_arrivals_7d": new_arrivals_7d,
        "avg_price": round(avg_price, 2) if avg_price else 0,
        "out_of_stock": out_of_stock,
    })


@app.get("/api/analytics/items_per_brand_store")
def api_items_per_brand_store():
    conn = db_connect()
    cur = conn.cursor()
    lsf = _latest_snapshot_filter()
    cur.execute(f"""
        SELECT store, UPPER(TRIM(brand)) as brand, COUNT(DISTINCT item_number) as cnt
        FROM snapshots
        WHERE brand IS NOT NULL AND brand != '' {lsf}
        GROUP BY store, brand
        ORDER BY store, cnt DESC
    """)
    rows = [{"store": r[0], "brand": r[1], "count": r[2]} for r in cur.fetchall()]
    conn.close()
    return JSONResponse(rows)


@app.get("/api/analytics/most_expensive_per_brand_store")
def api_most_expensive_per_brand_store():
    conn = db_connect()
    cur = conn.cursor()
    lsf = _latest_snapshot_filter()
    cur.execute(f"""
        SELECT store, UPPER(TRIM(brand)) as brand, title, item_number, MAX(price) as price, link, image_url
        FROM snapshots
        WHERE price IS NOT NULL AND brand IS NOT NULL AND brand != '' {lsf}
        GROUP BY store, brand
        ORDER BY store, price DESC
    """)
    rows = [{"store": r[0], "brand": r[1], "title": r[2],
             "item_number": r[3], "price": r[4], "link": r[5], "image": r[6]}
            for r in cur.fetchall()]
    conn.close()
    return JSONResponse(rows)


@app.get("/api/analytics/items_per_theme_store")
def api_items_per_theme_store():
    conn = db_connect()
    cur = conn.cursor()
    lsf = _latest_snapshot_filter()
    cur.execute(f"""
        SELECT store, TRIM(theme) as theme, COUNT(DISTINCT item_number) as cnt
        FROM snapshots
        WHERE theme IS NOT NULL AND theme != '' {lsf}
        GROUP BY store, theme
        ORDER BY cnt DESC
    """)
    rows = [{"store": r[0], "theme": r[1], "count": r[2]} for r in cur.fetchall()]
    conn.close()
    return JSONResponse(rows)


@app.get("/api/analytics/new_arrivals_per_store")
def api_new_arrivals_per_store():
    conn = db_connect()
    cur = conn.cursor()
    cur.execute("""
        SELECT store, COUNT(DISTINCT item_number) as cnt
        FROM alerts
        WHERE alert_type='new_arrival'
        GROUP BY store
        ORDER BY cnt DESC
    """)
    rows = [{"store": r[0], "count": r[1]} for r in cur.fetchall()]
    conn.close()
    return JSONResponse(rows)


@app.get("/api/analytics/discounts_per_brand_store")
def api_discounts_per_brand_store():
    conn = db_connect()
    cur = conn.cursor()
    lsf = _latest_snapshot_filter()
    cur.execute(f"""
        SELECT store, UPPER(TRIM(brand)) as brand,
               COUNT(*) as cnt,
               AVG(ROUND((compare_at - price) / compare_at * 100, 1)) as avg_pct
        FROM snapshots
        WHERE compare_at IS NOT NULL AND compare_at > price
          AND brand IS NOT NULL AND brand != '' {lsf}
        GROUP BY store, brand
        ORDER BY cnt DESC
    """)
    rows = [{"store": r[0], "brand": r[1], "count": r[2], "avg_discount_pct": round(r[3], 1) if r[3] else 0}
            for r in cur.fetchall()]
    conn.close()
    return JSONResponse(rows)


# ── Advanced Analysis API ──────────────────────────────────────────────────────

@app.get("/advanced", response_class=HTMLResponse)
def advanced_page():
    f = os.path.join(os.path.dirname(os.path.abspath(__file__)), "advanced.html")
    with open(f, "r", encoding="utf-8") as fh:
        return fh.read()


@app.get("/api/advanced/price_spread")
def api_price_spread():
    """Items available in 2+ stores with min/max/spread prices."""
    conn = db_connect()
    cur = conn.cursor()
    lsf = _latest_snapshot_filter()
    cur.execute(f"""
        SELECT item_number, title, brand, theme, category,
               COUNT(DISTINCT store) as store_count,
               MIN(price) as min_price,
               MAX(price) as max_price,
               ROUND(MAX(price) - MIN(price), 2) as spread,
               ROUND((MAX(price) - MIN(price)) / MIN(price) * 100, 1) as spread_pct,
               GROUP_CONCAT(store || ':' || ROUND(price,2) || ':' || availability, '|') as store_data
        FROM snapshots
        WHERE price IS NOT NULL AND price > 0 {lsf}
        GROUP BY item_number
        HAVING COUNT(DISTINCT store) >= 2
        ORDER BY spread DESC
    """)
    rows = []
    for r in cur.fetchall():
        stores = {}
        for chunk in (r["store_data"] or "").split("|"):
            parts = chunk.split(":")
            if len(parts) >= 3:
                stores[parts[0]] = {"price": float(parts[1]), "availability": parts[2]}
        rows.append({
            "item_number": r["item_number"],
            "title": r["title"] or "",
            "brand": r["brand"] or "",
            "theme": r["theme"] or "",
            "category": r["category"] or "",
            "store_count": r["store_count"],
            "min_price": r["min_price"],
            "max_price": r["max_price"],
            "spread": r["spread"],
            "spread_pct": r["spread_pct"],
            "stores": stores,
        })
    conn.close()
    return JSONResponse(rows)


@app.get("/api/advanced/store_behavior")
def api_store_behavior():
    """Per-store pricing behavior on comparable items."""
    conn = db_connect()
    cur = conn.cursor()
    lsf = _latest_snapshot_filter()

    # Get all comparable items (in 2+ stores)
    cur.execute(f"""
        SELECT item_number
        FROM snapshots
        WHERE price IS NOT NULL AND price > 0 {lsf}
        GROUP BY item_number
        HAVING COUNT(DISTINCT store) >= 2
    """)
    comparable_items = [r[0] for r in cur.fetchall()]

    if not comparable_items:
        conn.close()
        return JSONResponse([])

    placeholders = ",".join("?" * len(comparable_items))
    cur.execute(f"""
        SELECT s.store,
               COUNT(*) as items_carried,
               AVG(s.price) as avg_price,
               SUM(CASE WHEN s.price = m.min_price THEN 1 ELSE 0 END) as cheapest_count,
               SUM(CASE WHEN s.price = m.max_price THEN 1 ELSE 0 END) as priciest_count,
               AVG(ROUND((s.price - m.min_price) / NULLIF(m.min_price,0) * 100, 1)) as avg_premium_pct
        FROM snapshots s
        JOIN (
            SELECT item_number, MIN(price) as min_price, MAX(price) as max_price
            FROM snapshots
            WHERE price IS NOT NULL AND price > 0
              AND item_number IN ({placeholders})
            GROUP BY item_number
        ) m ON s.item_number = m.item_number
        WHERE s.price IS NOT NULL AND s.price > 0
          AND s.item_number IN ({placeholders})
        GROUP BY s.store
        ORDER BY avg_premium_pct ASC
    """, comparable_items + comparable_items)

    rows = []
    for r in cur.fetchall():
        total = r["items_carried"] or 1
        rows.append({
            "store": r["store"],
            "items_carried": r["items_carried"],
            "avg_price": round(r["avg_price"] or 0, 2),
            "cheapest_count": r["cheapest_count"] or 0,
            "cheapest_pct": round((r["cheapest_count"] or 0) / total * 100, 1),
            "priciest_count": r["priciest_count"] or 0,
            "priciest_pct": round((r["priciest_count"] or 0) / total * 100, 1),
            "avg_premium_pct": round(r["avg_premium_pct"] or 0, 1),
        })
    conn.close()
    return JSONResponse(rows)


@app.get("/api/advanced/deal_detector")
def api_deal_detector():
    """Sets discounted in one store but full price in others."""
    conn = db_connect()
    cur = conn.cursor()
    lsf = _latest_snapshot_filter()
    cur.execute(f"""
        SELECT s.item_number, s.title, s.brand, s.theme, s.store,
               s.price, s.compare_at,
               ROUND((s.compare_at - s.price) / s.compare_at * 100, 1) as discount_pct,
               m.min_other_price, m.stores_full_price,
               ROUND(m.min_other_price - s.price, 2) as saving_vs_others
        FROM snapshots s
        JOIN (
            SELECT item_number,
                   MIN(CASE WHEN (compare_at IS NULL OR compare_at <= price) THEN price END) as min_other_price,
                   COUNT(CASE WHEN (compare_at IS NULL OR compare_at <= price) THEN 1 END) as stores_full_price
            FROM snapshots
            WHERE price IS NOT NULL AND price > 0 {lsf}
            GROUP BY item_number
        ) m ON s.item_number = m.item_number
        WHERE s.compare_at IS NOT NULL AND s.compare_at > s.price
          AND m.min_other_price IS NOT NULL
          AND m.min_other_price > s.price
          AND s.price IS NOT NULL {lsf}
        ORDER BY saving_vs_others DESC
    """)
    rows = []
    for r in cur.fetchall():
        rows.append({
            "item_number": r["item_number"],
            "title": r["title"] or "",
            "brand": r["brand"] or "",
            "theme": r["theme"] or "",
            "store": r["store"],
            "price": r["price"],
            "compare_at": r["compare_at"],
            "discount_pct": r["discount_pct"],
            "min_other_price": r["min_other_price"],
            "stores_full_price": r["stores_full_price"],
            "saving_vs_others": r["saving_vs_others"],
        })
    conn.close()
    return JSONResponse(rows)


@app.post("/api/advanced/export")
async def api_advanced_export(request: Request):
    """Export advanced analysis data as Excel."""
    import io
    from fastapi.responses import StreamingResponse as SR
    body = await request.json()
    tab  = body.get("tab", "data")
    rows = body.get("rows", [])
    if not rows:
        return JSONResponse({"error": "No data"})

    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = {"spread":"Price Spread","behavior":"Store Behavior","deals":"Deal Detector"}.get(tab, tab)

    headers = list(rows[0].keys())
    hdr_fill = PatternFill("solid", fgColor="1E293B")
    hdr_font = Font(bold=True, color="94A3B8", size=9)
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h.replace("_"," ").title())
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for ri, row in enumerate(rows, 2):
        for ci, key in enumerate(headers, 1):
            val = row.get(key, "")
            if isinstance(val, str):
                try: val = float(val)
                except: pass
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = Font(size=9)

    ws.freeze_panes = "A2"
    for ci, key in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(ci)].width = 18 if "title" in key else 12

    ts = datetime.now().strftime("%Y%m%d_%H%M")
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return SR(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=lego_{tab}_{ts}.xlsx"}
    )


@app.get("/api/stores/logos")
def api_store_logos():
    """Return logo URLs for all stores."""
    return JSONResponse(get_all_store_logos())


@app.post("/api/stores/fetch_logos")
def api_fetch_all_logos():
    """Trigger logo fetch for all stores (background)."""
    conn = db_connect()
    cur = conn.cursor()
    cur.execute("SELECT name, base_url FROM stores WHERE enabled=1")
    rows = cur.fetchall()
    conn.close()
    hardcoded = [
        ("BRICKSHOP",       "https://brickshop.me"),
        ("PlayOne",         "https://playone.com.lb"),
        ("Ayoub Computers", "https://ayoubcomputers.com"),
    ]
    import threading as _flt
    for name, url in list({r["name"]: r["base_url"] for r in rows}.items()) + hardcoded:
        _flt.Thread(target=fetch_store_logo, args=(url, name), daemon=True).start()
    return JSONResponse({"ok": True, "message": "Fetching logos in background"})


# ── RadarList ──────────────────────────────────────────────────────────────────

@app.get("/radarlist", response_class=HTMLResponse)
def radarlist_page():
    f = os.path.join(os.path.dirname(os.path.abspath(__file__)), "radarlist.html")
    return HTMLResponse(open(f, encoding='utf-8').read())

@app.get("/api/radarlist")
def api_radarlist_get():
    """Get all items on RadarList with current prices."""
    conn = db_connect()
    cur = conn.cursor()
    cur.execute("SELECT * FROM radarlist ORDER BY added_at DESC")
    items = [dict(r) for r in cur.fetchall()]

    # Get current prices for each item
    lsf = _latest_snapshot_filter()
    for item in items:
        cur.execute(f"""
            SELECT store, price, compare_at, availability, link, image_url
            FROM snapshots
            WHERE item_number=? {lsf}
            ORDER BY price ASC
        """, (item["item_number"],))
        prices = [dict(r) for r in cur.fetchall()]
        item["current_prices"] = prices
        item["min_price"] = min((p["price"] for p in prices if p["price"]), default=None)
        item["min_store"] = next((p["store"] for p in prices if p["price"] == item["min_price"]), None)
        item["image_url"] = next((p["image_url"] for p in prices if p.get("image_url")), None)
        # Price drop detection
        if item["added_price"] and item["min_price"]:
            item["price_drop"] = round(item["added_price"] - item["min_price"], 2)
            item["price_drop_pct"] = round((item["added_price"] - item["min_price"]) / item["added_price"] * 100, 1)
        else:
            item["price_drop"] = 0
            item["price_drop_pct"] = 0

    conn.close()
    return JSONResponse(items)

@app.post("/api/radarlist/add")
async def api_radarlist_add(request: Request):
    body = await request.json()
    item_number = body.get("item_number", "").strip()
    if not item_number:
        return JSONResponse({"ok": False, "error": "item_number required"})
    conn = db_connect()
    cur = conn.cursor()
    try:
        # Get latest price info
        lsf = _latest_snapshot_filter()
        cur.execute(f"""
            SELECT title, brand, theme, price, store
            FROM snapshots WHERE item_number=? {lsf}
            ORDER BY price ASC LIMIT 1
        """, (item_number,))
        row = cur.fetchone()
        title  = body.get("title") or (row["title"] if row else "")
        brand  = body.get("brand") or (row["brand"] if row else "")
        theme  = body.get("theme") or (row["theme"] if row else "")
        price  = row["price"] if row else None
        store  = row["store"] if row else None
        cur.execute("""
            INSERT INTO radarlist(item_number, title, brand, theme, added_price, added_store)
            VALUES(?,?,?,?,?,?)
            ON CONFLICT(item_number) DO NOTHING
        """, (item_number, title, brand, theme, price, store))
        conn.commit()
        added = cur.rowcount > 0
    except Exception as e:
        conn.close()
        return JSONResponse({"ok": False, "error": str(e)})
    conn.close()
    return JSONResponse({"ok": True, "added": added})

@app.post("/api/radarlist/remove")
async def api_radarlist_remove(request: Request):
    body = await request.json()
    item_number = body.get("item_number", "").strip()
    conn = db_connect()
    cur = conn.cursor()
    cur.execute("DELETE FROM radarlist WHERE item_number=?", (item_number,))
    conn.commit()
    conn.close()
    return JSONResponse({"ok": True})

@app.get("/api/radarlist/ids")
def api_radarlist_ids():
    """Return just the item_numbers on RadarList — for fast star state loading."""
    conn = db_connect()
    cur = conn.cursor()
    cur.execute("SELECT item_number FROM radarlist")
    ids = [r[0] for r in cur.fetchall()]
    conn.close()
    return JSONResponse(ids)


# ── AI Assistant ───────────────────────────────────────────────────────────────

def build_context() -> str:
    """Build a compact context snapshot for the AI."""
    conn = db_connect()
    cur = conn.cursor()
    lsf = _latest_snapshot_filter()

    # RadarList with current prices (compact)
    cur.execute("SELECT item_number, title, added_price, added_store FROM radarlist ORDER BY added_at DESC")
    radar = [dict(r) for r in cur.fetchall()]
    radar_lines = []
    for item in radar[:10]:
        cur.execute(f"SELECT store, price FROM snapshots WHERE item_number=? {lsf} ORDER BY price ASC", (item["item_number"],))
        prices = [f"{r['store']}:${r['price']}" for r in cur.fetchall() if r['price']]
        min_price = min((float(p.split('$')[1]) for p in prices), default=None)
        drop = round(item["added_price"] - min_price, 2) if item["added_price"] and min_price else 0
        radar_lines.append(f"- #{item['item_number']} {item['title']} | added@${item['added_price']} | now: {', '.join(prices)} | drop:${drop}")

    # Top deals (compact)
    cur.execute(f"""
        SELECT item_number, title, store, price, compare_at,
               ROUND((compare_at-price)/compare_at*100,1) as pct
        FROM snapshots WHERE compare_at>price AND compare_at>0 {lsf}
        ORDER BY pct DESC LIMIT 8
    """)
    deal_lines = [f"- #{r['item_number']} {r['title']} | {r['store']} ${r['price']} (was ${r['compare_at']}, -{r['pct']}%)" for r in cur.fetchall()]

    # Store summary (compact)
    cur.execute(f"SELECT store, COUNT(*) as cnt, ROUND(AVG(price),2) as avg_p FROM snapshots WHERE price>0 {lsf} GROUP BY store ORDER BY avg_p ASC")
    store_lines = [f"- {r['store']}: {r['cnt']} items, avg ${r['avg_p']}" for r in cur.fetchall()]

    last_updated = meta_get("last_updated") or "unknown"
    conn.close()

    return f"""You are BrickRadar AI, a LEGO price assistant for stores in Lebanon.
Last refresh: {last_updated}

RADARLIST ({len(radar_lines)} items):
{chr(10).join(radar_lines) if radar_lines else "Empty"}

TOP DEALS:
{chr(10).join(deal_lines) if deal_lines else "None"}

STORES (cheapest to priciest avg):
{chr(10).join(store_lines)}

Be concise. Format prices as $XX.XX. If data not in context, say so."""


@app.get("/api/ai/test")
def api_ai_test():
    """Debug: test AI config and context building."""
    result = {
        "anthropic_key_set": bool(ANTHROPIC_API_KEY),
        "groq_key_set": bool(GROQ_API_KEY),
        "active_provider": "anthropic" if ANTHROPIC_API_KEY else ("groq" if GROQ_API_KEY else "none"),
    }
    try:
        ctx = build_context()
        result["context_length"] = len(ctx)
        result["context_ok"] = True
    except Exception as e:
        import traceback
        result["context_ok"] = False
        result["context_error"] = str(e)
        result["traceback"] = traceback.format_exc()
    return JSONResponse(result)

@app.post("/api/ai/chat")
async def api_ai_chat(request: Request):
    if not ANTHROPIC_API_KEY and not GROQ_API_KEY:
        return JSONResponse({"ok": False, "error": "No AI API key configured. Add ANTHROPIC_API_KEY or GROQ_API_KEY to your .env file."}, status_code=500)

    body = await request.json()
    messages = body.get("messages", [])
    if not messages:
        return JSONResponse({"ok": False, "error": "No messages"}, status_code=400)

    try:
        context = build_context()
    except Exception as e:
        print(f"[AI] build_context error: {e}")
        import traceback; traceback.print_exc()
        return JSONResponse({"ok": False, "error": f"Context error: {e}"}, status_code=500)

    # Choose provider — prefer Anthropic, fall back to Groq
    use_groq = not ANTHROPIC_API_KEY and bool(GROQ_API_KEY)

    async def stream():
        try:
            async with httpx.AsyncClient(timeout=60) as client:
                if use_groq:
                    # Groq API (OpenAI-compatible) with streaming
                    groq_messages = [{"role": "system", "content": context}] + messages
                    async with client.stream(
                        "POST",
                        "https://api.groq.com/openai/v1/chat/completions",
                        headers={
                            "Authorization": f"Bearer {GROQ_API_KEY}",
                            "Content-Type": "application/json",
                        },
                        json={
                            "model": "llama-3.1-8b-instant",
                            "messages": groq_messages,
                            "max_tokens": 1024,
                            "stream": True,
                        }
                    ) as resp:
                        if resp.status_code != 200:
                            error_body = await resp.aread()
                            err = error_body.decode()[:300]
                            print(f"[AI] Groq error {resp.status_code}: {err}")
                            yield f"data: {json.dumps({'text': f'⚠ Groq API error {resp.status_code}: {err}'})}\n\n"
                            yield "data: [DONE]\n\n"
                            return
                        async for line in resp.aiter_lines():
                            if line.startswith("data: "):
                                data = line[6:]
                                if data == "[DONE]": break
                                try:
                                    evt = json.loads(data)
                                    text = evt.get("choices", [{}])[0].get("delta", {}).get("content", "")
                                    if text:
                                        yield f"data: {json.dumps({'text': text})}\n\n"
                                except Exception:
                                    pass
                else:
                    # Anthropic API with streaming
                    async with client.stream("POST", "https://api.anthropic.com/v1/messages",
                        headers={
                            "x-api-key": ANTHROPIC_API_KEY,
                            "anthropic-version": "2023-06-01",
                            "content-type": "application/json",
                        },
                        json={
                            "model": "claude-haiku-4-5-20251001",
                            "max_tokens": 1024,
                            "system": context,
                            "messages": messages,
                            "stream": True,
                        }
                    ) as resp:
                        if resp.status_code != 200:
                            error_body = await resp.aread()
                            err_text = error_body.decode()[:300]
                            print(f"[AI] Anthropic error {resp.status_code}: {err_text}")
                            yield f"data: {json.dumps({'text': f'⚠ API error {resp.status_code}: {err_text}'})}\n\n"
                            yield "data: [DONE]\n\n"
                            return
                        async for line in resp.aiter_lines():
                            if line.startswith("data: "):
                                data = line[6:]
                                if data == "[DONE]": break
                                try:
                                    evt = json.loads(data)
                                    if evt.get("type") == "content_block_delta":
                                        text = evt.get("delta", {}).get("text", "")
                                        if text:
                                            yield f"data: {json.dumps({'text': text})}\n\n"
                                except Exception:
                                    pass
        except Exception as e:
            print(f"[AI] stream error: {e}")
            import traceback; traceback.print_exc()
            yield f"data: {json.dumps({'text': f'⚠ Error: {str(e)}'})}\n\n"
        yield "data: [DONE]\n\n"

    return StreamingResponse(stream(), media_type="text/event-stream")


@app.post("/api/ai/analyze_store")
async def api_ai_analyze_store(request: Request):
    """Use AI to analyze a store and suggest scraper configuration."""
    if not ANTHROPIC_API_KEY and not GROQ_API_KEY:
        return JSONResponse({"ok": False, "error": "No AI API key configured."})

    body = await request.json()
    url = body.get("url", "")
    platform = body.get("platform", "")
    product_count = body.get("product_count", 0)
    samples = body.get("samples", [])

    prompt = f"""You are analyzing a LEGO/toy store to suggest the best scraper configuration for BrickRadar.

Store URL: {url}
Detected platform: {platform}
Products found: {product_count}
Sample products: {json.dumps(samples[:5], indent=2)}

Based on this data, suggest the optimal configuration. Respond ONLY with a valid JSON object, no explanation, no markdown:
{{
  "store_name": "suggested store name (short, clean, e.g. KLAPTAP)",
  "collection_slug": "slug if store has mixed products and needs filtering (e.g. lego), or null",
  "new_arrivals_slug": "slug for new arrivals collection if detectable from samples, or null",
  "vat_multiplier": 1.0,
  "lego_only": false,
  "warnings": ["any issues or notes about this store"],
  "confidence": "high/medium/low"
}}

Rules:
- store_name: extract from domain, capitalize properly
- collection_slug: only suggest if store clearly has non-LEGO products mixed in
- lego_only: true only if store has many non-LEGO brands mixed in the samples
- vat_multiplier: 1.0 unless you detect ex-VAT pricing patterns
- warnings: note if site has Cloudflare, limited products, or unusual structure"""

    try:
        use_groq = not ANTHROPIC_API_KEY and bool(GROQ_API_KEY)
        async with httpx.AsyncClient(timeout=30) as client:
            if use_groq:
                resp = await client.post(
                    "https://api.groq.com/openai/v1/chat/completions",
                    headers={"Authorization": f"Bearer {GROQ_API_KEY}", "Content-Type": "application/json"},
                    json={"model": "llama-3.1-8b-instant", "messages": [{"role": "user", "content": prompt}], "max_tokens": 512}
                )
                data = resp.json()
                text = data["choices"][0]["message"]["content"].strip()
            else:
                resp = await client.post(
                    "https://api.anthropic.com/v1/messages",
                    headers={"x-api-key": ANTHROPIC_API_KEY, "anthropic-version": "2023-06-01", "content-type": "application/json"},
                    json={"model": "claude-haiku-4-5-20251001", "max_tokens": 512, "messages": [{"role": "user", "content": prompt}]}
                )
                data = resp.json()
                text = data["content"][0]["text"].strip()

        # Parse JSON from response
        import re as _re
        json_match = _re.search(r'\{[\s\S]*\}', text)
        if not json_match:
            return JSONResponse({"ok": False, "error": "AI returned unexpected format"})
        suggestion = json.loads(json_match.group())
        return JSONResponse({"ok": True, "suggestion": suggestion})

    except Exception as e:
        print(f"[AI analyze_store] error: {e}")
        return JSONResponse({"ok": False, "error": str(e)})


@app.post("/api/ai/discover-stores")
async def api_discover_stores(request: Request):
    """Use DuckDuckGo search + AI to find real LEGO stores in a given region."""
    if not ANTHROPIC_API_KEY and not GROQ_API_KEY:
        return JSONResponse({"ok": False, "error": "No AI API key configured."}, status_code=500)

    body = await request.json()
    region = body.get("region", "Lebanon").strip()

    hardcoded = ["thebrickmania.com", "bricking.com", "klaptap.com", "brickshop.me",
                 "playone.com.lb", "joueclubliban.com", "thetoystorelb.com"]

    async def do_ddg_search(query: str) -> list:
        """Search DuckDuckGo and return list of {title, url, snippet} results."""
        results = []
        try:
            async with httpx.AsyncClient(timeout=15, headers={
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
            }) as client:
                # DuckDuckGo HTML search
                r = await client.get("https://html.duckduckgo.com/html/",
                    params={"q": query, "kl": "wt-wt"},
                    follow_redirects=True
                )
                soup = BeautifulSoup(r.text, "lxml")
                for result in soup.select(".result")[:8]:
                    title_el = result.select_one(".result__title")
                    url_el = result.select_one(".result__url")
                    snippet_el = result.select_one(".result__snippet")
                    if title_el and url_el:
                        url = url_el.get_text(strip=True)
                        if not url.startswith("http"):
                            url = "https://" + url
                        results.append({
                            "title": title_el.get_text(strip=True),
                            "url": url,
                            "snippet": snippet_el.get_text(strip=True) if snippet_el else ""
                        })
        except Exception as e:
            print(f"[DDG] Search error: {e}")
        return results

    async def stream():
        try:
            yield f"data: {json.dumps({'text': '', 'status': 'searching'})}\n\n"

            # Run 2 searches
            queries = [
                f"LEGO store online shop {region} buy sets",
                f"toy store {region} LEGO official retailer website"
            ]
            all_results = []
            async with httpx.AsyncClient(timeout=15, headers={"User-Agent": "Mozilla/5.0"}) as client:
                for query in queries:
                    try:
                        r = await client.get("https://html.duckduckgo.com/html/",
                            params={"q": query, "kl": "wt-wt"},
                            follow_redirects=True
                        )
                        soup = BeautifulSoup(r.text, "lxml")
                        for result in soup.select(".result")[:6]:
                            title_el = result.select_one(".result__title")
                            url_el = result.select_one(".result__url")
                            snippet_el = result.select_one(".result__snippet")
                            if title_el and url_el:
                                url = url_el.get_text(strip=True).strip()
                                if not url.startswith("http"):
                                    url = "https://" + url
                                # Skip already tracked
                                if not any(h in url for h in hardcoded):
                                    all_results.append({
                                        "title": title_el.get_text(strip=True),
                                        "url": url,
                                        "snippet": snippet_el.get_text(strip=True) if snippet_el else ""
                                    })
                    except Exception as e:
                        print(f"[DDG] error: {e}")

            # Deduplicate by domain
            seen = set()
            unique_results = []
            for r in all_results:
                try:
                    import urllib.parse
                    domain = urllib.parse.urlparse(r["url"]).netloc.replace("www.", "")
                    if domain and domain not in seen:
                        seen.add(domain)
                        unique_results.append(r)
                except:
                    pass

            yield f"data: {json.dumps({'text': '', 'status': f'found {len(unique_results)} results, analyzing...'})}\n\n"

            if not unique_results:
                yield f"data: {json.dumps({'text': '[]', 'status': 'done'})}\n\n"
                yield "data: [DONE]\n\n"
                return

            # Pass real results to AI for filtering and formatting
            prompt = f"""You are analyzing web search results to find legitimate LEGO stores in {region}.

Here are the search results:
{json.dumps(unique_results, indent=2)}

From these results, identify stores that:
- Actually sell LEGO products online (not just mention LEGO)
- Have a real e-commerce website
- Are based in or specifically serve {region}
- Are NOT social media pages, news articles, or directories

Already tracked (exclude): {', '.join(hardcoded)}

Return ONLY a JSON array, no other text:
[
  {{
    "name": "Store Name",
    "url": "https://exact-url.com",
    "platform_guess": "shopify/woocommerce/bigcommerce/unknown",
    "notes": "one line about the store"
  }}
]

If none qualify, return an empty array: []"""

            use_groq = not ANTHROPIC_API_KEY and bool(GROQ_API_KEY)
            full_text = ""

            async with httpx.AsyncClient(timeout=60) as client:
                if use_groq:
                    async with client.stream("POST",
                        "https://api.groq.com/openai/v1/chat/completions",
                        headers={"Authorization": f"Bearer {GROQ_API_KEY}", "Content-Type": "application/json"},
                        json={"model": "llama-3.1-8b-instant",
                              "messages": [{"role": "user", "content": prompt}],
                              "max_tokens": 1024, "stream": True}
                    ) as resp:
                        async for line in resp.aiter_lines():
                            if line.startswith("data: "):
                                data = line[6:]
                                if data == "[DONE]": break
                                try:
                                    evt = json.loads(data)
                                    text = evt.get("choices", [{}])[0].get("delta", {}).get("content", "")
                                    if text: full_text += text
                                except: pass
                else:
                    async with client.stream("POST", "https://api.anthropic.com/v1/messages",
                        headers={"x-api-key": ANTHROPIC_API_KEY, "anthropic-version": "2023-06-01", "content-type": "application/json"},
                        json={"model": "claude-haiku-4-5-20251001", "max_tokens": 1024,
                              "messages": [{"role": "user", "content": prompt}], "stream": True}
                    ) as resp:
                        async for line in resp.aiter_lines():
                            if line.startswith("data: "):
                                data = line[6:]
                                if data == "[DONE]": break
                                try:
                                    evt = json.loads(data)
                                    if evt.get("type") == "content_block_delta":
                                        text = evt.get("delta", {}).get("text", "")
                                        if text: full_text += text
                                except: pass

            yield f"data: {json.dumps({'text': full_text, 'status': 'done'})}\n\n"

        except Exception as e:
            print(f"[Discover] error: {e}")
            import traceback; traceback.print_exc()
            yield f"data: {json.dumps({'text': '[]', 'status': 'error', 'error': str(e)})}\n\n"
        yield "data: [DONE]\n\n"

    return StreamingResponse(stream(), media_type="text/event-stream")


# ── Dashboard Export ───────────────────────────────────────────────────────────

@app.get("/api/export")
def api_export(request: Request):
    """Export filtered dashboard data as CSV, Excel, or JSON.
    Queries snapshots directly — no dependency on merged catalog."""
    import io, csv
    from fastapi.responses import StreamingResponse as SR

    qp = dict(request.query_params)
    fmt = qp.get("fmt", "csv").lower()

    store_names = get_all_store_names()
    raw_stores = request.query_params.getlist("stores")
    stores = order_stores([s for s in raw_stores if s in store_names]) or store_names

    selected_category = qp.get("category", "All")
    selected_theme    = qp.get("theme", "All")
    selected_brand    = qp.get("brand", "All")
    search_item       = qp.get("search_item", "").strip()
    only_deals        = qp.get("only_deals", "0") == "1"
    only_instock      = qp.get("only_instock", "0") == "1"

    # Query latest snapshot per store directly
    conn = db_connect()
    cur = conn.cursor()
    placeholders = ",".join("?" * len(stores))
    cur.execute(f"""
        SELECT s.store, s.item_number, s.title, s.brand, s.theme, s.category,
               s.price, s.compare_at, s.availability, s.link
        FROM snapshots s
        INNER JOIN (
            SELECT store, MAX(captured_at) as max_at FROM snapshots GROUP BY store
        ) latest ON s.store = latest.store AND s.captured_at = latest.max_at
        WHERE s.store IN ({placeholders})
        ORDER BY s.item_number, s.store
    """, stores)
    db_rows = cur.fetchall()
    conn.close()

    # Group by item_number
    from collections import defaultdict
    items = defaultdict(lambda: {"item_number":"","brand":"","title":"","theme":"","category":"","stores":{}})
    for r in db_rows:
        key = r["item_number"] or r["title"] or ""
        it = items[key]
        it["item_number"] = r["item_number"] or ""
        it["brand"]    = (r["brand"] or "").strip().upper()
        it["title"]    = r["title"] or ""
        it["theme"]    = r["theme"] or ""
        it["category"] = r["category"] or ""
        disc_pct = 0
        if r["compare_at"] and r["price"] and r["compare_at"] > r["price"]:
            disc_pct = round((r["compare_at"] - r["price"]) / r["compare_at"] * 100, 1)
        it["stores"][r["store"]] = {
            "price": r["price"],
            "availability": r["availability"] or "",
            "discount_pct": disc_pct,
            "link": r["link"] or "",
        }

    # Apply filters
    rows_out = []
    for it in items.values():
        if selected_category != "All" and it["category"] != selected_category: continue
        if selected_theme     != "All" and it["theme"]    != selected_theme:    continue
        if selected_brand     != "All" and it["brand"]    != selected_brand.upper(): continue
        if search_item and search_item not in it["item_number"]: continue
        if only_deals and not any((v["discount_pct"] or 0) > 0 for v in it["stores"].values()): continue
        if only_instock and not any(v["availability"] in ("available","in_stock") for v in it["stores"].values()): continue

        prices = [v["price"] for v in it["stores"].values() if v["price"] is not None]
        lp = min(prices) if prices else None

        row = {
            "item_number": it["item_number"],
            "brand":       it["brand"],
            "title":       it["title"],
            "theme":       it["theme"],
            "category":    it["category"],
            "lowest_price": f"{lp:.2f}" if lp is not None else "",
        }
        for s in stores:
            off = it["stores"].get(s, {})
            row[f"{s}_price"]        = f"{off['price']:.2f}" if off.get("price") is not None else ""
            row[f"{s}_avail"]        = off.get("availability", "")
            row[f"{s}_discount_pct"] = f"{off['discount_pct']:.0f}" if off.get("discount_pct") else ""
            row[f"{s}_link"]         = off.get("link", "")
        rows_out.append(row)

    if not rows_out:
        return JSONResponse({"error": "No data for current filters"})

    headers_row = list(rows_out[0].keys())
    ts = datetime.now().strftime("%Y%m%d_%H%M")

    # ── CSV ──
    if fmt == "csv":
        buf = io.StringIO()
        w = csv.DictWriter(buf, fieldnames=headers_row)
        w.writeheader()
        w.writerows(rows_out)
        buf.seek(0)
        return SR(
            iter([buf.getvalue().encode("utf-8")]),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f"attachment; filename=lego_tracker_{ts}.csv"}
        )

    # ── Excel ──
    if fmt == "xlsx":
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "LEGO Tracker"

        hdr_fill = PatternFill("solid", fgColor="1E293B")
        hdr_font = Font(bold=True, color="94A3B8", size=9)
        for ci, h in enumerate(headers_row, 1):
            cell = ws.cell(row=1, column=ci, value=h.replace("_", " ").title())
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        deal_fill = PatternFill("solid", fgColor="1A3020")
        for ri, row in enumerate(rows_out, 2):
            for ci, key in enumerate(headers_row, 1):
                val = row[key]
                if key.endswith("_price") or key == "lowest_price":
                    try: val = float(val)
                    except: pass
                elif key.endswith("_discount_pct"):
                    try: val = int(val)
                    except: pass
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.font = Font(size=9)
                if key == "lowest_price" and val:
                    cell.font = Font(size=9, bold=True, color="F59E0B")
                if key.endswith("_discount_pct") and val:
                    cell.fill = deal_fill

        ws.row_dimensions[1].height = 28
        ws.freeze_panes = "A2"
        col_widths = {"item_number":14,"brand":10,"title":42,"theme":18,"category":16,"lowest_price":13}
        for ci, key in enumerate(headers_row, 1):
            ws.column_dimensions[get_column_letter(ci)].width = col_widths.get(key, 16 if "link" in key else 11)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return SR(
            iter([buf.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=lego_tracker_{ts}.xlsx"}
        )

    # ── JSON ──
    return JSONResponse(rows_out)


@app.get("/export/pdf", response_class=HTMLResponse)
def export_pdf_page(request: Request):
    """Returns a print-ready HTML page with ALL filtered rows — open and Ctrl+P to PDF."""
    qp = dict(request.query_params)
    store_names = get_all_store_names()
    raw_stores = request.query_params.getlist("stores")
    stores = order_stores([s for s in raw_stores if s in store_names]) or store_names

    selected_category = qp.get("category", "All")
    selected_theme    = qp.get("theme", "All")
    selected_brand    = qp.get("brand", "All")
    search_item       = qp.get("search_item", "").strip()
    only_deals        = qp.get("only_deals", "0") == "1"
    only_instock      = qp.get("only_instock", "0") == "1"

    conn = db_connect()
    cur = conn.cursor()
    placeholders = ",".join("?" * len(stores))
    cur.execute(f"""
        SELECT s.store, s.item_number, s.title, s.brand, s.theme, s.category,
               s.price, s.compare_at, s.availability, s.link
        FROM snapshots s
        INNER JOIN (
            SELECT store, MAX(captured_at) as max_at FROM snapshots GROUP BY store
        ) latest ON s.store = latest.store AND s.captured_at = latest.max_at
        WHERE s.store IN ({placeholders})
        ORDER BY s.item_number, s.store
    """, stores)
    db_rows = cur.fetchall()
    conn.close()

    from collections import defaultdict
    items = defaultdict(lambda: {"item_number":"","brand":"","title":"","theme":"","category":"","stores":{}})
    for r in db_rows:
        key = r["item_number"] or r["title"] or ""
        it = items[key]
        it["item_number"] = r["item_number"] or ""
        it["brand"]    = (r["brand"] or "").strip().upper()
        it["title"]    = r["title"] or ""
        it["theme"]    = r["theme"] or ""
        it["category"] = r["category"] or ""
        disc_pct = 0
        if r["compare_at"] and r["price"] and r["compare_at"] > r["price"]:
            disc_pct = round((r["compare_at"] - r["price"]) / r["compare_at"] * 100, 1)
        it["stores"][r["store"]] = {"price": r["price"], "discount_pct": disc_pct, "availability": r["availability"] or ""}

    rows_out = []
    for it in items.values():
        if selected_category != "All" and it["category"] != selected_category: continue
        if selected_theme     != "All" and it["theme"]    != selected_theme:    continue
        if selected_brand     != "All" and it["brand"]    != selected_brand.upper(): continue
        if search_item and search_item not in it["item_number"]: continue
        if only_deals and not any((v["discount_pct"] or 0) > 0 for v in it["stores"].values()): continue
        if only_instock and not any(v["availability"] in ("available","in_stock") for v in it["stores"].values()): continue
        prices = [v["price"] for v in it["stores"].values() if v["price"] is not None]
        it["lowest_price"] = min(prices) if prices else None
        rows_out.append(it)

    store_headers = "".join(f"<th>{s}</th>" for s in stores)
    ts = datetime.now().strftime("%Y-%m-%d %H:%M")

    def store_cells(it):
        cells = ""
        for s in stores:
            off = it["stores"].get(s)
            if off and off["price"] is not None:
                disc = f' <span style="color:#b45309">-{off["discount_pct"]:.0f}%</span>' if off["discount_pct"] else ""
                cells += f'<td>${off["price"]:.2f}{disc}</td>'
            else:
                cells += "<td>—</td>"
        return cells

    rows_html = ""
    for it in rows_out:
        lp = f'<b>${it["lowest_price"]:.2f}</b>' if it["lowest_price"] else "—"
        rows_html += f"""<tr>
          <td>{it["item_number"]}</td>
          <td>{it["brand"]}</td>
          <td>{it["title"]}</td>
          <td>{it["theme"]}</td>
          <td>{it["category"]}</td>
          {store_cells(it)}
          <td>{lp}</td>
        </tr>"""

    filters_desc = " | ".join(filter(None, [
        f"Stores: {', '.join(stores)}" if stores != store_names else "",
        f"Category: {selected_category}" if selected_category != "All" else "",
        f"Theme: {selected_theme}" if selected_theme != "All" else "",
        f"Brand: {selected_brand}" if selected_brand != "All" else "",
        f"Search: {search_item}" if search_item else "",
        "Deals only" if only_deals else "",
        "In stock only" if only_instock else "",
    ])) or "All products"

    html = f"""<!doctype html><html><head><meta charset="UTF-8">
<title>LEGO Tracker Export — {ts}</title>
<style>
  body {{ font-family: Arial, sans-serif; font-size: 8pt; color: #111; margin: 0; padding: 1cm; }}
  h1 {{ font-size: 13pt; margin: 0 0 2px; }}
  .meta {{ font-size: 7.5pt; color: #666; margin-bottom: 8px; }}
  table {{ width: 100%; border-collapse: collapse; }}
  th {{ background: #1e293b; color: #fff; padding: 4px 6px; text-align: left; font-size: 7.5pt; }}
  td {{ padding: 3px 6px; border-bottom: 1px solid #e5e7eb; vertical-align: top; }}
  tr:nth-child(even) td {{ background: #f9fafb; }}
  @page {{ margin: 1.5cm; size: A4 landscape; }}
  @media screen {{ body {{ max-width: 1400px; margin: auto; padding: 20px; }} }}
  .print-btn {{ margin-bottom: 12px; padding: 8px 18px; background: #4f46e5; color: #fff; border: none; border-radius: 6px; cursor: pointer; font-size: 10pt; }}
  @media print {{ .print-btn {{ display: none; }} }}
</style>
</head><body>
<button class="print-btn" onclick="window.print()">&#128196; Print / Save as PDF</button>
<h1>LEGO Tracker — Product Export</h1>
<div class="meta">Generated: {ts} &nbsp;|&nbsp; {len(rows_out)} products &nbsp;|&nbsp; {filters_desc}</div>
<table>
  <thead><tr><th>Item #</th><th>Brand</th><th>Title</th><th>Theme</th><th>Category</th>{store_headers}<th>Lowest</th></tr></thead>
  <tbody>{rows_html}</tbody>
</table>
</body></html>"""
    return html
