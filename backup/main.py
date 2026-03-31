"""
main.py — thin orchestrator for RadarOS / BrickRadar.
Wires core/ and modules/brickradar/ together into a FastAPI app.
"""

import io
import json
import os
import re
import sqlite3
import time
import threading
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

from fastapi import FastAPI, Request
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates

# ── Core imports ───────────────────────────────────────────────────────────────
from core.db import (
    db_connect, db_init,
    meta_get, meta_set,
    alerts_unread_count, alerts_mark_read,
    persist_snapshot, compute_alerts,
    radarlist_add, radarlist_remove, radarlist_get_ids,
)
from core.engine import merge_catalogs, refresh_all as _core_refresh_all
from core.models import StoreOffer
from core.utils import safe_float, extract_item_number, compute_discount_pct, utc_now_iso
from core.scrapers.shopify import fetch_shopify_store, fetch_new_arrival_items
from core.scrapers.bigcommerce import fetch_bigcommerce_store
from core.ai import build_context, stream_chat, stream_discover_stores, analyze_store

# ── Module imports ─────────────────────────────────────────────────────────────
from modules.brickradar.config import (
    MODULE,
    SHOPIFY_STORES,
    BIGCOMMERCE_STORES,
    OFFICIAL_STORES,
    INTERNATIONAL_STORES,
    NEW_ARRIVAL_COLLECTIONS,
    HARDCODED_STORE_URLS,
    LEGO_THEMES,
)
from modules.brickradar.scrapers import (
    fetch_brickshop,
    fetch_playone,
    fetch_html_stores,
    normalize_theme_category_from_shopify,
)
from modules.brickradar.scrapers_official import (
    fetch_cada,
    fetch_mouldking,
    fetch_lego_com,
)

# ── Paths ──────────────────────────────────────────────────────────────────────
APP_DIR       = os.path.dirname(os.path.abspath(__file__))
TEMPLATES_DIR = os.path.join(APP_DIR, "Brickradar", "app", "templates")
STATIC_DIR          = os.path.join(APP_DIR, "Brickradar", "app", "static")
PLATFORM_STATIC_DIR = os.path.join(APP_DIR, "static", "platform")
PLATFORM_TMPL_DIR   = os.path.join(APP_DIR, "templates")
os.makedirs(PLATFORM_STATIC_DIR, exist_ok=True)
os.makedirs(PLATFORM_TMPL_DIR, exist_ok=True)
DB_PATH       = os.path.join(APP_DIR, "Brickradar", "app", "data", "lego_tracker.sqlite3")
TEMPLATE_FILE = "dashboard.html"

# ── API keys ───────────────────────────────────────────────────────────────────
from dotenv import load_dotenv
load_dotenv()
ANTHROPIC_API_KEY   = os.getenv("ANTHROPIC_API_KEY", "")
ADMIN_EMAIL         = os.getenv("ADMIN_EMAIL", "khwarizmesarl@gmail.com")
ADMIN_PASSWORD_HASH = os.getenv("ADMIN_PASSWORD_HASH", "")
ADMIN_PASSWORD_SALT = os.getenv("ADMIN_PASSWORD_SALT", "")
ADMIN_SESSION_HOURS = int(os.getenv("ADMIN_SESSION_HOURS", "24"))
GMAIL_APP_PASSWORD  = os.getenv("GMAIL_APP_PASSWORD", "")
_admin_sessions: dict = {}
_otp_store:      dict = {}
_login_attempts: dict = {}
GROQ_API_KEY      = os.getenv("GROQ_API_KEY", "")

# ── Init ───────────────────────────────────────────────────────────────────────
db_init(DB_PATH)

app       = FastAPI()
templates = Jinja2Templates(directory=TEMPLATES_DIR)
templates.env.filters["tojson"] = lambda v: json.dumps(v)

if os.path.isdir(STATIC_DIR):
    app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")
if os.path.isdir(PLATFORM_STATIC_DIR):
    app.mount("/assets", StaticFiles(directory=PLATFORM_STATIC_DIR), name="platform_static")

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                  "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0 Safari/537.36",
}

# ── Helpers ────────────────────────────────────────────────────────────────────

def get_all_store_names() -> List[str]:
    names = list(SHOPIFY_STORES.keys()) + ["BRICKSHOP", "PlayOne"]
    names += list(BIGCOMMERCE_STORES.keys())
    conn = db_connect(DB_PATH)
    cur  = conn.cursor()
    cur.execute("SELECT name FROM stores WHERE enabled=1")
    for r in cur.fetchall():
        if r["name"] not in names:
            names.append(r["name"])
    conn.close()
    from core.utils import order_stores
    return order_stores(names)

def _lsf(source_type: str = "local") -> str:
    """Latest snapshot filter by source_type. Defaults to local."""
    if source_type == "local":
        return f"AND source_type='{source_type}' AND captured_at IN (SELECT MAX(captured_at) FROM snapshots GROUP BY store)"
    return f"AND source_type='{source_type}' AND id IN (SELECT MAX(id) FROM snapshots WHERE source_type='{source_type}' GROUP BY store, item_number)"


# ── Dashboard ──────────────────────────────────────────────────────────────────

@app.get("/", response_class=HTMLResponse)
def dashboard(request: Request):
    qp          = request.query_params
    source_tier = qp.get("tier", "local")   # local | official | international

    # Filter store names by tier
    all_store_names = get_all_store_names()

    # Build tier → store name mapping from config
    _local_stores       = (
        set(SHOPIFY_STORES.keys()) |
        set(BIGCOMMERCE_STORES.keys()) |
        {"BRICKSHOP", "PlayOne", "Brix & Figures", "Thetoystorelb", "Joueclubliban"}
    )
    _official_stores    = set(OFFICIAL_STORES.keys())
    _intl_stores        = set(INTERNATIONAL_STORES.keys())

    if source_tier == "official":
        store_names = [s for s in all_store_names if s in _official_stores]
        if not store_names:
            store_names = list(_official_stores)
    elif source_tier == "international":
        store_names = [s for s in all_store_names if s in _intl_stores]
        if not store_names:
            store_names = list(_intl_stores)
    else:  # local (default)
        store_names = [s for s in all_store_names if s in _local_stores] or all_store_names

    conn = db_connect(DB_PATH)
    cur  = conn.cursor()
    cur.execute("""
        SELECT s.* FROM snapshots s
        INNER JOIN (
            SELECT store, item_number, MAX(id) AS max_id
            FROM snapshots GROUP BY store, item_number
        ) t ON s.id = t.max_id
    """)
    rows = cur.fetchall()
    conn.close()

    latest_by_store_item: Dict[Tuple[str, str], sqlite3.Row] = {}
    for r in rows:
        item_key = r["item_number"] or f"SID_{r['store']}_{r['id']}"
        latest_by_store_item[(r["store"], item_key)] = r

    catalogs = []
    for s in store_names:
        cat: Dict[str, Dict[str, Any]] = {}
        for (st, item), r in latest_by_store_item.items():
            if st != s or not item:
                continue
            offer = StoreOffer(
                price=r["price"],
                availability=r["availability"] or "N/A",
                link=r["link"] or "",
                discount_pct=compute_discount_pct(r["price"], r["compare_at"]),
            )
            try:
                img_list = json.loads(r["images_json"]) if r["images_json"] else []
            except Exception:
                img_list = []
            if not img_list and r["image_url"]:
                img_list = [r["image_url"]]
            cat[item] = {
                "item_number": item,
                "title":    r["title"] or "",
                "theme":    r["theme"] or "",
                "category": r["category"] or "",
                "image_url":   r["image_url"] or "",
                "image_list":  img_list,
                "vendor":  "",
                "brand":   (r["brand"] or "").strip().upper() or "LEGO",
                "compare_at": r["compare_at"],
                "stores":  {s: offer},
            }
        catalogs.append(cat)

    merged = merge_catalogs(catalogs)

    qp = request.query_params
    selected_category   = qp.get("category",  "All")
    selected_theme      = qp.get("theme",      "All")
    selected_brand      = qp.get("brand",      "All")
    search_item         = qp.get("search_item", "").strip()
    only_deals          = qp.get("only_deals",      "0") == "1"
    only_alerts         = qp.get("only_alerts",     "0") == "1"
    only_comparable     = qp.get("only_comparable",  "0") == "1"
    only_instock        = qp.get("only_instock",    "0") == "1"
    alert_type_filter   = qp.get("alert_type", "all")
    sort                = qp.get("sort",  "item")
    order               = qp.get("order", "asc")
    per_page            = int(qp.get("per_page", "1000") or "1000")
    page                = max(1, int(qp.get("page", "1") or "1"))

    raw_selected_stores = request.query_params.getlist("stores")
    if raw_selected_stores:
        from core.utils import order_stores
        stores = order_stores([s for s in raw_selected_stores if s in store_names]) or store_names
        selected_stores = stores
    else:
        stores          = store_names
        selected_stores = []

    all_categories = ["All"] + sorted({(r.get("category") or "").strip() for r in merged.values() if (r.get("category") or "").strip()})
    all_themes     = ["All"] + sorted({(r.get("theme")    or "").strip() for r in merged.values() if (r.get("theme")    or "").strip()})
    all_brands     = ["All"] + sorted({(r.get("brand")    or "").strip().upper() for r in merged.values() if (r.get("brand") or "").strip()})

    # Build alert lookup
    ALERT_PRIORITY = {"price_drop": 0, "new_arrival": 1, "new_in_store": 2, "price_increase": 3}
    _ac = db_connect(DB_PATH)
    _acc = _ac.cursor()
    _acc.execute("SELECT item_number, store, alert_type, old_price, new_price FROM alerts WHERE unread=1 ORDER BY id DESC")
    item_store_alerts: Dict[str, Dict[str, Dict]] = {}
    for ar in _acc.fetchall():
        iid, sname, atype = ar["item_number"], ar["store"], ar["alert_type"] or "price_change"
        if iid not in item_store_alerts:
            item_store_alerts[iid] = {}
        cur_p = ALERT_PRIORITY.get(item_store_alerts[iid].get(sname, {}).get("type", "new_in_store"), 99)
        new_p = ALERT_PRIORITY.get(atype, 99)
        if sname not in item_store_alerts[iid] or new_p < cur_p:
            item_store_alerts[iid][sname] = {"type": atype, "store": sname, "old_price": ar["old_price"], "new_price": ar["new_price"]}
    _ac.close()

    def _get_item_alert(item_number, selected):
        store_map = item_store_alerts.get(item_number, {})
        best, best_p = None, 99
        for s in selected:
            a = store_map.get(s)
            if a:
                p = ALERT_PRIORITY.get(a["type"], 99)
                if p < best_p:
                    best_p, best = p, a
        return best

    def _get_item_alert_types(item_number, selected):
        store_map = item_store_alerts.get(item_number, {})
        return {store_map[s]["type"] for s in selected if s in store_map}

    filtered = []
    for rec in merged.values():
        if selected_category != "All" and (rec.get("category") or "").strip() != selected_category:
            continue
        if selected_theme != "All" and (rec.get("theme") or "").strip() != selected_theme:
            continue
        if selected_brand != "All" and (rec.get("brand") or "").strip().upper() != selected_brand.upper():
            continue
        if search_item and search_item not in (rec.get("item_number") or ""):
            continue

        rec2 = dict(rec)
        rec2["stores"] = {s: rec["stores"].get(s) for s in stores if s in rec["stores"]}
        if not any(rec2["stores"].values()):
            continue

        raw_img = rec2.get("image_url") or ""
        rec2["image_proxy"]      = re.sub(r"_[0-9]+x[0-9]*\.", ".", raw_img) if raw_img else ""
        rec2["image_list_clean"] = [re.sub(r"_[0-9]+x[0-9]*\.", ".", u) for u in (rec2.get("image_list") or [])[:4]]
        rec2["image_list_json"]  = json.dumps(rec2["image_list_clean"])

        lp, ls = None, None
        for s in stores:
            off = rec2["stores"].get(s)
            if off and off.price is not None and (lp is None or off.price < lp):
                lp, ls = off.price, s
        rec2["lowest_price"] = lp
        rec2["lowest_store"] = ls

        if only_deals and not any((off.discount_pct or 0) > 0 for off in rec2["stores"].values() if off):
            continue

        _plain_item       = rec2["item_number"].split("|")[-1] if "|" in (rec2["item_number"] or "") else rec2["item_number"]
        rec2["alert"]       = _get_item_alert(_plain_item, stores)
        rec2["alert_types"] = _get_item_alert_types(_plain_item, stores)

        if only_alerts:
            if not rec2["alert_types"]:
                continue
            if alert_type_filter != "all" and alert_type_filter not in rec2["alert_types"]:
                continue
        if only_comparable and not all(rec2["stores"].get(s) and rec2["stores"][s].price is not None for s in stores):
            continue
        if only_instock and not any("in stock" in (rec2["stores"].get(s) and rec2["stores"][s].availability or "").lower() for s in stores):
            continue

        filtered.append(rec2)

    def sort_key(r):
        if sort == "brand":    return (r.get("brand") or "").upper()
        if sort == "item":     return int(r["item_number"]) if str(r["item_number"]).isdigit() else 10**9
        if sort == "title":    return (r.get("title") or "").lower()
        if sort == "theme":    return (r.get("theme") or "").lower()
        if sort == "category": return (r.get("category") or "").lower()
        if sort == "price":    return r.get("lowest_price") if r.get("lowest_price") is not None else 10**12
        if sort.startswith("store:"):
            off = (r.get("stores") or {}).get(sort.split(":", 1)[1])
            return off.price if (off and off.price is not None) else 10**12
        return r.get("item_number") or ""

    filtered.sort(key=sort_key, reverse=(order == "desc"))

    total       = len(filtered)
    total_pages = max(1, (total + per_page - 1) // per_page)
    page        = min(page, total_pages)
    page_rows   = filtered[(page - 1) * per_page : page * per_page]

    return templates.TemplateResponse(TEMPLATE_FILE, {
        "request": request,
        "rows": page_rows, "total": total,
        "page": page, "per_page": per_page, "total_pages": total_pages,
        "sort": sort, "order": order,
        "stores": stores, "store_names": store_names,
        "all_categories": all_categories, "all_themes": all_themes, "all_brands": all_brands,
        "selected_category": selected_category, "selected_theme": selected_theme, "selected_brand": selected_brand,
        "search_item": search_item, "compare": "all",
        "selected_stores": selected_stores, "raw_selected_stores": raw_selected_stores,
        "only_deals": only_deals, "only_alerts": only_alerts,
        "only_comparable": only_comparable, "only_instock": only_instock,
        "alert_type_filter": alert_type_filter,
        "last_updated": meta_get(DB_PATH, "last_updated") or "never",
        "alerts_unread": alerts_unread_count(DB_PATH),
        "price_range": "all", "min_price": "", "max_price": "",
        "source_tier": source_tier,
    })


# ── Official Brands Catalog ────────────────────────────────────────────────────

@app.get("/official", response_class=HTMLResponse)
def official_catalog(request: Request):
    qp            = request.query_params
    search_q      = qp.get("q", "").strip().lower()
    selected_brand = qp.get("brand", "All")
    selected_theme = qp.get("theme", "All")
    avail_filter  = qp.get("avail", "all")
    sort          = qp.get("sort", "title")
    per_page      = int(qp.get("per_page", "96") or "96")
    page          = max(1, int(qp.get("page", "1") or "1"))

    # Fetch latest snapshots for official stores only
    conn = db_connect(DB_PATH)
    cur  = conn.cursor()
    cur.execute("""
        SELECT s.store, s.item_number, s.title, s.brand, s.theme, s.category,
               s.price, s.compare_at, s.availability, s.link, s.image_url
        FROM snapshots s
        INNER JOIN (
            SELECT store, item_number, MAX(id) AS max_id
            FROM snapshots WHERE source_type='official'
            GROUP BY store, item_number
        ) t ON s.id = t.max_id
    """)
    rows = cur.fetchall()
    conn.close()

    # Build product list
    STORE_TO_BRAND = {
        "CaDA Official":  "CADA",
        "Mould King":     "MOULD KING",
        "LEGO Official":  "LEGO",
    }
    products = []
    for r in rows:
        brand = STORE_TO_BRAND.get(r["store"], (r["store"] or "").upper())
        theme = (r["theme"] or r["category"] or "").strip()
        title = (r["title"] or "").strip()
        avail = (r["availability"] or "").lower()

        products.append({
            "item_number": r["item_number"] or "",
            "title":       title,
            "brand":       brand,
            "theme":       theme,
            "category":    r["category"] or "",
            "image_url":   r["image_url"] or "",
            "stores": {
                r["store"]: StoreOffer(
                    price=r["price"],
                    availability=r["availability"] or "N/A",
                    link=r["link"] or "",
                    discount_pct=compute_discount_pct(r["compare_at"], r["price"]),
                )
            },
        })

    # Build filter lists BEFORE filtering (so dropdowns always show all options)
    all_brands = ["All"] + sorted({p["brand"] for p in products if p["brand"]})
    all_themes = ["All"] + sorted({p["theme"] for p in products if p["theme"]})

    # Filters
    if search_q:
        products = [p for p in products if search_q in p["title"].lower() or search_q in p["item_number"].lower()]
    if selected_brand != "All":
        products = [p for p in products if p["brand"] == selected_brand.upper()]
    if selected_theme != "All":
        products = [p for p in products if p["theme"] == selected_theme]
    if avail_filter == "instock":
        products = [p for p in products if any("in" in (o.availability or "").lower() for o in p["stores"].values())]
    elif avail_filter == "outstock":
        products = [p for p in products if any("out" in (o.availability or "").lower() for o in p["stores"].values())]

    # Sort
    if sort == "price_asc":
        products.sort(key=lambda p: min((o.price or 9999 for o in p["stores"].values()), default=9999))
    elif sort == "price_desc":
        products.sort(key=lambda p: min((o.price or 0 for o in p["stores"].values()), default=0), reverse=True)
    else:
        products.sort(key=lambda p: (p["brand"], p["title"]))

    # Paginate
    total       = len(products)
    total_pages = max(1, (total + per_page - 1) // per_page)
    page        = min(page, total_pages)
    page_rows   = products[(page - 1) * per_page : page * per_page]

    return templates.TemplateResponse("official.html", {
        "request":        request,
        "rows":           page_rows,
        "total":          total,
        "page":           page,
        "per_page":       per_page,
        "total_pages":    total_pages,
        "sort":           sort,
        "search_q":       search_q,
        "selected_brand": selected_brand,
        "selected_theme": selected_theme,
        "avail_filter":   avail_filter,
        "all_brands":     all_brands,
        "all_themes":     all_themes,
        "last_updated":   meta_get(DB_PATH, "last_updated") or "never",
        "alerts_unread":  alerts_unread_count(DB_PATH),
    })


@app.get("/international", response_class=HTMLResponse)
def international_catalog(request: Request):
    from modules.brickradar.config import INTERNATIONAL_STORES, INTERNATIONAL_COUNTRIES
    qp               = request.query_params
    search_q         = qp.get("q", "").strip().lower()
    selected_source  = qp.get("source", "All")
    selected_sub_tier = qp.get("sub_tier", "All")
    selected_country = qp.get("country", "All")
    selected_theme   = qp.get("theme", "All")
    sort             = qp.get("sort", "title")
    per_page         = int(qp.get("per_page", "96") or "96")
    page             = max(1, int(qp.get("page", "1") or "1"))

    # Fetch latest snapshots for international stores only
    conn = db_connect(DB_PATH)
    cur  = conn.cursor()
    cur.execute("""
        SELECT s.store, s.item_number, s.title, s.brand, s.theme,
               s.price, s.compare_at, s.availability, s.link, s.image_url
        FROM snapshots s
        INNER JOIN (
            SELECT store, item_number, MAX(id) AS max_id
            FROM snapshots WHERE source_type='international'
            GROUP BY store, item_number
        ) t ON s.id = t.max_id
    """)
    rows = cur.fetchall()
    conn.close()

    from collections import defaultdict
    items = defaultdict(lambda: {"item_number": "", "title": "", "theme": "", "image_url": "", "stores": {}})
    for r in rows:
        key = r["item_number"] or r["title"]
        it  = items[key]
        it["item_number"] = r["item_number"] or ""
        it["title"]       = (r["title"] or "").strip()
        it["theme"]       = (r["theme"] or "").strip()
        it["image_url"]   = r["image_url"] or it["image_url"]
        it["stores"][r["store"]] = StoreOffer(
            price=r["price"],
            availability=r["availability"] or "N/A",
            link=r["link"] or "",
            discount_pct=compute_discount_pct(r["compare_at"], r["price"]),
        )

    products = list(items.values())

    # Build store->sub_tier, country, currency from DB
    _conn2 = db_connect(DB_PATH)
    _cur2  = _conn2.cursor()
    _cur2.execute("SELECT name, sub_tier, country_code, currency FROM stores WHERE source_type='international'")
    store_sub_tier = {}
    store_country  = {}
    store_currency = {}
    for _r in _cur2.fetchall():
        store_sub_tier[_r["name"]] = _r["sub_tier"] or "regional"
        store_country[_r["name"]]  = _r["country_code"] or "AE"
        store_currency[_r["name"]] = _r["currency"] or "USD"
    _conn2.close()

    # Build dropdowns BEFORE filtering so they never collapse
    all_themes_pre    = ["All"] + sorted({p["theme"] for p in products if p["theme"]})
    all_sources_pre   = sorted({s for p in products for s in p["stores"]})
    all_countries_pre = {k: v for k, v in INTERNATIONAL_COUNTRIES.items()
                         if k in {store_country.get(s) for p in products for s in p["stores"]}}

    # Filters
    if search_q:
        products = [p for p in products if search_q in p["title"].lower() or search_q in p["item_number"].lower()]
    if selected_source != "All":
        products = [p for p in products if selected_source in p["stores"]]
    if selected_sub_tier != "All":
        products = [p for p in products if any(store_sub_tier.get(s, "regional") == selected_sub_tier for s in p["stores"])]
    if selected_country != "All":
        products = [p for p in products if any(store_country.get(s, "") == selected_country for s in p["stores"])]
    if selected_theme != "All":
        products = [p for p in products if p["theme"] == selected_theme]

    # Sort
    if sort == "price_asc":
        products.sort(key=lambda p: min((o.price or 9999 for o in p["stores"].values()), default=9999))
    elif sort == "price_desc":
        products.sort(key=lambda p: min((o.price or 0 for o in p["stores"].values()), default=0), reverse=True)
    else:
        products.sort(key=lambda p: p["title"].lower())

    all_sources   = all_sources_pre
    all_themes    = all_themes_pre
    all_countries = all_countries_pre
    intl_stores   = all_sources

    total       = len(products)
    total_pages = max(1, (total + per_page - 1) // per_page)
    page        = min(page, total_pages)
    page_rows   = products[(page - 1) * per_page : page * per_page]

    return templates.TemplateResponse("international.html", {
        "request":            request,
        "rows":               page_rows,
        "total":              total,
        "page":               page,
        "per_page":           per_page,
        "total_pages":        total_pages,
        "sort":               sort,
        "search_q":           search_q,
        "selected_source":    selected_source,
        "selected_sub_tier":  selected_sub_tier,
        "selected_country":   selected_country,
        "selected_theme":     selected_theme,
        "all_sources":        all_sources,
        "all_themes":         all_themes,
        "all_countries":      all_countries,
        "intl_stores":        intl_stores,
        "store_currency":     store_currency,
        "last_updated":       meta_get(DB_PATH, "last_updated") or "never",
        "alerts_unread":      alerts_unread_count(DB_PATH),
    })



# ── Refresh / SSE ──────────────────────────────────────────────────────────────

@app.post("/api/refresh")
def api_refresh():
    import queue as _queue
    results = {}
    q = _queue.Queue()

    def _progress(msg):
        q.put(msg)

    import threading as _t
    done_ev = _t.Event()

    def _run():
        _core_refresh_all(
            db_path=DB_PATH,
            shopify_stores={k: {**v, "normalize_theme_fn": normalize_theme_category_from_shopify} for k, v in SHOPIFY_STORES.items()},
            bigcommerce_stores=BIGCOMMERCE_STORES,
            html_stores=[],
            fetch_shopify_fn=fetch_shopify_store,
            fetch_bigcommerce_fn=fetch_bigcommerce_store,
            fetch_html_stores_fn=fetch_html_stores,
            progress_fn=_progress,
        )
        done_ev.set()

    _t.Thread(target=_run, daemon=True).start()
    done_ev.wait(timeout=600)
    return JSONResponse({"ok": True})


@app.post("/api/refresh/local")
def api_refresh_local():
    import threading
    def _run():
        from core.engine import refresh_all
        refresh_all(
            db_path=DB_PATH,
            shopify_stores={k: {**v, "normalize_theme_fn": normalize_theme_category_from_shopify} for k, v in SHOPIFY_STORES.items()},
            bigcommerce_stores=BIGCOMMERCE_STORES,
            fetch_html_stores_fn=fetch_html_stores,
            source_type_filter="local",
        )
    threading.Thread(target=_run, daemon=True).start()
    return JSONResponse({"ok": True, "message": "Local refresh started"})


@app.post("/api/refresh/official")
async def api_refresh_official(request: Request):
    import threading
    from datetime import datetime as _dt
    body       = {}
    try: body  = await request.json()
    except: pass
    store_filter = body.get("store", None)  # None = all official stores

    def _run():
        try:
            from modules.brickradar.scrapers_official import fetch_official_stores
            from core.db import persist_snapshot
            captured_at = _dt.utcnow().isoformat()

            # Load country codes from DB for persist
            conn = db_connect(DB_PATH)
            cur  = conn.cursor()
            cur.execute("SELECT name, country_code FROM stores WHERE source_type='official' AND enabled=1")
            country_map = {r["name"]: r["country_code"] for r in cur.fetchall()}
            conn.close()

            catalogs_by_store = {}
            def _progress(msg): print(f"[official] {msg}")

            # fetch_official_stores returns list of catalogs — need store names too
            # Use per-store approach so we can map country_code correctly
            import sqlite3 as _sq
            conn2 = _sq.connect(DB_PATH)
            conn2.row_factory = _sq.Row
            cur2  = conn2.cursor()
            cur2.execute("SELECT name, base_url, platform, collection_slug, vat_multiplier, lego_only FROM stores WHERE source_type='official' AND enabled=1")
            db_stores = [dict(r) for r in cur2.fetchall()]
            conn2.close()

            from modules.brickradar.scrapers_official import fetch_store_by_platform
            for s in db_stores:
                name = s["name"]
                if store_filter and name != store_filter:
                    continue
                try:
                    catalog = fetch_store_by_platform(
                        name=name,
                        base_url=s["base_url"],
                        platform=s["platform"],
                        collection_slug=s.get("collection_slug") or "",
                        db_path=DB_PATH,
                        vat_multiplier=float(s.get("vat_multiplier") or 1.0),
                        lego_only=bool(s.get("lego_only")),
                    )
                    if catalog:
                        persist_snapshot(DB_PATH, captured_at, name, catalog,
                                         source_type="official",
                                         country_code=country_map.get(name, "CN"))
                        print(f"[{name}] refreshed {len(catalog)} products")
                except Exception as e:
                    print(f"[{name}] ERROR: {e}")
        except Exception as e:
            print(f"[official refresh] ERROR: {e}")

    threading.Thread(target=_run, daemon=True).start()
    store_msg = store_filter or "all official stores"
    return JSONResponse({"ok": True, "message": f"Official refresh started for {store_msg}"})


@app.post("/api/refresh/international")
async def api_refresh_international(request: Request):
    import threading
    from datetime import datetime as _dt
    body = {}
    try: body = await request.json()
    except: pass
    store_filter = body.get("store", None)

    def _run():
        try:
            from modules.brickradar.scrapers_official import fetch_store_by_platform
            from core.db import persist_snapshot
            import sqlite3 as _sq
            captured_at = _dt.utcnow().isoformat()

            conn = _sq.connect(DB_PATH)
            conn.row_factory = _sq.Row
            cur  = conn.cursor()
            cur.execute("SELECT name, base_url, platform, collection_slug, vat_multiplier, lego_only, country_code FROM stores WHERE source_type='international' AND enabled=1")
            db_stores = [dict(r) for r in cur.fetchall()]
            conn.close()

            for s in db_stores:
                name = s["name"]
                if store_filter and name != store_filter:
                    continue
                platform = (s.get("platform") or "").lower()
                if platform not in ("shopify", "woocommerce", "ueeshop"):
                    print(f"[{name}] platform '{platform}' not yet scrapable — skipping")
                    continue
                try:
                    catalog = fetch_store_by_platform(
                        name=name,
                        base_url=s["base_url"],
                        platform=platform,
                        collection_slug=s.get("collection_slug") or "",
                        db_path=DB_PATH,
                        vat_multiplier=float(s.get("vat_multiplier") or 1.0),
                        lego_only=bool(s.get("lego_only")),
                    )
                    if catalog:
                        persist_snapshot(DB_PATH, captured_at, name, catalog,
                                         source_type="international",
                                         country_code=s.get("country_code", "AE"))
                        print(f"[{name}] refreshed {len(catalog)} products")
                except Exception as e:
                    print(f"[{name}] ERROR: {e}")
        except Exception as e:
            print(f"[international refresh] ERROR: {e}")

    threading.Thread(target=_run, daemon=True).start()
    store_msg = store_filter or "all international stores"
    return JSONResponse({"ok": True, "message": f"International refresh started for {store_msg}"})


@app.get("/api/refresh/stream")
def api_refresh_stream(stores: str = ""):
    selected = [s.strip() for s in stores.split(",") if s.strip()] if stores else []
    import queue as _queue

    SEQUENTIAL_AFTER = {"Brix & Figures", "Thetoystorelb"}

    def event_stream():
        captured_at = utc_now_iso()
        tasks = []
        for sname, cfg in SHOPIFY_STORES.items():
            tasks.append(("shopify", sname, cfg))

        conn = db_connect(DB_PATH)
        cur  = conn.cursor()
        cur.execute("SELECT name, base_url, platform, vat_multiplier, new_arrivals_collection, collection_slug, lego_only FROM stores WHERE enabled=1")
        for ds in cur.fetchall():
            sname = ds["name"]
            if sname in SHOPIFY_STORES:
                continue
            if ds["platform"] == "shopify":
                tasks.append(("shopify_db", sname, {
                    "url": ds["base_url"], "vat_multiplier": ds["vat_multiplier"],
                    "new_arrivals_collection": ds["new_arrivals_collection"],
                    "collection_slug": ds["collection_slug"],
                    "lego_only": bool(ds["lego_only"]),
                }))
        conn.close()

        tasks.append(("brickshop", None, None))
        tasks.append(("playone",   None, None))
        for bc_name, bc_cfg in BIGCOMMERCE_STORES.items():
            tasks.append(("bigcommerce", bc_name, bc_cfg))
        for off_name, off_cfg in OFFICIAL_STORES.items():
            tasks.append(("official", off_name, off_cfg))

        if selected:
            def _label(t): return t[1] or ("BRICKSHOP" if t[0] == "brickshop" else ("PlayOne" if t[0] == "playone" else t[0]))
            tasks = [t for t in tasks if _label(t) in selected]

        total = len(tasks)
        q     = _queue.Queue()

        def _run_task(task, stagger_idx: int = 0):
            if stagger_idx > 0:
                time.sleep(stagger_idx * 5)
            kind, sname, cfg = task
            label = sname or kind
            try:
                if kind in ("shopify", "shopify_db"):
                    base_url = cfg.get("url") or cfg.get("base_url", "")
                    new_items = fetch_new_arrival_items(label, base_url, cfg.get("new_arrivals_collection") or "")
                    result = fetch_shopify_store(
                        store_name=label, base_url=base_url,
                        vat_multiplier=float(cfg.get("vat_multiplier", 1.0)),
                        new_items=new_items,
                        collection_slug=cfg.get("collection_slug") or "",
                        lego_only=cfg.get("lego_only", False),
                        normalize_theme_fn=normalize_theme_category_from_shopify,
                    )
                    q.put(("ok", label, result))
                elif kind == "brickshop":
                    q.put(("ok", "BRICKSHOP", fetch_brickshop()))
                elif kind == "bigcommerce":
                    q.put(("ok", label, fetch_bigcommerce_store(
                        label, cfg["url"], cfg.get("collection_slug", ""),
                        cfg.get("lego_only", True), float(cfg.get("vat_multiplier", 1.0))
                    )))
                elif kind == "official":
                    _fn = {"CaDA Official": fetch_cada, "Mould King": lambda: fetch_mouldking(DB_PATH), "LEGO Official": fetch_lego_com}.get(label)
                    if _fn:
                        q.put(("ok", label, _fn()))
                    else:
                        q.put(("err", label, "No scraper for official store"))
                else:
                    q.put(("ok", "PlayOne", fetch_playone()))
            except Exception as e:
                q.put(("err", label, str(e)))

        yield f"data: {json.dumps({'type': 'start', 'total': total})}\n\n"

        # Keepalive ping
        _stop_ping = threading.Event()
        def _ping():
            while not _stop_ping.wait(20):
                q.put(("ping", None, None))
        threading.Thread(target=_ping, daemon=True).start()

        # Background persist
        _sse_q    = _queue.Queue()
        _persist_done = threading.Event()

        def _background_persist():
            done_bg = 0
            while done_bg < total:
                try:
                    status, name, data = q.get(timeout=600)
                except Exception:
                    print(f"[PERSIST] queue timeout after {done_bg}/{total}")
                    break
                if status == "ping":
                    _sse_q.put(("ping", None, None)); continue
                done_bg += 1
                if status == "ok":
                    count = len(data) if isinstance(data, dict) else 0
                    # Determine source_type and country_code for this store
                    st_cfg = (
                        SHOPIFY_STORES.get(name) or
                        BIGCOMMERCE_STORES.get(name) or
                        OFFICIAL_STORES.get(name) or
                        INTERNATIONAL_STORES.get(name) or {}
                    )
                    src_type     = st_cfg.get("source_type", "local")
                    country_code = st_cfg.get("country_code", "LB")
                    if name in ("BRICKSHOP", "PlayOne"):
                        src_type, country_code = "local", "LB"
                    persist_snapshot(DB_PATH, captured_at, name, data, src_type, country_code)
                    compute_alerts(DB_PATH, captured_at, name, data)
                    _sse_q.put(("ok", name, count, done_bg))
                else:
                    _sse_q.put(("err", name, data, done_bg))
            meta_set(DB_PATH, "last_updated", captured_at)
            _sse_q.put(("done", None, None, total))
            _persist_done.set()

        threading.Thread(target=_background_persist, daemon=True).start()

        # Submit tasks
        main_tasks = [t for t in tasks if (t[1] or t[0]) not in SEQUENTIAL_AFTER]
        late_tasks = [t for t in tasks if (t[1] or t[0]) in SEQUENTIAL_AFTER]
        pool = ThreadPoolExecutor(max_workers=len(main_tasks) + 1)
        shopify_idx = 0
        for t in main_tasks:
            if t[0] in ("brickshop", "playone", "official"):
                pool.submit(_run_task, t, 0)
            else:
                pool.submit(_run_task, t, shopify_idx); shopify_idx += 1

        if late_tasks:
            def _run_late():
                time.sleep(shopify_idx * 5 + 10)
                for lt in late_tasks:
                    _run_task(lt, 0); time.sleep(10)
            pool.submit(_run_late)
        pool.shutdown(wait=False)

        while True:
            msg  = _sse_q.get()
            kind = msg[0]
            if kind == "ping":
                yield ": keepalive\n\n"
            elif kind == "ok":
                _, name, count, done_n = msg
                yield f"data: {json.dumps({'type': 'store', 'name': name, 'count': count, 'done': done_n, 'total': total})}\n\n"
            elif kind == "err":
                _, name, err, done_n = msg
                yield f"data: {json.dumps({'type': 'error', 'name': name, 'error': err, 'done': done_n, 'total': total})}\n\n"
            elif kind == "done":
                break

        _stop_ping.set()
        yield f"data: {json.dumps({'type': 'done'})}\n\n"

    return StreamingResponse(event_stream(), media_type="text/event-stream")


# ── Alerts ─────────────────────────────────────────────────────────────────────

@app.get("/api/compare/{item_number}")
def api_compare(item_number: str):
    conn = db_connect(DB_PATH)
    cur  = conn.cursor()
    cur.execute("""
        SELECT s.store, s.title, s.price, s.availability, s.link,
               s.image_url, s.source_type, s.country_code, st.currency
        FROM snapshots s
        LEFT JOIN stores st ON st.name = s.store
        INNER JOIN (SELECT store, MAX(id) as max_id FROM snapshots WHERE item_number=? GROUP BY store) t ON s.id=t.max_id
        WHERE s.item_number=? ORDER BY s.source_type, s.price ASC
    """, (item_number, item_number))
    rows = cur.fetchall()
    conn.close()
    title=""; image_url=""; results=[]
    for r in rows:
        if r["title"] and not title: title=r["title"]
        if r["image_url"] and not image_url: image_url=r["image_url"]
        results.append({"store":r["store"],"price":r["price"],"currency":r["currency"] or "USD",
                        "availability":r["availability"] or "","link":r["link"] or "",
                        "source_type":r["source_type"] or "local","country_code":r["country_code"] or ""})
    return JSONResponse({"item_number":item_number,"title":title,"image_url":image_url,"results":results})


@app.post("/api/alerts/mark_read")
def api_mark_alerts_read():
    alerts_mark_read(DB_PATH)
    return JSONResponse({"ok": True, "unread": alerts_unread_count(DB_PATH)})


# ── Stores page ────────────────────────────────────────────────────────────────

@app.get("/stores", response_class=HTMLResponse)
def stores_page():
    f = os.path.join(APP_DIR, "Brickradar", "app", "stores.html")
    return HTMLResponse(open(f, encoding="utf-8").read())


@app.get("/api/stores/logos")
def api_store_logos():
    """Return logo URLs for all stores."""
    static_dir = os.path.join(STATIC_DIR, "logos")
    os.makedirs(static_dir, exist_ok=True)
    logos = {}
    all_urls = dict(HARDCODED_STORE_URLS)
    for bc_name, bc_cfg in BIGCOMMERCE_STORES.items():
        all_urls[bc_name] = bc_cfg["url"]
    conn = db_connect(DB_PATH)
    cur  = conn.cursor()
    cur.execute("SELECT name, base_url FROM stores WHERE enabled=1")
    for r in cur.fetchall():
        all_urls[r["name"]] = r["base_url"]
    conn.close()
    for name in all_urls:
        safe_name = re.sub(r"[^a-zA-Z0-9_-]", "_", name)
        for ext in [".png", ".jpg", ".svg", ".ico", ".webp", ".avif"]:
            p = os.path.join(static_dir, f"{safe_name}{ext}")
            if os.path.exists(p):
                logos[name] = f"/static/logos/{safe_name}{ext}"
                break
    return JSONResponse(logos)


@app.post("/api/stores/fetch_logos")
def api_fetch_all_logos():
    """Trigger logo fetch for all stores in background."""
    import threading as _flt
    import urllib.parse

    def _fetch_logo(base_url: str, store_name: str):
        import httpx as _hx
        static_dir = os.path.join(STATIC_DIR, "logos")
        os.makedirs(static_dir, exist_ok=True)
        safe_name = re.sub(r"[^a-zA-Z0-9_-]", "_", store_name)
        for ext in [".png", ".jpg", ".svg", ".ico", ".webp", ".avif"]:
            if os.path.exists(os.path.join(static_dir, f"{safe_name}{ext}")):
                return
        try:
            with _hx.Client(timeout=10, follow_redirects=True, headers=HEADERS) as client:
                r = client.get(base_url.rstrip("/") + "/")
                if r.status_code != 200:
                    return
                from bs4 import BeautifulSoup as _BS
                soup = _BS(r.text, "lxml")
                logo_url = None
                for sel, attr in [('meta[property="og:image"]', "content"),
                                   ('link[rel*="apple-touch-icon"]', "href"),
                                   ('link[rel*="icon"]', "href")]:
                    el = soup.select_one(sel)
                    if el and el.get(attr):
                        logo_url = el[attr]; break
                if not logo_url:
                    logo_url = base_url.rstrip("/") + "/favicon.ico"
                if logo_url.startswith("//"):
                    logo_url = "https:" + logo_url
                elif not logo_url.startswith("http"):
                    logo_url = urllib.parse.urljoin(base_url, logo_url)
                lr = client.get(logo_url)
                if lr.status_code == 200 and len(lr.content) > 100:
                    ct = lr.headers.get("content-type", "")
                    ext = ".svg" if "svg" in ct else ".ico" if "ico" in ct or logo_url.endswith(".ico") else ".jpg" if "jpeg" in ct or "jpg" in ct else ".webp" if "webp" in ct else ".png"
                    with open(os.path.join(static_dir, f"{safe_name}{ext}"), "wb") as f:
                        f.write(lr.content)
        except Exception as e:
            print(f"[Logo] {store_name}: {e}")

    all_urls = dict(HARDCODED_STORE_URLS)
    for bc_name, bc_cfg in BIGCOMMERCE_STORES.items():
        all_urls[bc_name] = bc_cfg["url"]
    conn = db_connect(DB_PATH)
    cur  = conn.cursor()
    cur.execute("SELECT name, base_url FROM stores WHERE enabled=1")
    for r in cur.fetchall():
        all_urls[r["name"]] = r["base_url"]
    conn.close()
    for name, url in all_urls.items():
        _flt.Thread(target=_fetch_logo, args=(url, name), daemon=True).start()
    return JSONResponse({"ok": True, "message": "Fetching logos in background"})


@app.get("/api/stores")
def api_get_stores(request: Request):
    qp          = request.query_params
    source_type = qp.get("source_type", "")
    sub_tier    = qp.get("sub_tier", "")
    country_code = qp.get("country_code", "")

    conn = db_connect(DB_PATH)
    cur  = conn.cursor()
    cur.execute("SELECT store, COUNT(DISTINCT item_number) as cnt, MAX(captured_at) as last_seen FROM snapshots GROUP BY store")
    snap_counts = {r["store"]: {"count": r["cnt"], "last_seen": r["last_seen"]} for r in cur.fetchall()}
    cur.execute("SELECT * FROM stores ORDER BY name")
    db_rows = {r["name"]: dict(r) for r in cur.fetchall()}
    conn.close()

    all_stores = []
    for name, row in db_rows.items():
        snap = snap_counts.get(name, {})
        row["product_count"] = snap.get("count", 0)
        row["last_scraped"]  = snap.get("last_seen", row.get("last_scraped", ""))
        row["hardcoded"]     = False
        all_stores.append(row)

    # Filter by tier/country if requested
    if source_type:
        all_stores = [s for s in all_stores if (s.get("source_type") or "") == source_type]
    if sub_tier:
        all_stores = [s for s in all_stores if (s.get("sub_tier") or "") == sub_tier]
    if country_code:
        all_stores = [s for s in all_stores if (s.get("country_code") or "") == country_code]

    return JSONResponse(all_stores)


@app.post("/api/stores/test")
async def api_test_store(request: Request):
    import httpx as _httpx
    body            = await request.json()
    raw_url         = (body.get("url") or "").strip().rstrip("/")
    if not raw_url.startswith("http"):
        raw_url = "https://" + raw_url.lstrip("http://").lstrip("https://")
    url             = raw_url
    collection_slug = (body.get("collection_slug") or "").strip() or None
    result = {
        "url": url, "platform": None, "product_count": 0, "samples": [],
        "error": None, "collection_slug": collection_slug,
        "scrapable": False, "tier": None, "tier_cost": None,
        "block_type": None, "diagnosis": None,
    }

    def _products_url(base, page, slug=None):
        if slug: return f"{base}/collections/{slug}/products.json?limit=250&page={page}"
        return f"{base}/products.json?limit=250&page={page}"

    try:
        import re as _re
        # Step 1: fetch homepage to detect platform
        try:
            home = _httpx.get(url, timeout=12, follow_redirects=True, headers=HEADERS)
            html = home.text.lower()
            status = home.status_code
        except Exception as e:
            result["error"] = str(e)
            result["diagnosis"] = "Could not reach the store — check the URL or try again."
            result["tier"] = "unreachable"
            return JSONResponse(result)

        # Detect platform from HTML signatures
        if "cdn.shopify" in html:
            detected_platform = "shopify"
        elif "woocommerce" in html:
            detected_platform = "woocommerce"
        elif "bigcommerce" in html:
            detected_platform = "bigcommerce"
        elif "mage/" in html or "magento" in html:
            detected_platform = "magento"
        elif "wixsite" in html or "wix.com" in html:
            detected_platform = "wix"
        elif "squarespace" in html:
            detected_platform = "squarespace"
        elif "shopware" in html:
            detected_platform = "shopware"
        else:
            detected_platform = "unknown"

        # Detect block type
        block_type = None
        if status == 403:
            if "cloudflare" in html or "__cf_" in home.headers.get("server","").lower():
                block_type = "cloudflare"
            else:
                block_type = "forbidden"
        elif status == 429:
            block_type = "rate_limited"
        elif "datadome" in html or "_dd_" in html:
            block_type = "datadome"
        elif "kasada" in html:
            block_type = "kasada"
        elif status != 200:
            block_type = f"http_{status}"

        result["platform"] = detected_platform

        # Step 2: Try Shopify products.json
        try:
            r = _httpx.get(_products_url(url, 1, collection_slug), timeout=12, follow_redirects=True, headers=HEADERS)
            if r.status_code == 200 and r.text.strip().startswith("{"):
                products = r.json().get("products") or []
                if products:
                    result["platform"] = "shopify"
                    result["scrapable"] = True
                    result["tier"] = "free"
                    result["tier_cost"] = 0
                    result["diagnosis"] = f"Shopify store detected — {len(products)} products visible via API. Ready to add."
                    page, total = 1, 0
                    with _httpx.Client(timeout=12, follow_redirects=True, headers=HEADERS) as client:
                        while True:
                            pr = client.get(_products_url(url, page, collection_slug))
                            if pr.status_code != 200: break
                            prods = pr.json().get("products") or []
                            if not prods: break
                            total += len(prods)
                            if len(prods) < 250 or page >= 5: break
                            page += 1
                    result["product_count"] = total
                    for p in products[:5]:
                        title = (p.get("title") or "").strip()
                        item_number = extract_item_number(title)
                        if not item_number:
                            for v in (p.get("variants") or []):
                                item_number = extract_item_number(v.get("sku") or "")
                                if item_number: break
                        variants = p.get("variants") or []
                        price = safe_float(variants[0].get("price")) if variants else None
                        images = p.get("images") or []
                        result["samples"].append({"title": title, "item_number": item_number or "—", "price": price, "image": images[0].get("src","") if images else ""})
                    return JSONResponse(result)
        except Exception:
            pass

        # Step 3: Try WooCommerce
        try:
            woo = _httpx.get(f"{url}/wp-json/wc/store/v1/products?per_page=5", timeout=10, follow_redirects=True, headers=HEADERS)
            if woo.status_code == 200 and woo.text.strip().startswith("["):
                prods = woo.json()
                if prods:
                    result["platform"] = "woocommerce"
                    result["scrapable"] = True
                    result["tier"] = "free"
                    result["tier_cost"] = 0
                    result["product_count"] = len(prods)
                    result["diagnosis"] = f"WooCommerce store detected — API accessible. Ready to add."
                    for p in prods[:5]:
                        result["samples"].append({
                            "title": p.get("name",""), "item_number": "—",
                            "price": p.get("prices",{}).get("price",""), "image": ""
                        })
                    return JSONResponse(result)
        except Exception:
            pass

        # Step 4: Classify blocked stores by tier
        if detected_platform in ("magento", "shopware", "unknown") or block_type in ("cloudflare", "forbidden", "http_403"):
            if block_type in ("datadome", "kasada") or detected_platform in ("magento",) and block_type == "cloudflare":
                result["tier"] = "tier3"
                result["tier_cost"] = 5.0
                result["block_type"] = block_type or "advanced_antibot"
                result["diagnosis"] = f"Advanced anti-bot protection detected ({block_type or 'DataDome/Kasada'}). Requires premium scraping service (~$5/mo). Submit a request and we'll set it up for you."
            elif block_type == "cloudflare" or detected_platform in ("magento", "shopware"):
                result["tier"] = "tier2"
                result["tier_cost"] = 2.0
                result["block_type"] = block_type or detected_platform
                result["diagnosis"] = f"{detected_platform.title()} store with {'Cloudflare protection' if block_type=='cloudflare' else 'platform-level restrictions'}. Requires paid scraping proxy (~$2/mo). Submit a request and we'll set it up for you."
            elif detected_platform in ("wix", "squarespace"):
                result["tier"] = "tier4"
                result["tier_cost"] = None
                result["block_type"] = detected_platform
                result["diagnosis"] = f"{detected_platform.title()} stores do not expose product data via API. Unfortunately this store cannot be scraped."
            else:
                result["tier"] = "tier2"
                result["tier_cost"] = 2.0
                result["block_type"] = block_type or "blocked"
                result["diagnosis"] = f"Store is accessible (HTTP {status}) but no standard API found. Platform: {detected_platform}. May require a custom scraper (~$2/mo)."
        else:
            result["tier"] = "tier2"
            result["tier_cost"] = 2.0
            result["diagnosis"] = f"Platform detected: {detected_platform}. No accessible product API found. Submit a request for custom scraping support."

        result["error"] = result["diagnosis"]
        return JSONResponse(result)

    except Exception as e:
        result["error"] = str(e)
        result["diagnosis"] = "Unexpected error during diagnosis."
        result["tier"] = "unreachable"
        return JSONResponse(result)


@app.post("/api/stores/request")
async def api_store_request(request: Request):
    body = await request.json()
    conn = db_connect(DB_PATH)
    conn.execute("""
        INSERT INTO store_requests(url, platform, tier, tier_cost, block_type, name, email, notes)
        VALUES (?,?,?,?,?,?,?,?)
    """, (
        body.get("url",""), body.get("platform",""), body.get("tier",""),
        body.get("tier_cost"), body.get("block_type",""),
        body.get("name",""), body.get("email",""), body.get("notes","")
    ))
    conn.commit()
    conn.close()
    print(f"[Store Request] {body.get('email')} requested {body.get('url')} ({body.get('tier')})")
    return JSONResponse({"ok": True})


@app.post("/api/stores/add")
async def api_add_store(request: Request):
    body            = await request.json()
    name            = (body.get("name") or "").strip()
    url             = (body.get("url")  or "").strip().rstrip("/")
    platform        = body.get("platform") or "shopify"
    vat             = float(body.get("vat_multiplier") or 1.0)
    new_arrivals    = (body.get("new_arrivals_collection") or "").strip() or None
    collection_slug = (body.get("collection_slug") or "").strip() or None
    lego_only       = int(bool(body.get("lego_only", False)))
    source_type     = (body.get("source_type") or "local").strip()
    sub_tier        = (body.get("sub_tier") or source_type).strip()
    country_code    = (body.get("country_code") or "LB").strip()
    currency        = (body.get("currency") or "USD").strip()
    if not name or not url:
        return JSONResponse({"ok": False, "error": "Name and URL required"}, status_code=400)
    conn = db_connect(DB_PATH)
    cur  = conn.cursor()
    try:
        cur.execute("""
            INSERT INTO stores(name, base_url, platform, vat_multiplier, new_arrivals_collection,
                               collection_slug, lego_only, enabled, source_type, sub_tier, country_code, currency)
            VALUES(?,?,?,?,?,?,?,1,?,?,?,?)
        """, (name, url, platform, vat, new_arrivals, collection_slug, lego_only,
              source_type, sub_tier, country_code, currency))
        conn.commit()
    except Exception as e:
        conn.close()
        return JSONResponse({"ok": False, "error": str(e)}, status_code=400)
    conn.close()
    return JSONResponse({"ok": True})


@app.post("/api/stores/toggle")
async def api_toggle_store(request: Request):
    body     = await request.json()
    store_id = body.get("id")
    conn     = db_connect(DB_PATH)
    conn.execute("UPDATE stores SET enabled = 1 - enabled WHERE id=?", (store_id,))
    conn.commit()
    cur = conn.execute("SELECT enabled FROM stores WHERE id=?", (store_id,))
    row = cur.fetchone()
    conn.close()
    return JSONResponse({"ok": True, "enabled": row["enabled"] if row else 0})


@app.post("/api/stores/delete")
async def api_delete_store(request: Request):
    body     = await request.json()
    store_id = body.get("id")
    conn     = db_connect(DB_PATH)
    conn.execute("DELETE FROM stores WHERE id=?", (store_id,))
    conn.commit()
    conn.close()
    return JSONResponse({"ok": True})


# ── Analytics ──────────────────────────────────────────────────────────────────

@app.get("/analytics/official", response_class=HTMLResponse)
def analytics_official_page():
    f = os.path.join(APP_DIR, "Brickradar", "app", "templates", "analytics_official.html")
    return HTMLResponse(open(f, encoding="utf-8").read() if os.path.exists(f) else "<h1>Not found</h1>")

@app.get("/analytics/international", response_class=HTMLResponse)
def analytics_international_page():
    f = os.path.join(APP_DIR, "Brickradar", "app", "templates", "analytics_international.html")
    return HTMLResponse(open(f, encoding="utf-8").read() if os.path.exists(f) else "<h1>Not found</h1>")

@app.get("/analytics", response_class=HTMLResponse)
def analytics_page():
    f = os.path.join(APP_DIR, "Brickradar", "app", "analytics.html")
    return HTMLResponse(open(f, encoding="utf-8").read())


@app.get("/api/analytics/kpis")
def api_analytics_kpis(source_type: str = "local"):
    conn = db_connect(DB_PATH)
    cur  = conn.cursor()
    lsf  = _lsf(source_type)
    cur.execute(f"SELECT COUNT(DISTINCT item_number) FROM snapshots WHERE 1=1 {lsf}")
    total_products = cur.fetchone()[0]
    cur.execute(f"SELECT COUNT(DISTINCT store) FROM snapshots WHERE 1=1 {lsf}")
    total_stores = cur.fetchone()[0]
    cur.execute(f"SELECT COUNT(*) FROM snapshots WHERE compare_at IS NOT NULL AND compare_at > price {lsf}")
    discounted = cur.fetchone()[0]
    cur.execute("SELECT COUNT(DISTINCT item_number) FROM alerts WHERE alert_type='new_arrival' AND created_at >= datetime('now','-7 days')")
    new_arrivals_7d = cur.fetchone()[0]
    cur.execute(f"SELECT AVG(price) FROM snapshots WHERE price IS NOT NULL {lsf}")
    avg_price = cur.fetchone()[0]
    cur.execute(f"SELECT COUNT(*) FROM snapshots WHERE availability NOT IN ('available','in_stock','') AND availability IS NOT NULL {lsf}")
    out_of_stock = cur.fetchone()[0]
    conn.close()
    return JSONResponse({"total_products": total_products, "total_stores": total_stores,
                         "discounted": discounted, "new_arrivals_7d": new_arrivals_7d,
                         "avg_price": round(avg_price, 2) if avg_price else 0, "out_of_stock": out_of_stock})


@app.get("/api/analytics/items_per_brand_store")
def api_items_per_brand_store(source_type: str = "local"):
    conn = db_connect(DB_PATH)
    cur  = conn.cursor()
    cur.execute(f"SELECT store, UPPER(TRIM(brand)) as brand, COUNT(DISTINCT item_number) as cnt FROM snapshots WHERE brand IS NOT NULL AND brand != '' {_lsf()} GROUP BY store, brand ORDER BY store, cnt DESC")
    rows = [{"store": r[0], "brand": r[1], "count": r[2]} for r in cur.fetchall()]
    conn.close()
    return JSONResponse(rows)


@app.get("/api/analytics/most_expensive_per_brand_store")
def api_most_expensive_per_brand_store(source_type: str = "local"):
    conn = db_connect(DB_PATH)
    cur  = conn.cursor()
    cur.execute(f"SELECT store, UPPER(TRIM(brand)) as brand, title, item_number, MAX(price) as price, link, image_url FROM snapshots WHERE price IS NOT NULL AND brand IS NOT NULL AND brand != '' {_lsf()} GROUP BY store, brand ORDER BY store, price DESC")
    rows = [{"store": r[0], "brand": r[1], "title": r[2], "item_number": r[3], "price": r[4], "link": r[5], "image": r[6]} for r in cur.fetchall()]
    conn.close()
    return JSONResponse(rows)


@app.get("/api/analytics/items_per_theme_store")
def api_items_per_theme_store(source_type: str = "local"):
    conn = db_connect(DB_PATH)
    cur  = conn.cursor()
    cur.execute(f"SELECT store, TRIM(theme) as theme, COUNT(DISTINCT item_number) as cnt FROM snapshots WHERE theme IS NOT NULL AND theme != '' {_lsf()} GROUP BY store, theme ORDER BY cnt DESC")
    rows = [{"store": r[0], "theme": r[1], "count": r[2]} for r in cur.fetchall()]
    conn.close()
    return JSONResponse(rows)


@app.get("/api/analytics/new_arrivals_per_store")
def api_new_arrivals_per_store():
    conn = db_connect(DB_PATH)
    cur  = conn.cursor()
    cur.execute("SELECT store, COUNT(DISTINCT item_number) as cnt FROM alerts WHERE alert_type='new_arrival' GROUP BY store ORDER BY cnt DESC")
    rows = [{"store": r[0], "count": r[1]} for r in cur.fetchall()]
    conn.close()
    return JSONResponse(rows)


@app.get("/api/analytics/discounts_per_brand_store")
def api_discounts_per_brand_store(source_type: str = "local"):
    conn = db_connect(DB_PATH)
    cur  = conn.cursor()
    cur.execute(f"SELECT store, UPPER(TRIM(brand)) as brand, COUNT(*) as cnt, AVG(ROUND((compare_at-price)/compare_at*100,1)) as avg_pct FROM snapshots WHERE compare_at IS NOT NULL AND compare_at > price AND brand IS NOT NULL AND brand != '' {_lsf()} GROUP BY store, brand ORDER BY cnt DESC")
    rows = [{"store": r[0], "brand": r[1], "count": r[2], "avg_discount_pct": round(r[3], 1) if r[3] else 0} for r in cur.fetchall()]
    conn.close()
    return JSONResponse(rows)


# ── Advanced analysis ──────────────────────────────────────────────────────────

@app.get("/advanced", response_class=HTMLResponse)
def advanced_page():
    f = os.path.join(APP_DIR, "Brickradar", "app", "advanced.html")
    return HTMLResponse(open(f, encoding="utf-8").read())


@app.get("/api/advanced/price_spread")
def api_price_spread():
    conn = db_connect(DB_PATH)
    cur  = conn.cursor()
    cur.execute(f"""
        SELECT item_number, title, brand, theme, category,
               COUNT(DISTINCT store) as store_count,
               MIN(price) as min_price, MAX(price) as max_price,
               ROUND(MAX(price)-MIN(price),2) as spread,
               ROUND((MAX(price)-MIN(price))/MIN(price)*100,1) as spread_pct,
               GROUP_CONCAT(store||':'||ROUND(price,2)||':'||availability,'|') as store_data
        FROM snapshots WHERE price IS NOT NULL AND price > 0 {_lsf()}
        GROUP BY item_number HAVING COUNT(DISTINCT store) >= 2 ORDER BY spread DESC
    """)
    rows = []
    for r in cur.fetchall():
        stores = {}
        for chunk in (r["store_data"] or "").split("|"):
            parts = chunk.split(":")
            if len(parts) >= 3:
                stores[parts[0]] = {"price": float(parts[1]), "availability": parts[2]}
        rows.append({"item_number": r["item_number"], "title": r["title"] or "", "brand": r["brand"] or "",
                     "theme": r["theme"] or "", "category": r["category"] or "",
                     "store_count": r["store_count"], "min_price": r["min_price"], "max_price": r["max_price"],
                     "spread": r["spread"], "spread_pct": r["spread_pct"], "stores": stores})
    conn.close()
    return JSONResponse(rows)


@app.get("/api/advanced/store_behavior")
def api_store_behavior():
    conn = db_connect(DB_PATH)
    cur  = conn.cursor()
    cur.execute(f"SELECT item_number FROM snapshots WHERE price IS NOT NULL AND price > 0 {_lsf()} GROUP BY item_number HAVING COUNT(DISTINCT store) >= 2")
    comparable_items = [r[0] for r in cur.fetchall()]
    if not comparable_items:
        conn.close()
        return JSONResponse([])
    ph = ",".join("?" * len(comparable_items))
    cur.execute(f"""
        SELECT s.store, COUNT(*) as items_carried, AVG(s.price) as avg_price,
               SUM(CASE WHEN s.price=m.min_price THEN 1 ELSE 0 END) as cheapest_count,
               SUM(CASE WHEN s.price=m.max_price THEN 1 ELSE 0 END) as priciest_count,
               AVG(ROUND((s.price-m.min_price)/NULLIF(m.min_price,0)*100,1)) as avg_premium_pct
        FROM snapshots s
        JOIN (SELECT item_number, MIN(price) as min_price, MAX(price) as max_price
              FROM snapshots WHERE price IS NOT NULL AND price > 0 AND item_number IN ({ph}) GROUP BY item_number) m
        ON s.item_number=m.item_number
        WHERE s.price IS NOT NULL AND s.price > 0 AND s.item_number IN ({ph})
        GROUP BY s.store ORDER BY avg_premium_pct ASC
    """, comparable_items + comparable_items)
    rows = []
    for r in cur.fetchall():
        total = r["items_carried"] or 1
        rows.append({"store": r["store"], "items_carried": r["items_carried"],
                     "avg_price": round(r["avg_price"] or 0, 2),
                     "cheapest_count": r["cheapest_count"] or 0,
                     "cheapest_pct": round((r["cheapest_count"] or 0) / total * 100, 1),
                     "priciest_count": r["priciest_count"] or 0,
                     "priciest_pct": round((r["priciest_count"] or 0) / total * 100, 1),
                     "avg_premium_pct": round(r["avg_premium_pct"] or 0, 1)})
    conn.close()
    return JSONResponse(rows)


@app.get("/api/advanced/deal_detector")
def api_deal_detector():
    conn = db_connect(DB_PATH)
    cur  = conn.cursor()
    cur.execute(f"""
        SELECT s.item_number, s.title, s.brand, s.theme, s.store,
               s.price, s.compare_at,
               ROUND((s.compare_at-s.price)/s.compare_at*100,1) as discount_pct,
               m.min_other_price, m.stores_full_price,
               ROUND(m.min_other_price-s.price,2) as saving_vs_others
        FROM snapshots s
        JOIN (SELECT item_number,
                     MIN(CASE WHEN (compare_at IS NULL OR compare_at<=price) THEN price END) as min_other_price,
                     COUNT(CASE WHEN (compare_at IS NULL OR compare_at<=price) THEN 1 END) as stores_full_price
              FROM snapshots WHERE price IS NOT NULL AND price > 0 {_lsf()} GROUP BY item_number) m
        ON s.item_number=m.item_number
        WHERE s.compare_at IS NOT NULL AND s.compare_at > s.price
          AND m.min_other_price IS NOT NULL AND m.min_other_price > s.price
          AND s.price IS NOT NULL {_lsf()}
        ORDER BY saving_vs_others DESC
    """)
    rows = [{"item_number": r["item_number"], "title": r["title"] or "", "brand": r["brand"] or "",
             "theme": r["theme"] or "", "store": r["store"], "price": r["price"],
             "compare_at": r["compare_at"], "discount_pct": r["discount_pct"],
             "min_other_price": r["min_other_price"], "stores_full_price": r["stores_full_price"],
             "saving_vs_others": r["saving_vs_others"]} for r in cur.fetchall()]
    conn.close()
    return JSONResponse(rows)


@app.post("/api/advanced/export")
async def api_advanced_export(request: Request):
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    body = await request.json()
    tab  = body.get("tab", "data")
    rows = body.get("rows", [])
    if not rows:
        return JSONResponse({"error": "No data"})
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = {"spread": "Price Spread", "behavior": "Store Behavior", "deals": "Deal Detector"}.get(tab, tab)
    headers  = list(rows[0].keys())
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h.replace("_", " ").title())
        cell.fill = PatternFill("solid", fgColor="1E293B")
        cell.font = Font(bold=True, color="94A3B8", size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for ri, row in enumerate(rows, 2):
        for ci, key in enumerate(headers, 1):
            val = row.get(key, "")
            if isinstance(val, str):
                try: val = float(val)
                except: pass
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = Font(size=9)
    ws.freeze_panes = "A2"
    for ci, key in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(ci)].width = 18 if "title" in key else 12
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    return StreamingResponse(iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=lego_{tab}_{ts}.xlsx"})


# ── RadarList ──────────────────────────────────────────────────────────────────


import hashlib as _hashlib
import secrets as _secrets
from datetime import datetime as _dt, timedelta as _td

def _verify_admin_password(password: str) -> bool:
    if not ADMIN_PASSWORD_HASH or not ADMIN_PASSWORD_SALT:
        return password == "Kh@R1z_Br!ck#2026"
    h = _hashlib.sha256((ADMIN_PASSWORD_SALT + password).encode()).hexdigest()
    return h == ADMIN_PASSWORD_HASH

def _create_session() -> str:
    token = _secrets.token_hex(32)
    _admin_sessions[token] = _dt.utcnow() + _td(hours=ADMIN_SESSION_HOURS)
    return token

def _check_session(request: Request) -> bool:
    token = request.cookies.get("admin_session")
    if not token or token not in _admin_sessions: return False
    if _dt.utcnow() > _admin_sessions[token]:
        del _admin_sessions[token]; return False
    return True

def _send_otp(email: str) -> str:
    otp = str(_secrets.randbelow(900000) + 100000)
    _otp_store[email] = (otp, _dt.utcnow() + _td(minutes=5))
    if GMAIL_APP_PASSWORD:
        try:
            import smtplib, ssl
            from email.mime.text import MIMEText
            msg = MIMEText(f"RadarOS Admin OTP: {otp}\n\nValid 5 minutes.")
            msg["Subject"] = "RadarOS Admin Login"
            msg["From"] = ADMIN_EMAIL; msg["To"] = ADMIN_EMAIL
            with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=ssl.create_default_context()) as s:
                s.login(ADMIN_EMAIL, GMAIL_APP_PASSWORD)
                s.sendmail(ADMIN_EMAIL, ADMIN_EMAIL, msg.as_string())
        except Exception as e:
            print(f"[Admin] OTP email failed: {e}")
    print(f"[Admin] OTP: {otp}")
    return otp

def _verify_otp(email: str, otp: str) -> bool:
    if email not in _otp_store: return False
    stored, expiry = _otp_store[email]
    if _dt.utcnow() > expiry: del _otp_store[email]; return False
    if otp == stored: del _otp_store[email]; return True
    return False

def _open_tmpl(filename: str) -> str:
    """Open template from platform templates dir first, then BrickRadar."""
    platform_path = os.path.join(PLATFORM_TMPL_DIR, filename)
    brickradar_path = os.path.join(APP_DIR, "Brickradar", "app", "templates", filename)
    for p in [platform_path, brickradar_path]:
        if os.path.exists(p):
            return open(p, encoding="utf-8").read()
    return f"<h1>{filename} not found</h1>"

@app.get("/admin/login", response_class=HTMLResponse)
def admin_login_page(): return HTMLResponse(_open_tmpl("admin_login.html"))

@app.post("/admin/login")
async def admin_login(request: Request):
    from fastapi.responses import RedirectResponse
    body = await request.json()
    ip   = request.client.host
    now  = _dt.utcnow()
    attempts, last = _login_attempts.get(ip, (0, now))
    if attempts >= 5 and (now - last).seconds < 300:
        return JSONResponse({"ok": False, "error": "Too many attempts. Wait 5 minutes."}, status_code=429)
    if not _verify_admin_password(body.get("password", "")):
        _login_attempts[ip] = (attempts + 1, now)
        return JSONResponse({"ok": False, "error": "Invalid password."})
    _login_attempts.pop(ip, None)
    if GMAIL_APP_PASSWORD:
        _send_otp(ADMIN_EMAIL)
        return JSONResponse({"ok": True, "otp_required": True})
    token = _create_session()
    resp  = JSONResponse({"ok": True, "otp_required": False})
    resp.set_cookie("admin_session", token, httponly=True, max_age=ADMIN_SESSION_HOURS*3600)
    return resp

@app.post("/admin/verify-otp")
async def admin_verify_otp(request: Request):
    body = await request.json()
    if _verify_otp(ADMIN_EMAIL, body.get("otp", "")):
        token = _create_session()
        resp  = JSONResponse({"ok": True})
        resp.set_cookie("admin_session", token, httponly=True, max_age=ADMIN_SESSION_HOURS*3600)
        return resp
    return JSONResponse({"ok": False, "error": "Invalid or expired OTP."})

@app.get("/admin/logout")
def admin_logout():
    from fastapi.responses import RedirectResponse
    resp = RedirectResponse("/admin/login")
    resp.delete_cookie("admin_session")
    return resp

@app.get("/admin", response_class=HTMLResponse)
def admin_dashboard(request: Request):
    if not _check_session(request):
        from fastapi.responses import RedirectResponse
        return RedirectResponse("/admin/login")
    return HTMLResponse(_open_tmpl("admin.html"))

@app.get("/api/admin/stats")
def api_admin_stats(request: Request):
    if not _check_session(request): return JSONResponse({"error":"Unauthorized"}, status_code=401)
    conn = db_connect(DB_PATH); cur = conn.cursor()
    stats = {
        "total_snapshots":  cur.execute("SELECT COUNT(*) FROM snapshots").fetchone()[0],
        "total_stores":     cur.execute("SELECT COUNT(*) FROM stores WHERE enabled=1").fetchone()[0],
        "total_requests":   cur.execute("SELECT COUNT(*) FROM store_requests").fetchone()[0],
        "pending_requests": cur.execute("SELECT COUNT(*) FROM store_requests WHERE status='pending'").fetchone()[0],
        "db_size_mb":       round(os.path.getsize(DB_PATH)/1024/1024, 2),
        "snapshots_by_tier":{r[0]:r[1] for r in cur.execute("SELECT source_type,COUNT(*) FROM snapshots GROUP BY source_type").fetchall()},
    }
    conn.close(); return JSONResponse(stats)

@app.get("/api/admin/requests")
def api_admin_requests(request: Request, status: str = ""):
    if not _check_session(request): return JSONResponse({"error":"Unauthorized"}, status_code=401)
    conn = db_connect(DB_PATH)
    q = "SELECT * FROM store_requests" + (f" WHERE status='{status}'" if status else "") + " ORDER BY created_at DESC"
    rows = [dict(r) for r in conn.execute(q).fetchall()]
    conn.close(); return JSONResponse(rows)

@app.post("/api/admin/requests/{req_id}/status")
async def api_admin_update_request(req_id: int, request: Request):
    if not _check_session(request): return JSONResponse({"error":"Unauthorized"}, status_code=401)
    body = await request.json()
    conn = db_connect(DB_PATH)
    conn.execute("UPDATE store_requests SET status=?, admin_notes=? WHERE id=?",
                 (body.get("status","pending"), body.get("admin_notes",""), req_id))
    conn.commit(); conn.close(); return JSONResponse({"ok": True})

@app.get("/api/admin/pricing")
def api_admin_pricing(request: Request):
    if not _check_session(request): return JSONResponse({"error":"Unauthorized"}, status_code=401)
    conn = db_connect(DB_PATH)
    rows = [dict(r) for r in conn.execute("SELECT * FROM pricing_config ORDER BY tier").fetchall()]
    conn.close(); return JSONResponse(rows)

@app.post("/api/admin/pricing")
async def api_admin_update_pricing(request: Request):
    if not _check_session(request): return JSONResponse({"error":"Unauthorized"}, status_code=401)
    body = await request.json()
    conn = db_connect(DB_PATH)
    for item in body:
        conn.execute("UPDATE pricing_config SET price=?,description=? WHERE tier=?",
                     (item["price"], item.get("description",""), item["tier"]))
    conn.commit(); conn.close(); return JSONResponse({"ok": True})


@app.get("/api/admin/plans")
def api_admin_get_plans(request: Request):
    if not _check_session(request): return JSONResponse({"error":"Unauthorized"}, status_code=401)
    conn  = db_connect(DB_PATH)
    rows  = [dict(r) for r in conn.execute("SELECT * FROM plans ORDER BY price").fetchall()]
    conn.close(); return JSONResponse(rows)


@app.post("/api/admin/plans")
async def api_admin_create_plan(request: Request):
    if not _check_session(request): return JSONResponse({"error":"Unauthorized"}, status_code=401)
    body = await request.json()
    conn = db_connect(DB_PATH)
    conn.execute("""INSERT INTO plans(name,price,currency,store_limit,refresh_limit,tier_access,description)
                    VALUES(?,?,?,?,?,?,?)""",
                 (body["name"], body.get("price",0), body.get("currency","USD"),
                  body.get("store_limit",3), body.get("refresh_limit",1),
                  body.get("tier_access","local"), body.get("description","")))
    conn.commit(); conn.close()
    return JSONResponse({"ok":True})

@app.put("/api/admin/plans/{plan_id}")
async def api_admin_update_plan(plan_id: int, request: Request):
    if not _check_session(request): return JSONResponse({"error":"Unauthorized"}, status_code=401)
    body = await request.json()
    conn = db_connect(DB_PATH)
    conn.execute("""UPDATE plans SET name=?,price=?,currency=?,store_limit=?,refresh_limit=?,tier_access=?,description=?
                    WHERE id=?""",
                 (body["name"], body.get("price",0), body.get("currency","USD"),
                  body.get("store_limit",3), body.get("refresh_limit",1),
                  body.get("tier_access","local"), body.get("description",""), plan_id))
    conn.commit(); conn.close()
    return JSONResponse({"ok":True})

@app.post("/api/admin/plans/{plan_id}/status")
async def api_admin_toggle_plan(plan_id: int, request: Request):
    if not _check_session(request): return JSONResponse({"error":"Unauthorized"}, status_code=401)
    body = await request.json()
    conn = db_connect(DB_PATH)
    conn.execute("UPDATE plans SET is_active=? WHERE id=?", (body.get("is_active",1), plan_id))
    conn.commit(); conn.close()
    return JSONResponse({"ok":True})


# ── Public User Auth & Registration ──────────────────────────────────────────

@app.get("/api/geo")
async def api_geo(request: Request):
    """Detect user country from IP using ipapi.co"""
    import httpx as _httpx
    ip = request.client.host
    # Use forwarded IP if behind proxy
    forwarded = request.headers.get("X-Forwarded-For","").split(",")[0].strip()
    if forwarded: ip = forwarded
    # Skip localhost
    if ip in ("127.0.0.1","::1","localhost"):
        return JSONResponse({"country_code":"LB","country":"Lebanon","timezone":"Asia/Beirut","vpn_suspected":False,"ip":ip,"confidence":"low"})
    try:
        r = _httpx.get(f"https://ipapi.co/{ip}/json/", timeout=5)
        d = r.json()
        return JSONResponse({
            "country_code": d.get("country_code","XX"),
            "country":      d.get("country_name","Unknown"),
            "timezone":     d.get("timezone",""),
            "city":         d.get("city",""),
            "org":          d.get("org",""),
            "vpn_suspected": "hosting" in d.get("org","").lower() or "vpn" in d.get("org","").lower(),
            "ip":           ip,
            "confidence":   "high"
        })
    except Exception as e:
        return JSONResponse({"country_code":"XX","country":"Unknown","timezone":"","vpn_suspected":False,"ip":ip,"confidence":"low"})

@app.get("/api/modules")
def api_get_modules():
    conn = db_connect(DB_PATH)
    rows = [dict(r) for r in conn.execute("SELECT * FROM modules WHERE is_active=1 ORDER BY id").fetchall()]
    conn.close()
    return JSONResponse(rows)

@app.get("/api/country-plans/{country_code}/{module_slug}")
def api_country_plans(country_code: str, module_slug: str):
    """Get available plans for a country and module."""
    conn = db_connect(DB_PATH)
    # Try country first, fall back to XX (default)
    rows = conn.execute("""
        SELECT p.*, cp.is_available, cp.trial_days, cp.trial_stores, cp.notes
        FROM country_plans cp
        JOIN plans p ON cp.plan_id=p.id
        JOIN modules m ON cp.module_id=m.id
        WHERE cp.country_code=? AND m.slug=? AND p.is_active=1
        ORDER BY p.price
    """, (country_code.upper(), module_slug)).fetchall()
    if not rows:
        rows = conn.execute("""
            SELECT p.*, cp.is_available, cp.trial_days, cp.trial_stores, cp.notes
            FROM country_plans cp
            JOIN plans p ON cp.plan_id=p.id
            JOIN modules m ON cp.module_id=m.id
            WHERE cp.country_code='XX' AND m.slug=? AND p.is_active=1
            ORDER BY p.price
        """, (module_slug,)).fetchall()
    conn.close()
    return JSONResponse([dict(r) for r in rows])

@app.post("/api/register")
async def api_register(request: Request):
    import base64 as _b64, secrets as _sec
    body        = await request.json()
    name        = body.get("name","").strip()
    email       = body.get("email","").strip().lower()
    password    = body.get("password","").strip()
    country     = body.get("country_code","XX").upper()
    plan_id     = int(body.get("plan_id", 1))
    module_slug = body.get("module_slug","brickradar")

    if not name or not email or not password:
        return JSONResponse({"ok":False,"error":"Name, email and password are required"})
    if len(password) < 8:
        return JSONResponse({"ok":False,"error":"Password must be at least 8 characters"})

    conn = db_connect(DB_PATH)
    # Check email exists
    if conn.execute("SELECT id FROM users WHERE email=?", (email,)).fetchone():
        conn.close()
        return JSONResponse({"ok":False,"error":"Email already registered"})

    # Get module
    mod = conn.execute("SELECT id FROM modules WHERE slug=?", (module_slug,)).fetchone()
    if not mod:
        conn.close()
        return JSONResponse({"ok":False,"error":"Invalid module"})
    module_id = mod[0]

    # Hash password
    salt = _b64.b64encode(os.urandom(16)).decode()
    h    = _hashlib.sha256((salt+password).encode()).hexdigest()
    verify_token = _sec.token_hex(32)

    # Create user
    conn.execute("""INSERT INTO users(name,email,password_hash,password_salt,plan_id,country_code,email_verified,verify_token)
                    VALUES(?,?,?,?,?,?,0,?)""",
                 (name, email, h, salt, plan_id, country, verify_token))
    uid = conn.execute("SELECT last_insert_rowid()").fetchone()[0]

    # Get trial config
    cp = conn.execute("""SELECT trial_days, trial_stores FROM country_plans
                         WHERE country_code=? AND plan_id=? AND module_id=?""",
                      (country, plan_id, module_id)).fetchone()
    trial_ends = None
    if cp and cp[0]:
        from datetime import datetime as _dtt, timedelta as _tdd
        trial_ends = (_dtt.utcnow() + _tdd(days=cp[0])).isoformat()

    # Create subscription
    conn.execute("INSERT INTO subscriptions(user_id,plan_id) VALUES(?,?)", (uid, plan_id))
    conn.execute("""INSERT INTO user_modules(user_id,module_id,plan_id,trial_ends,setup_done)
                    VALUES(?,?,?,?,0)""", (uid, module_id, plan_id, trial_ends))
    conn.commit()
    conn.close()

    # Send welcome email
    if GMAIL_APP_PASSWORD:
        try:
            import smtplib, ssl
            from email.mime.multipart import MIMEMultipart
            from email.mime.text import MIMEText
            verify_url = f"http://localhost:8001/verify?token={verify_token}"
            msg = MIMEMultipart("alternative")
            msg["Subject"] = "Welcome to RadarOS — Verify your email"
            msg["From"]    = ADMIN_EMAIL
            msg["To"]      = email
            html_body = f"""
            <div style="font-family:sans-serif;max-width:500px;margin:0 auto;">
              <h2>Welcome to RadarOS, {name}!</h2>
              <p>Your account has been created successfully.</p>
              <p><strong>Module:</strong> {module_slug.title()}<br>
              <strong>Plan:</strong> Plan #{plan_id}</p>
              {"<p><strong>Trial:</strong> "+trial_ends[:10]+" end date</p>" if trial_ends else ""}
              <p><a href="{verify_url}" style="background:#6366f1;color:#fff;padding:10px 20px;border-radius:8px;text-decoration:none;display:inline-block;margin-top:1rem;">Verify Email & Setup Account</a></p>
              <p style="color:#94a3b8;font-size:0.85rem;">RadarOS — Market Intelligence Platform</p>
            </div>"""
            msg.attach(MIMEText(html_body, "html"))
            ctx = ssl.create_default_context()
            with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=ctx) as s:
                s.login(ADMIN_EMAIL, GMAIL_APP_PASSWORD)
                s.sendmail(ADMIN_EMAIL, email, msg.as_string())
            print(f"[Register] Welcome email sent to {email}")
        except Exception as e:
            print(f"[Register] Email failed: {e}")

    return JSONResponse({"ok":True,"user_id":uid,"verify_token":verify_token})

@app.get("/api/user/login")
async def api_user_login_get():
    return JSONResponse({"error":"Use POST"}, status_code=405)

@app.post("/api/user/login")
async def api_user_login(request: Request):
    import secrets as _sec
    body     = await request.json()
    email    = body.get("email","").strip().lower()
    password = body.get("password","").strip()
    conn     = db_connect(DB_PATH)
    user     = conn.execute("SELECT * FROM users WHERE email=?", (email,)).fetchone()
    conn.close()
    if not user:
        return JSONResponse({"ok":False,"error":"Invalid email or password"})
    h = _hashlib.sha256((user["password_salt"]+password).encode()).hexdigest()
    if h != user["password_hash"]:
        return JSONResponse({"ok":False,"error":"Invalid email or password"})
    if user["status"] == "suspended":
        return JSONResponse({"ok":False,"error":"Account suspended. Contact support."})
    # Create session
    token   = _sec.token_hex(32)
    expires = _dt.utcnow() + _td(hours=24)
    # Store in memory (will upgrade to DB later)
    _admin_sessions[f"user_{token}"] = {"user_id":user["id"],"expires":expires,"type":"user"}
    resp = JSONResponse({"ok":True,"user_id":user["id"],"setup_done":user["setup_done"],"name":user["name"]})
    resp.set_cookie("user_session", token, httponly=True, max_age=86400)
    return resp

@app.get("/api/user/me")
def api_user_me(request: Request):
    token = request.cookies.get("user_session")
    if not token or f"user_{token}" not in _admin_sessions:
        return JSONResponse({"error":"Not logged in"}, status_code=401)
    sess = _admin_sessions[f"user_{token}"]
    if _dt.utcnow() > sess["expires"]:
        return JSONResponse({"error":"Session expired"}, status_code=401)
    conn = db_connect(DB_PATH)
    user = dict(conn.execute("""
        SELECT u.*, p.name as plan_name, p.price as plan_price, p.features,
               m.name as module_name, m.slug as module_slug
        FROM users u
        LEFT JOIN plans p ON u.plan_id=p.id
        LEFT JOIN user_modules um ON um.user_id=u.id
        LEFT JOIN modules m ON um.module_id=m.id
        WHERE u.id=? ORDER BY um.created_at DESC LIMIT 1
    """, (sess["user_id"],)).fetchone() or {})
    conn.close()
    if not user: return JSONResponse({"error":"User not found"}, status_code=404)
    user.pop("password_hash",""); user.pop("password_salt",""); user.pop("verify_token","")
    return JSONResponse(user)

@app.get("/login", response_class=HTMLResponse)
def user_login_page():
    return HTMLResponse(_open_tmpl("login.html"))

@app.get("/register", response_class=HTMLResponse)
def user_register_page():
    return HTMLResponse(_open_tmpl("register.html"))

@app.get("/account", response_class=HTMLResponse)
def user_account_page():
    return HTMLResponse(_open_tmpl("account.html"))

@app.get("/setup", response_class=HTMLResponse)
def user_setup_page():
    return HTMLResponse(_open_tmpl("setup.html"))

@app.get("/radarlist", response_class=HTMLResponse)
def radarlist_page():
    f = os.path.join(APP_DIR, "Brickradar", "app", "radarlist.html")
    return HTMLResponse(open(f, encoding="utf-8").read())


@app.get("/api/radarlist")
def api_radarlist_get():
    conn = db_connect(DB_PATH)
    cur  = conn.cursor()
    cur.execute("SELECT * FROM radarlist ORDER BY added_at DESC")
    items = [dict(r) for r in cur.fetchall()]
    lsf   = _lsf()
    for item in items:
        cur.execute(f"SELECT store, price, compare_at, availability, link, image_url FROM snapshots WHERE item_number=? {lsf} ORDER BY price ASC", (item["item_number"],))
        prices = [dict(r) for r in cur.fetchall()]
        item["current_prices"] = prices
        item["min_price"]  = min((p["price"] for p in prices if p["price"]), default=None)
        item["min_store"]  = next((p["store"] for p in prices if p["price"] == item["min_price"]), None)
        item["image_url"]  = next((p["image_url"] for p in prices if p.get("image_url")), None)
        if item["added_price"] and item["min_price"]:
            item["price_drop"]     = round(item["added_price"] - item["min_price"], 2)
            item["price_drop_pct"] = round((item["added_price"] - item["min_price"]) / item["added_price"] * 100, 1)
        else:
            item["price_drop"] = 0
            item["price_drop_pct"] = 0
    conn.close()
    return JSONResponse(items)


@app.post("/api/radarlist/add")
async def api_radarlist_add(request: Request):
    body        = await request.json()
    item_number = body.get("item_number", "").strip()
    if not item_number:
        return JSONResponse({"ok": False, "error": "item_number required"})
    conn = db_connect(DB_PATH)
    cur  = conn.cursor()
    try:
        cur.execute(f"SELECT title, brand, theme, price, store FROM snapshots WHERE item_number=? {_lsf()} ORDER BY price ASC LIMIT 1", (item_number,))
        row   = cur.fetchone()
        title = body.get("title") or (row["title"] if row else "")
        brand = body.get("brand") or (row["brand"] if row else "")
        theme = body.get("theme") or (row["theme"] if row else "")
        price = row["price"] if row else None
        store = row["store"] if row else None
        cur.execute("INSERT INTO radarlist(item_number,title,brand,theme,added_price,added_store) VALUES(?,?,?,?,?,?) ON CONFLICT(item_number) DO NOTHING",
                    (item_number, title, brand, theme, price, store))
        conn.commit()
        added = cur.rowcount > 0
    except Exception as e:
        conn.close()
        return JSONResponse({"ok": False, "error": str(e)})
    conn.close()
    return JSONResponse({"ok": True, "added": added})


@app.post("/api/radarlist/remove")
async def api_radarlist_remove(request: Request):
    body        = await request.json()
    item_number = body.get("item_number", "").strip()
    conn        = db_connect(DB_PATH)
    conn.execute("DELETE FROM radarlist WHERE item_number=?", (item_number,))
    conn.commit()
    conn.close()
    return JSONResponse({"ok": True})


@app.get("/api/radarlist/ids")
def api_radarlist_ids():
    conn = db_connect(DB_PATH)
    cur  = conn.cursor()
    cur.execute("SELECT item_number FROM radarlist")
    ids  = [r[0] for r in cur.fetchall()]
    conn.close()
    return JSONResponse(ids)


# ── AI routes ──────────────────────────────────────────────────────────────────

@app.get("/api/ai/test")
def api_ai_test():
    result = {
        "anthropic_key_set":  bool(ANTHROPIC_API_KEY),
        "groq_key_set":       bool(GROQ_API_KEY),
        "active_provider":    "anthropic" if ANTHROPIC_API_KEY else ("groq" if GROQ_API_KEY else "none"),
    }
    try:
        ctx = build_context(DB_PATH)
        result["context_length"] = len(ctx)
        result["context_ok"]     = True
    except Exception as e:
        import traceback
        result["context_ok"]    = False
        result["context_error"] = str(e)
        result["traceback"]     = traceback.format_exc()
    return JSONResponse(result)


@app.post("/api/ai/chat")
async def api_ai_chat(request: Request):
    if not ANTHROPIC_API_KEY and not GROQ_API_KEY:
        return JSONResponse({"ok": False, "error": "No AI API key configured."}, status_code=500)
    body     = await request.json()
    messages = body.get("messages", [])
    page     = body.get("page", "dashboard")
    if not messages:
        return JSONResponse({"ok": False, "error": "No messages"}, status_code=400)
    return StreamingResponse(
        stream_chat(DB_PATH, messages, page, ANTHROPIC_API_KEY, GROQ_API_KEY, MODULE["name"]),
        media_type="text/event-stream",
    )


@app.post("/api/ai/analyze_store")
async def api_ai_analyze_store(request: Request):
    if not ANTHROPIC_API_KEY and not GROQ_API_KEY:
        return JSONResponse({"ok": False, "error": "No AI API key configured."})
    body   = await request.json()
    result = await analyze_store(
        url=body.get("url", ""), platform=body.get("platform", ""),
        product_count=body.get("product_count", 0), samples=body.get("samples", []),
        anthropic_api_key=ANTHROPIC_API_KEY, groq_api_key=GROQ_API_KEY,
    )
    return JSONResponse(result)


# ── DB admin routes ────────────────────────────────────────────────────────────

ADMIN_PIN = os.getenv("ADMIN_PIN", "1234")

@app.get("/api/db/stats")
def db_stats():
    try:
        size_mb = round(os.path.getsize(DB_PATH) / 1024 / 1024, 2)
    except Exception:
        size_mb = 0
    conn = db_connect(DB_PATH)
    cur  = conn.cursor()
    try:
        cur.execute("SELECT COUNT(*) FROM snapshots")
        total = cur.fetchone()[0]
        cur.execute("SELECT store, COUNT(*) as rows, MAX(captured_at) as last_refresh FROM snapshots GROUP BY store ORDER BY rows DESC")
        stores = [{"store": r[0], "rows": r[1], "last_refresh": r[2]} for r in cur.fetchall()]
        return JSONResponse({"db_size_mb": size_mb, "total_snapshots": total, "stores": stores})
    except Exception as e:
        return JSONResponse({"error": str(e)})
    finally:
        conn.close()


@app.post("/api/db/query")
async def db_custom_query(request: Request):
    body      = await request.json()
    sql       = (body.get("sql") or "").strip()
    pin       = (body.get("pin") or "").strip()
    if not sql:
        return JSONResponse({"error": "No query provided"})
    sql_upper = sql.upper().lstrip()
    is_write  = any(sql_upper.startswith(k) for k in ("UPDATE", "DELETE", "INSERT", "DROP", "ALTER", "CREATE"))
    if is_write:
        if pin != ADMIN_PIN:
            return JSONResponse({"error": "❌ Admin PIN required for write operations", "pin_required": True})
    elif not (sql_upper.startswith("SELECT") or sql_upper.startswith("WITH")):
        return JSONResponse({"error": "Only SELECT queries are allowed without admin PIN"})
    conn = db_connect(DB_PATH)
    cur  = conn.cursor()
    try:
        cur.execute(sql)
        conn.commit()
        columns  = [d[0] for d in cur.description] if cur.description else []
        rows     = [list(r) for r in cur.fetchmany(200)] if cur.description else []
        affected = cur.rowcount if is_write else None
        return JSONResponse({"columns": columns, "rows": rows, "affected": affected})
    except Exception as e:
        return JSONResponse({"error": str(e)})
    finally:
        conn.close()


# ── Export ─────────────────────────────────────────────────────────────────────

@app.get("/api/export")
def api_export(request: Request):
    """Export filtered dashboard data as CSV, Excel, or JSON."""
    import csv
    from collections import defaultdict
    from core.utils import order_stores as _order_stores

    qp               = dict(request.query_params)
    fmt              = qp.get("fmt", "csv").lower()
    store_names      = get_all_store_names()
    raw_stores       = request.query_params.getlist("stores")
    stores           = _order_stores([s for s in raw_stores if s in store_names]) or store_names
    selected_category = qp.get("category", "All")
    selected_theme    = qp.get("theme",    "All")
    selected_brand    = qp.get("brand",    "All")
    search_item       = qp.get("search_item", "").strip()
    only_deals        = qp.get("only_deals",   "0") == "1"
    only_instock      = qp.get("only_instock", "0") == "1"

    conn = db_connect(DB_PATH)
    cur  = conn.cursor()
    ph   = ",".join("?" * len(stores))
    cur.execute(f"""
        SELECT s.store, s.item_number, s.title, s.brand, s.theme, s.category,
               s.price, s.compare_at, s.availability, s.link
        FROM snapshots s
        INNER JOIN (SELECT store, MAX(captured_at) as max_at FROM snapshots GROUP BY store) latest
        ON s.store=latest.store AND s.captured_at=latest.max_at
        WHERE s.store IN ({ph}) ORDER BY s.item_number, s.store
    """, stores)
    db_rows = cur.fetchall()
    conn.close()

    items = defaultdict(lambda: {"item_number": "", "brand": "", "title": "", "theme": "", "category": "", "stores": {}})
    for r in db_rows:
        key = r["item_number"] or r["title"] or ""
        it  = items[key]
        it["item_number"] = r["item_number"] or ""
        it["brand"]    = (r["brand"] or "").strip().upper()
        it["title"]    = r["title"] or ""
        it["theme"]    = r["theme"] or ""
        it["category"] = r["category"] or ""
        disc_pct = 0
        if r["compare_at"] and r["price"] and r["compare_at"] > r["price"]:
            disc_pct = round((r["compare_at"] - r["price"]) / r["compare_at"] * 100, 1)
        it["stores"][r["store"]] = {"price": r["price"], "availability": r["availability"] or "", "discount_pct": disc_pct, "link": r["link"] or ""}

    rows_out = []
    for it in items.values():
        if selected_category != "All" and it["category"] != selected_category: continue
        if selected_theme     != "All" and it["theme"]    != selected_theme:    continue
        if selected_brand     != "All" and it["brand"]    != selected_brand.upper(): continue
        if search_item and search_item not in it["item_number"]: continue
        if only_deals   and not any((v["discount_pct"] or 0) > 0 for v in it["stores"].values()): continue
        if only_instock and not any(v["availability"] in ("available", "in_stock") for v in it["stores"].values()): continue
        prices = [v["price"] for v in it["stores"].values() if v["price"] is not None]
        lp     = min(prices) if prices else None
        row    = {"item_number": it["item_number"], "brand": it["brand"], "title": it["title"],
                  "theme": it["theme"], "category": it["category"],
                  "lowest_price": f"{lp:.2f}" if lp is not None else ""}
        for s in stores:
            off = it["stores"].get(s, {})
            row[f"{s}_price"]        = f"{off['price']:.2f}" if off.get("price") is not None else ""
            row[f"{s}_avail"]        = off.get("availability", "")
            row[f"{s}_discount_pct"] = f"{off['discount_pct']:.0f}" if off.get("discount_pct") else ""
            row[f"{s}_link"]         = off.get("link", "")
        rows_out.append(row)

    if not rows_out:
        return JSONResponse({"error": "No data for current filters"})

    headers_row = list(rows_out[0].keys())
    ts          = datetime.now().strftime("%Y%m%d_%H%M")

    if fmt == "csv":
        buf = io.StringIO()
        w   = csv.DictWriter(buf, fieldnames=headers_row)
        w.writeheader(); w.writerows(rows_out); buf.seek(0)
        return StreamingResponse(iter([buf.getvalue().encode("utf-8")]),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f"attachment; filename=lego_tracker_{ts}.csv"})

    if fmt == "xlsx":
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "LEGO Tracker"
        for ci, h in enumerate(headers_row, 1):
            cell = ws.cell(row=1, column=ci, value=h.replace("_", " ").title())
            cell.fill = PatternFill("solid", fgColor="1E293B")
            cell.font = Font(bold=True, color="94A3B8", size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        deal_fill = PatternFill("solid", fgColor="1A3020")
        for ri, row in enumerate(rows_out, 2):
            for ci, key in enumerate(headers_row, 1):
                val = row[key]
                if key.endswith("_price") or key == "lowest_price":
                    try: val = float(val)
                    except: pass
                elif key.endswith("_discount_pct"):
                    try: val = int(val)
                    except: pass
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.font = Font(size=9)
                if key == "lowest_price" and val:
                    cell.font = Font(size=9, bold=True, color="F59E0B")
                if key.endswith("_discount_pct") and val:
                    cell.fill = deal_fill
        ws.row_dimensions[1].height = 28
        ws.freeze_panes = "A2"
        col_widths = {"item_number": 14, "brand": 10, "title": 42, "theme": 18, "category": 16, "lowest_price": 13}
        for ci, key in enumerate(headers_row, 1):
            ws.column_dimensions[get_column_letter(ci)].width = col_widths.get(key, 16 if "link" in key else 11)
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        return StreamingResponse(iter([buf.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=lego_tracker_{ts}.xlsx"})

    return JSONResponse(rows_out)


@app.get("/export/pdf", response_class=HTMLResponse)
def export_pdf_page(request: Request):
    """Print-ready HTML page — open and Ctrl+P to save as PDF."""
    from collections import defaultdict
    from core.utils import order_stores as _order_stores

    qp               = dict(request.query_params)
    store_names      = get_all_store_names()
    raw_stores       = request.query_params.getlist("stores")
    stores           = _order_stores([s for s in raw_stores if s in store_names]) or store_names
    selected_category = qp.get("category", "All")
    selected_theme    = qp.get("theme",    "All")
    selected_brand    = qp.get("brand",    "All")
    search_item       = qp.get("search_item", "").strip()
    only_deals        = qp.get("only_deals",   "0") == "1"
    only_instock      = qp.get("only_instock", "0") == "1"

    conn = db_connect(DB_PATH)
    cur  = conn.cursor()
    ph   = ",".join("?" * len(stores))
    cur.execute(f"""
        SELECT s.store, s.item_number, s.title, s.brand, s.theme, s.category,
               s.price, s.compare_at, s.availability, s.link
        FROM snapshots s
        INNER JOIN (SELECT store, MAX(captured_at) as max_at FROM snapshots GROUP BY store) latest
        ON s.store=latest.store AND s.captured_at=latest.max_at
        WHERE s.store IN ({ph}) ORDER BY s.item_number, s.store
    """, stores)
    db_rows = cur.fetchall()
    conn.close()

    items = defaultdict(lambda: {"item_number": "", "brand": "", "title": "", "theme": "", "category": "", "stores": {}})
    for r in db_rows:
        key = r["item_number"] or r["title"] or ""
        it  = items[key]
        it["item_number"] = r["item_number"] or ""
        it["brand"]    = (r["brand"] or "").strip().upper()
        it["title"]    = r["title"] or ""
        it["theme"]    = r["theme"] or ""
        it["category"] = r["category"] or ""
        disc_pct = 0
        if r["compare_at"] and r["price"] and r["compare_at"] > r["price"]:
            disc_pct = round((r["compare_at"] - r["price"]) / r["compare_at"] * 100, 1)
        it["stores"][r["store"]] = {"price": r["price"], "discount_pct": disc_pct, "availability": r["availability"] or ""}

    rows_out = []
    for it in items.values():
        if selected_category != "All" and it["category"] != selected_category: continue
        if selected_theme     != "All" and it["theme"]    != selected_theme:    continue
        if selected_brand     != "All" and it["brand"]    != selected_brand.upper(): continue
        if search_item and search_item not in it["item_number"]: continue
        if only_deals   and not any((v["discount_pct"] or 0) > 0 for v in it["stores"].values()): continue
        if only_instock and not any(v["availability"] in ("available", "in_stock") for v in it["stores"].values()): continue
        prices = [v["price"] for v in it["stores"].values() if v["price"] is not None]
        it["lowest_price"] = min(prices) if prices else None
        rows_out.append(it)

    def store_cells(it):
        cells = ""
        for s in stores:
            off = it["stores"].get(s)
            if off and off["price"] is not None:
                disc   = f' <span style="color:#b45309">-{off["discount_pct"]:.0f}%</span>' if off["discount_pct"] else ""
                cells += f'<td>${off["price"]:.2f}{disc}</td>'
            else:
                cells += "<td>—</td>"
        return cells

    store_headers = "".join(f"<th>{s}</th>" for s in stores)
    ts            = datetime.now().strftime("%Y-%m-%d %H:%M")
    rows_html     = ""
    for it in rows_out:
        lp = f'<b>${it["lowest_price"]:.2f}</b>' if it["lowest_price"] else "—"
        rows_html += f"""<tr>
          <td>{it["item_number"]}</td><td>{it["brand"]}</td><td>{it["title"]}</td>
          <td>{it["theme"]}</td><td>{it["category"]}</td>
          {store_cells(it)}<td>{lp}</td></tr>"""

    filters_desc = " | ".join(filter(None, [
        f"Stores: {', '.join(stores)}" if stores != store_names else "",
        f"Category: {selected_category}" if selected_category != "All" else "",
        f"Theme: {selected_theme}" if selected_theme != "All" else "",
        f"Brand: {selected_brand}" if selected_brand != "All" else "",
        f"Search: {search_item}" if search_item else "",
        "Deals only" if only_deals else "",
        "In stock only" if only_instock else "",
    ])) or "All products"

    return f"""<!doctype html><html><head><meta charset="UTF-8">
<title>LEGO Tracker Export — {ts}</title>
<style>
  body{{font-family:Arial,sans-serif;font-size:8pt;color:#111;margin:0;padding:1cm}}
  h1{{font-size:13pt;margin:0 0 2px}}.meta{{font-size:7.5pt;color:#666;margin-bottom:8px}}
  table{{width:100%;border-collapse:collapse}}
  th{{background:#1e293b;color:#fff;padding:4px 6px;text-align:left;font-size:7.5pt}}
  td{{padding:3px 6px;border-bottom:1px solid #e5e7eb;vertical-align:top}}
  tr:nth-child(even) td{{background:#f9fafb}}
  @page{{margin:1.5cm;size:A4 landscape}}
  @media screen{{body{{max-width:1400px;margin:auto;padding:20px}}}}
  .print-btn{{margin-bottom:12px;padding:8px 18px;background:#4f46e5;color:#fff;border:none;border-radius:6px;cursor:pointer;font-size:10pt}}
  @media print{{.print-btn{{display:none}}}}
</style></head><body>
<button class="print-btn" onclick="window.print()">&#128196; Print / Save as PDF</button>
<h1>LEGO Tracker — Product Export</h1>
<div class="meta">Generated: {ts} &nbsp;|&nbsp; {len(rows_out)} products &nbsp;|&nbsp; {filters_desc}</div>
<table><thead><tr><th>Item #</th><th>Brand</th><th>Title</th><th>Theme</th><th>Category</th>{store_headers}<th>Lowest</th></tr></thead>
<tbody>{rows_html}</tbody></table>
</body></html>"""


@app.post("/api/ai/discover-stores")
async def api_discover_stores(request: Request):
    if not ANTHROPIC_API_KEY and not GROQ_API_KEY:
        return JSONResponse({"ok": False, "error": "No AI API key configured."}, status_code=500)
    try:
        body = await request.json()
    except Exception:
        body = {}
    region  = (body.get("region") or "Lebanon").strip()
    tier    = (body.get("tier")   or "local").strip()
    country = (body.get("country") or "").strip()

    # Build list of already-tracked store domains from DB
    conn = db_connect(DB_PATH)
    cur  = conn.cursor()
    cur.execute("SELECT base_url FROM stores WHERE enabled=1")
    tracked = []
    for row in cur.fetchall():
        try:
            import urllib.parse as _up
            domain = _up.urlparse(row["base_url"]).netloc.replace("www.", "")
            if domain:
                tracked.append(domain)
        except: pass
    conn.close()

    return StreamingResponse(
        stream_discover_stores(
            region, tracked, ANTHROPIC_API_KEY, GROQ_API_KEY,
            tier=tier, country=country
        ),
        media_type="text/event-stream",
    )
