"""
main.py — thin orchestrator for RadarOS / BrickRadar.
Wires core/ and modules/brickradar/ together into a FastAPI app.
"""

import io
import json
import os
import re
import sqlite3
import time
import threading
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

from fastapi import FastAPI, Request
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates

# ── Core imports ───────────────────────────────────────────────────────────────
from core.db import (
    db_connect, db_init,
    meta_get, meta_set,
    alerts_unread_count, alerts_mark_read,
    persist_snapshot, compute_alerts,
    radarlist_add, radarlist_remove, radarlist_get_ids,
)
from core.engine import merge_catalogs, refresh_all as _core_refresh_all
from core.models import StoreOffer
from core.utils import safe_float, extract_item_number, compute_discount_pct, utc_now_iso
from core.scrapers.shopify import fetch_shopify_store, fetch_new_arrival_items
from core.scrapers.bigcommerce import fetch_bigcommerce_store
from core.ai import build_context, stream_chat, stream_discover_stores, analyze_store

# ── Module imports ─────────────────────────────────────────────────────────────
from modules.brickradar.config import (
    MODULE,
    SHOPIFY_STORES,
    BIGCOMMERCE_STORES,
    NEW_ARRIVAL_COLLECTIONS,
    HARDCODED_STORE_URLS,
    LEGO_THEMES,
)
from modules.brickradar.scrapers import (
    fetch_brickshop,
    fetch_playone,
    fetch_html_stores,
    normalize_theme_category_from_shopify,
)

# ── Paths ──────────────────────────────────────────────────────────────────────
APP_DIR       = os.path.dirname(os.path.abspath(__file__))
TEMPLATES_DIR = os.path.join(APP_DIR, "Brickradar", "app", "templates")
STATIC_DIR    = os.path.join(APP_DIR, "Brickradar", "app", "static")
DB_PATH       = os.path.join(APP_DIR, "Brickradar", "app", "data", "lego_tracker.sqlite3")
TEMPLATE_FILE = "dashboard.html"

# ── API keys ───────────────────────────────────────────────────────────────────
from dotenv import load_dotenv
load_dotenv()
ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY", "")
GROQ_API_KEY      = os.getenv("GROQ_API_KEY", "")

# ── Init ───────────────────────────────────────────────────────────────────────
db_init(DB_PATH)

app       = FastAPI()
templates = Jinja2Templates(directory=TEMPLATES_DIR)
templates.env.filters["tojson"] = lambda v: json.dumps(v)

if os.path.isdir(STATIC_DIR):
    app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                  "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0 Safari/537.36",
}

# ── Helpers ────────────────────────────────────────────────────────────────────

def get_all_store_names() -> List[str]:
    names = list(SHOPIFY_STORES.keys()) + ["BRICKSHOP", "PlayOne"]
    names += list(BIGCOMMERCE_STORES.keys())
    conn = db_connect(DB_PATH)
    cur  = conn.cursor()
    cur.execute("SELECT name FROM stores WHERE enabled=1")
    for r in cur.fetchall():
        if r["name"] not in names:
            names.append(r["name"])
    conn.close()
    from core.utils import order_stores
    return order_stores(names)

def _lsf() -> str:
    """Latest snapshot filter SQL snippet."""
    return "AND captured_at IN (SELECT MAX(captured_at) FROM snapshots GROUP BY store)"


# ── Dashboard ──────────────────────────────────────────────────────────────────

@app.get("/", response_class=HTMLResponse)
def dashboard(request: Request):
    store_names = get_all_store_names()

    conn = db_connect(DB_PATH)
    cur  = conn.cursor()
    cur.execute("""
        SELECT s.* FROM snapshots s
        INNER JOIN (
            SELECT store, item_number, MAX(id) AS max_id
            FROM snapshots GROUP BY store, item_number
        ) t ON s.id = t.max_id
    """)
    rows = cur.fetchall()
    conn.close()

    latest_by_store_item: Dict[Tuple[str, str], sqlite3.Row] = {}
    for r in rows:
        item_key = r["item_number"] or f"SID_{r['store']}_{r['id']}"
        latest_by_store_item[(r["store"], item_key)] = r

    catalogs = []
    for s in store_names:
        cat: Dict[str, Dict[str, Any]] = {}
        for (st, item), r in latest_by_store_item.items():
            if st != s or not item:
                continue
            offer = StoreOffer(
                price=r["price"],
                availability=r["availability"] or "N/A",
                link=r["link"] or "",
                discount_pct=compute_discount_pct(r["price"], r["compare_at"]),
            )
            try:
                img_list = json.loads(r["images_json"]) if r["images_json"] else []
            except Exception:
                img_list = []
            if not img_list and r["image_url"]:
                img_list = [r["image_url"]]
            cat[item] = {
                "item_number": item,
                "title":    r["title"] or "",
                "theme":    r["theme"] or "",
                "category": r["category"] or "",
                "image_url":   r["image_url"] or "",
                "image_list":  img_list,
                "vendor":  "",
                "brand":   (r["brand"] or "").strip().upper() or "LEGO",
                "compare_at": r["compare_at"],
                "stores":  {s: offer},
            }
        catalogs.append(cat)

    merged = merge_catalogs(catalogs)

    qp = request.query_params
    selected_category   = qp.get("category",  "All")
    selected_theme      = qp.get("theme",      "All")
    selected_brand      = qp.get("brand",      "All")
    search_item         = qp.get("search_item", "").strip()
    only_deals          = qp.get("only_deals",      "0") == "1"
    only_alerts         = qp.get("only_alerts",     "0") == "1"
    only_comparable     = qp.get("only_comparable",  "0") == "1"
    only_instock        = qp.get("only_instock",    "0") == "1"
    alert_type_filter   = qp.get("alert_type", "all")
    sort                = qp.get("sort",  "item")
    order               = qp.get("order", "asc")
    per_page            = int(qp.get("per_page", "1000") or "1000")
    page                = max(1, int(qp.get("page", "1") or "1"))

    raw_selected_stores = request.query_params.getlist("stores")
    if raw_selected_stores:
        from core.utils import order_stores
        stores = order_stores([s for s in raw_selected_stores if s in store_names]) or store_names
        selected_stores = stores
    else:
        stores          = store_names
        selected_stores = []

    all_categories = ["All"] + sorted({(r.get("category") or "").strip() for r in merged.values() if (r.get("category") or "").strip()})
    all_themes     = ["All"] + sorted({(r.get("theme")    or "").strip() for r in merged.values() if (r.get("theme")    or "").strip()})
    all_brands     = ["All"] + sorted({(r.get("brand")    or "").strip().upper() for r in merged.values() if (r.get("brand") or "").strip()})

    # Build alert lookup
    ALERT_PRIORITY = {"price_drop": 0, "new_arrival": 1, "new_in_store": 2, "price_increase": 3}
    _ac = db_connect(DB_PATH)
    _acc = _ac.cursor()
    _acc.execute("SELECT item_number, store, alert_type, old_price, new_price FROM alerts WHERE unread=1 ORDER BY id DESC")
    item_store_alerts: Dict[str, Dict[str, Dict]] = {}
    for ar in _acc.fetchall():
        iid, sname, atype = ar["item_number"], ar["store"], ar["alert_type"] or "price_change"
        if iid not in item_store_alerts:
            item_store_alerts[iid] = {}
        cur_p = ALERT_PRIORITY.get(item_store_alerts[iid].get(sname, {}).get("type", "new_in_store"), 99)
        new_p = ALERT_PRIORITY.get(atype, 99)
        if sname not in item_store_alerts[iid] or new_p < cur_p:
            item_store_alerts[iid][sname] = {"type": atype, "store": sname, "old_price": ar["old_price"], "new_price": ar["new_price"]}
    _ac.close()

    def _get_item_alert(item_number, selected):
        store_map = item_store_alerts.get(item_number, {})
        best, best_p = None, 99
        for s in selected:
            a = store_map.get(s)
            if a:
                p = ALERT_PRIORITY.get(a["type"], 99)
                if p < best_p:
                    best_p, best = p, a
        return best

    def _get_item_alert_types(item_number, selected):
        store_map = item_store_alerts.get(item_number, {})
        return {store_map[s]["type"] for s in selected if s in store_map}

    filtered = []
    for rec in merged.values():
        if selected_category != "All" and (rec.get("category") or "").strip() != selected_category:
            continue
        if selected_theme != "All" and (rec.get("theme") or "").strip() != selected_theme:
            continue
        if selected_brand != "All" and (rec.get("brand") or "").strip().upper() != selected_brand.upper():
            continue
        if search_item and search_item not in (rec.get("item_number") or ""):
            continue

        rec2 = dict(rec)
        rec2["stores"] = {s: rec["stores"].get(s) for s in stores if s in rec["stores"]}
        if not any(rec2["stores"].values()):
            continue

        raw_img = rec2.get("image_url") or ""
        rec2["image_proxy"]      = re.sub(r"_[0-9]+x[0-9]*\.", ".", raw_img) if raw_img else ""
        rec2["image_list_clean"] = [re.sub(r"_[0-9]+x[0-9]*\.", ".", u) for u in (rec2.get("image_list") or [])[:4]]
        rec2["image_list_json"]  = json.dumps(rec2["image_list_clean"])

        lp, ls = None, None
        for s in stores:
            off = rec2["stores"].get(s)
            if off and off.price is not None and (lp is None or off.price < lp):
                lp, ls = off.price, s
        rec2["lowest_price"] = lp
        rec2["lowest_store"] = ls

        if only_deals and not any((off.discount_pct or 0) > 0 for off in rec2["stores"].values() if off):
            continue

        _plain_item       = rec2["item_number"].split("|")[-1] if "|" in (rec2["item_number"] or "") else rec2["item_number"]
        rec2["alert"]       = _get_item_alert(_plain_item, stores)
        rec2["alert_types"] = _get_item_alert_types(_plain_item, stores)

        if only_alerts:
            if not rec2["alert_types"]:
                continue
            if alert_type_filter != "all" and alert_type_filter not in rec2["alert_types"]:
                continue
        if only_comparable and not all(rec2["stores"].get(s) and rec2["stores"][s].price is not None for s in stores):
            continue
        if only_instock and not any("in stock" in (rec2["stores"].get(s) and rec2["stores"][s].availability or "").lower() for s in stores):
            continue

        filtered.append(rec2)

    def sort_key(r):
        if sort == "brand":    return (r.get("brand") or "").upper()
        if sort == "item":     return int(r["item_number"]) if str(r["item_number"]).isdigit() else 10**9
        if sort == "title":    return (r.get("title") or "").lower()
        if sort == "theme":    return (r.get("theme") or "").lower()
        if sort == "category": return (r.get("category") or "").lower()
        if sort == "price":    return r.get("lowest_price") if r.get("lowest_price") is not None else 10**12
        if sort.startswith("store:"):
            off = (r.get("stores") or {}).get(sort.split(":", 1)[1])
            return off.price if (off and off.price is not None) else 10**12
        return r.get("item_number") or ""

    filtered.sort(key=sort_key, reverse=(order == "desc"))

    total       = len(filtered)
    total_pages = max(1, (total + per_page - 1) // per_page)
    page        = min(page, total_pages)
    page_rows   = filtered[(page - 1) * per_page : page * per_page]

    return templates.TemplateResponse(TEMPLATE_FILE, {
        "request": request,
        "rows": page_rows, "total": total,
        "page": page, "per_page": per_page, "total_pages": total_pages,
        "sort": sort, "order": order,
        "stores": stores, "store_names": store_names,
        "all_categories": all_categories, "all_themes": all_themes, "all_brands": all_brands,
        "selected_category": selected_category, "selected_theme": selected_theme, "selected_brand": selected_brand,
        "search_item": search_item, "compare": "all",
        "selected_stores": selected_stores, "raw_selected_stores": raw_selected_stores,
        "only_deals": only_deals, "only_alerts": only_alerts,
        "only_comparable": only_comparable, "only_instock": only_instock,
        "alert_type_filter": alert_type_filter,
        "last_updated": meta_get(DB_PATH, "last_updated") or "never",
        "alerts_unread": alerts_unread_count(DB_PATH),
        "price_range": "all", "min_price": "", "max_price": "",
    })


# ── Refresh / SSE ──────────────────────────────────────────────────────────────

@app.post("/api/refresh")
def api_refresh():
    import queue as _queue
    results = {}
    q = _queue.Queue()

    def _progress(msg):
        q.put(msg)

    import threading as _t
    done_ev = _t.Event()

    def _run():
        _core_refresh_all(
            db_path=DB_PATH,
            shopify_stores={k: {**v, "normalize_theme_fn": normalize_theme_category_from_shopify} for k, v in SHOPIFY_STORES.items()},
            bigcommerce_stores=BIGCOMMERCE_STORES,
            html_stores=[],
            fetch_shopify_fn=fetch_shopify_store,
            fetch_bigcommerce_fn=fetch_bigcommerce_store,
            fetch_html_stores_fn=fetch_html_stores,
            progress_fn=_progress,
        )
        done_ev.set()

    _t.Thread(target=_run, daemon=True).start()
    done_ev.wait(timeout=600)
    return JSONResponse({"ok": True})


@app.get("/api/refresh/stream")
def api_refresh_stream(stores: str = ""):
    selected = [s.strip() for s in stores.split(",") if s.strip()] if stores else []
    import queue as _queue

    SEQUENTIAL_AFTER = {"Brix & Figures", "Thetoystorelb"}

    def event_stream():
        captured_at = utc_now_iso()
        tasks = []
        for sname, cfg in SHOPIFY_STORES.items():
            tasks.append(("shopify", sname, cfg))

        conn = db_connect(DB_PATH)
        cur  = conn.cursor()
        cur.execute("SELECT name, base_url, platform, vat_multiplier, new_arrivals_collection, collection_slug, lego_only FROM stores WHERE enabled=1")
        for ds in cur.fetchall():
            sname = ds["name"]
            if sname in SHOPIFY_STORES:
                continue
            if ds["platform"] == "shopify":
                tasks.append(("shopify_db", sname, {
                    "url": ds["base_url"], "vat_multiplier": ds["vat_multiplier"],
                    "new_arrivals_collection": ds["new_arrivals_collection"],
                    "collection_slug": ds["collection_slug"],
                    "lego_only": bool(ds["lego_only"]),
                }))
        conn.close()

        tasks.append(("brickshop", None, None))
        tasks.append(("playone",   None, None))
        for bc_name, bc_cfg in BIGCOMMERCE_STORES.items():
            tasks.append(("bigcommerce", bc_name, bc_cfg))

        if selected:
            def _label(t): return t[1] or ("BRICKSHOP" if t[0] == "brickshop" else ("PlayOne" if t[0] == "playone" else t[0]))
            tasks = [t for t in tasks if _label(t) in selected]

        total = len(tasks)
        q     = _queue.Queue()

        def _run_task(task, stagger_idx: int = 0):
            if stagger_idx > 0:
                time.sleep(stagger_idx * 5)
            kind, sname, cfg = task
            label = sname or kind
            try:
                if kind in ("shopify", "shopify_db"):
                    base_url = cfg.get("url") or cfg.get("base_url", "")
                    new_items = fetch_new_arrival_items(label, base_url, cfg.get("new_arrivals_collection") or "")
                    result = fetch_shopify_store(
                        store_name=label, base_url=base_url,
                        vat_multiplier=float(cfg.get("vat_multiplier", 1.0)),
                        new_items=new_items,
                        collection_slug=cfg.get("collection_slug") or "",
                        lego_only=cfg.get("lego_only", False),
                        normalize_theme_fn=normalize_theme_category_from_shopify,
                    )
                    q.put(("ok", label, result))
                elif kind == "brickshop":
                    q.put(("ok", "BRICKSHOP", fetch_brickshop()))
                elif kind == "bigcommerce":
                    q.put(("ok", label, fetch_bigcommerce_store(
                        label, cfg["url"], cfg.get("collection_slug", ""),
                        cfg.get("lego_only", True), float(cfg.get("vat_multiplier", 1.0))
                    )))
                else:
                    q.put(("ok", "PlayOne", fetch_playone()))
            except Exception as e:
                q.put(("err", label, str(e)))

        yield f"data: {json.dumps({'type': 'start', 'total': total})}\n\n"

        # Keepalive ping
        _stop_ping = threading.Event()
        def _ping():
            while not _stop_ping.wait(20):
                q.put(("ping", None, None))
        threading.Thread(target=_ping, daemon=True).start()

        # Background persist
        _sse_q    = _queue.Queue()
        _persist_done = threading.Event()

        def _background_persist():
            done_bg = 0
            while done_bg < total:
                try:
                    status, name, data = q.get(timeout=600)
                except Exception:
                    print(f"[PERSIST] queue timeout after {done_bg}/{total}")
                    break
                if status == "ping":
                    _sse_q.put(("ping", None, None)); continue
                done_bg += 1
                if status == "ok":
                    count = len(data) if isinstance(data, dict) else 0
                    persist_snapshot(DB_PATH, captured_at, name, data)
                    compute_alerts(DB_PATH, captured_at, name, data)
                    _sse_q.put(("ok", name, count, done_bg))
                else:
                    _sse_q.put(("err", name, data, done_bg))
            meta_set(DB_PATH, "last_updated", captured_at)
            _sse_q.put(("done", None, None, total))
            _persist_done.set()

        threading.Thread(target=_background_persist, daemon=True).start()

        # Submit tasks
        main_tasks = [t for t in tasks if (t[1] or t[0]) not in SEQUENTIAL_AFTER]
        late_tasks = [t for t in tasks if (t[1] or t[0]) in SEQUENTIAL_AFTER]
        pool = ThreadPoolExecutor(max_workers=len(main_tasks) + 1)
        shopify_idx = 0
        for t in main_tasks:
            if t[0] in ("brickshop", "playone"):
                pool.submit(_run_task, t, 0)
            else:
                pool.submit(_run_task, t, shopify_idx); shopify_idx += 1

        if late_tasks:
            def _run_late():
                time.sleep(shopify_idx * 5 + 10)
                for lt in late_tasks:
                    _run_task(lt, 0); time.sleep(10)
            pool.submit(_run_late)
        pool.shutdown(wait=False)

        while True:
            msg  = _sse_q.get()
            kind = msg[0]
            if kind == "ping":
                yield ": keepalive\n\n"
            elif kind == "ok":
                _, name, count, done_n = msg
                yield f"data: {json.dumps({'type': 'store', 'name': name, 'count': count, 'done': done_n, 'total': total})}\n\n"
            elif kind == "err":
                _, name, err, done_n = msg
                yield f"data: {json.dumps({'type': 'error', 'name': name, 'error': err, 'done': done_n, 'total': total})}\n\n"
            elif kind == "done":
                break

        _stop_ping.set()
        yield f"data: {json.dumps({'type': 'done'})}\n\n"

    return StreamingResponse(event_stream(), media_type="text/event-stream")


# ── Alerts ─────────────────────────────────────────────────────────────────────

@app.post("/api/alerts/mark_read")
def api_mark_alerts_read():
    alerts_mark_read(DB_PATH)
    return JSONResponse({"ok": True, "unread": alerts_unread_count(DB_PATH)})


# ── Stores page ────────────────────────────────────────────────────────────────

@app.get("/stores", response_class=HTMLResponse)
def stores_page():
    f = os.path.join(APP_DIR, "Brickradar", "app", "stores.html")
    return HTMLResponse(open(f, encoding="utf-8").read())


@app.get("/api/stores/logos")
def api_store_logos():
    """Return logo URLs for all stores."""
    static_dir = os.path.join(STATIC_DIR, "logos")
    os.makedirs(static_dir, exist_ok=True)
    logos = {}
    all_urls = dict(HARDCODED_STORE_URLS)
    for bc_name, bc_cfg in BIGCOMMERCE_STORES.items():
        all_urls[bc_name] = bc_cfg["url"]
    conn = db_connect(DB_PATH)
    cur  = conn.cursor()
    cur.execute("SELECT name, base_url FROM stores WHERE enabled=1")
    for r in cur.fetchall():
        all_urls[r["name"]] = r["base_url"]
    conn.close()
    for name in all_urls:
        safe_name = re.sub(r"[^a-zA-Z0-9_-]", "_", name)
        for ext in [".png", ".jpg", ".svg", ".ico", ".webp", ".avif"]:
            p = os.path.join(static_dir, f"{safe_name}{ext}")
            if os.path.exists(p):
                logos[name] = f"/static/logos/{safe_name}{ext}"
                break
    return JSONResponse(logos)


@app.post("/api/stores/fetch_logos")
def api_fetch_all_logos():
    """Trigger logo fetch for all stores in background."""
    import threading as _flt
    import urllib.parse

    def _fetch_logo(base_url: str, store_name: str):
        import httpx as _hx
        static_dir = os.path.join(STATIC_DIR, "logos")
        os.makedirs(static_dir, exist_ok=True)
        safe_name = re.sub(r"[^a-zA-Z0-9_-]", "_", store_name)
        for ext in [".png", ".jpg", ".svg", ".ico", ".webp", ".avif"]:
            if os.path.exists(os.path.join(static_dir, f"{safe_name}{ext}")):
                return
        try:
            with _hx.Client(timeout=10, follow_redirects=True, headers=HEADERS) as client:
                r = client.get(base_url.rstrip("/") + "/")
                if r.status_code != 200:
                    return
                from bs4 import BeautifulSoup as _BS
                soup = _BS(r.text, "lxml")
                logo_url = None
                for sel, attr in [('meta[property="og:image"]', "content"),
                                   ('link[rel*="apple-touch-icon"]', "href"),
                                   ('link[rel*="icon"]', "href")]:
                    el = soup.select_one(sel)
                    if el and el.get(attr):
                        logo_url = el[attr]; break
                if not logo_url:
                    logo_url = base_url.rstrip("/") + "/favicon.ico"
                if logo_url.startswith("//"):
                    logo_url = "https:" + logo_url
                elif not logo_url.startswith("http"):
                    logo_url = urllib.parse.urljoin(base_url, logo_url)
                lr = client.get(logo_url)
                if lr.status_code == 200 and len(lr.content) > 100:
                    ct = lr.headers.get("content-type", "")
                    ext = ".svg" if "svg" in ct else ".ico" if "ico" in ct or logo_url.endswith(".ico") else ".jpg" if "jpeg" in ct or "jpg" in ct else ".webp" if "webp" in ct else ".png"
                    with open(os.path.join(static_dir, f"{safe_name}{ext}"), "wb") as f:
                        f.write(lr.content)
        except Exception as e:
            print(f"[Logo] {store_name}: {e}")

    all_urls = dict(HARDCODED_STORE_URLS)
    for bc_name, bc_cfg in BIGCOMMERCE_STORES.items():
        all_urls[bc_name] = bc_cfg["url"]
    conn = db_connect(DB_PATH)
    cur  = conn.cursor()
    cur.execute("SELECT name, base_url FROM stores WHERE enabled=1")
    for r in cur.fetchall():
        all_urls[r["name"]] = r["base_url"]
    conn.close()
    for name, url in all_urls.items():
        _flt.Thread(target=_fetch_logo, args=(url, name), daemon=True).start()
    return JSONResponse({"ok": True, "message": "Fetching logos in background"})


@app.get("/api/stores")
def api_get_stores():
    conn = db_connect(DB_PATH)
    cur  = conn.cursor()
    cur.execute("SELECT store, COUNT(DISTINCT item_number) as cnt, MAX(captured_at) as last_seen FROM snapshots GROUP BY store")
    snap_counts = {r["store"]: {"count": r["cnt"], "last_seen": r["last_seen"]} for r in cur.fetchall()}
    cur.execute("SELECT * FROM stores ORDER BY name")
    db_rows = {r["name"]: dict(r) for r in cur.fetchall()}
    conn.close()

    hardcoded = []
    for sname, cfg in SHOPIFY_STORES.items():
        hardcoded.append({"id": None, "name": sname, "base_url": cfg["url"], "platform": "shopify",
                           "vat_multiplier": cfg.get("vat_multiplier", 1.0),
                           "collection_slug": cfg.get("collection_slug"),
                           "new_arrivals_collection": NEW_ARRIVAL_COLLECTIONS.get(sname),
                           "enabled": 1, "hardcoded": True, "lego_only": cfg.get("lego_only", False)})
    hardcoded.append({"id": None, "name": "BRICKSHOP", "base_url": "https://brickshop.me", "platform": "woocommerce", "vat_multiplier": 1.0, "collection_slug": None, "new_arrivals_collection": None, "enabled": 1, "hardcoded": True, "lego_only": False})
    hardcoded.append({"id": None, "name": "PlayOne",   "base_url": "https://playone.com.lb", "platform": "html", "vat_multiplier": 1.0, "collection_slug": None, "new_arrivals_collection": None, "enabled": 1, "hardcoded": True, "lego_only": True})
    for bc_name, bc_cfg in BIGCOMMERCE_STORES.items():
        hardcoded.append({"id": None, "name": bc_name, "base_url": bc_cfg["url"], "platform": "bigcommerce",
                           "vat_multiplier": bc_cfg.get("vat_multiplier", 1.0),
                           "collection_slug": bc_cfg.get("collection_slug"),
                           "new_arrivals_collection": None, "enabled": 1, "hardcoded": True,
                           "lego_only": bc_cfg.get("lego_only", True)})

    all_stores = []
    hardcoded_names = set()
    for s in hardcoded:
        snap = snap_counts.get(s["name"], {})
        s["product_count"] = snap.get("count", 0)
        s["last_scraped"]  = snap.get("last_seen", "")
        all_stores.append(s)
        hardcoded_names.add(s["name"])

    for name, row in db_rows.items():
        if name in hardcoded_names:
            continue
        snap = snap_counts.get(name, {})
        row["product_count"] = snap.get("count", 0)
        row["last_scraped"]  = snap.get("last_seen", row.get("last_scraped", ""))
        row["hardcoded"]     = False
        all_stores.append(row)

    return JSONResponse(all_stores)


@app.post("/api/stores/test")
async def api_test_store(request: Request):
    import httpx as _httpx
    body            = await request.json()
    url             = ("https://" + (body.get("url") or "").strip().rstrip("/").lstrip("https://").lstrip("http://"))
    collection_slug = (body.get("collection_slug") or "").strip() or None
    result = {"url": url, "platform": None, "product_count": 0, "samples": [], "error": None, "collection_slug": collection_slug}

    def _products_url(base, page, slug=None):
        if slug: return f"{base}/collections/{slug}/products.json?limit=250&page={page}"
        return f"{base}/products.json?limit=250&page={page}"

    try:
        r = _httpx.get(_products_url(url, 1, collection_slug), timeout=15, follow_redirects=True, headers=HEADERS)
        if r.status_code == 200:
            products = r.json().get("products") or []
            if products:
                result["platform"] = "shopify"
                page, total = 1, 0
                with _httpx.Client(timeout=15, follow_redirects=True, headers=HEADERS) as client:
                    while True:
                        pr = client.get(_products_url(url, page, collection_slug))
                        if pr.status_code != 200: break
                        prods = pr.json().get("products") or []
                        if not prods: break
                        total += len(prods)
                        if len(prods) < 250 or page >= 5: break
                        page += 1
                result["product_count"] = total
                for p in products[:5]:
                    title       = (p.get("title") or "").strip()
                    item_number = extract_item_number(title)
                    if not item_number:
                        for v in (p.get("variants") or []):
                            item_number = extract_item_number(v.get("sku") or "")
                            if item_number: break
                    variants = p.get("variants") or []
                    price    = safe_float(variants[0].get("price")) if variants else None
                    images   = p.get("images") or []
                    result["samples"].append({"title": title, "item_number": item_number or "—", "price": price, "image": images[0].get("src", "") if images else ""})
            else:
                result["error"] = "Shopify endpoint returned no products"
        else:
            result["error"] = f"Not Shopify (HTTP {r.status_code})"
    except Exception as e:
        result["error"] = str(e)

    return JSONResponse(result)


@app.post("/api/stores/add")
async def api_add_store(request: Request):
    body            = await request.json()
    name            = (body.get("name") or "").strip()
    url             = (body.get("url")  or "").strip().rstrip("/")
    platform        = body.get("platform") or "shopify"
    vat             = float(body.get("vat_multiplier") or 1.0)
    new_arrivals    = (body.get("new_arrivals_collection") or "").strip() or None
    collection_slug = (body.get("collection_slug") or "").strip() or None
    lego_only       = int(bool(body.get("lego_only", False)))
    if not name or not url:
        return JSONResponse({"ok": False, "error": "Name and URL required"}, status_code=400)
    conn = db_connect(DB_PATH)
    cur  = conn.cursor()
    try:
        cur.execute("""
            INSERT INTO stores(name, base_url, platform, vat_multiplier, new_arrivals_collection, collection_slug, lego_only, enabled)
            VALUES(?,?,?,?,?,?,?,1)
        """, (name, url, platform, vat, new_arrivals, collection_slug, lego_only))
        conn.commit()
    except Exception as e:
        conn.close()
        return JSONResponse({"ok": False, "error": str(e)}, status_code=400)
    conn.close()
    return JSONResponse({"ok": True})


@app.post("/api/stores/toggle")
async def api_toggle_store(request: Request):
    body     = await request.json()
    store_id = body.get("id")
    conn     = db_connect(DB_PATH)
    conn.execute("UPDATE stores SET enabled = 1 - enabled WHERE id=?", (store_id,))
    conn.commit()
    cur = conn.execute("SELECT enabled FROM stores WHERE id=?", (store_id,))
    row = cur.fetchone()
    conn.close()
    return JSONResponse({"ok": True, "enabled": row["enabled"] if row else 0})


@app.post("/api/stores/delete")
async def api_delete_store(request: Request):
    body     = await request.json()
    store_id = body.get("id")
    conn     = db_connect(DB_PATH)
    conn.execute("DELETE FROM stores WHERE id=?", (store_id,))
    conn.commit()
    conn.close()
    return JSONResponse({"ok": True})


# ── Analytics ──────────────────────────────────────────────────────────────────

@app.get("/analytics", response_class=HTMLResponse)
def analytics_page():
    f = os.path.join(APP_DIR, "Brickradar", "app", "analytics.html")
    return HTMLResponse(open(f, encoding="utf-8").read())


@app.get("/api/analytics/kpis")
def api_analytics_kpis():
    conn = db_connect(DB_PATH)
    cur  = conn.cursor()
    lsf  = _lsf()
    cur.execute(f"SELECT COUNT(DISTINCT item_number) FROM snapshots WHERE 1=1 {lsf}")
    total_products = cur.fetchone()[0]
    cur.execute(f"SELECT COUNT(DISTINCT store) FROM snapshots WHERE 1=1 {lsf}")
    total_stores = cur.fetchone()[0]
    cur.execute(f"SELECT COUNT(*) FROM snapshots WHERE compare_at IS NOT NULL AND compare_at > price {lsf}")
    discounted = cur.fetchone()[0]
    cur.execute("SELECT COUNT(DISTINCT item_number) FROM alerts WHERE alert_type='new_arrival' AND created_at >= datetime('now','-7 days')")
    new_arrivals_7d = cur.fetchone()[0]
    cur.execute(f"SELECT AVG(price) FROM snapshots WHERE price IS NOT NULL {lsf}")
    avg_price = cur.fetchone()[0]
    cur.execute(f"SELECT COUNT(*) FROM snapshots WHERE availability NOT IN ('available','in_stock','') AND availability IS NOT NULL {lsf}")
    out_of_stock = cur.fetchone()[0]
    conn.close()
    return JSONResponse({"total_products": total_products, "total_stores": total_stores,
                         "discounted": discounted, "new_arrivals_7d": new_arrivals_7d,
                         "avg_price": round(avg_price, 2) if avg_price else 0, "out_of_stock": out_of_stock})


@app.get("/api/analytics/items_per_brand_store")
def api_items_per_brand_store():
    conn = db_connect(DB_PATH)
    cur  = conn.cursor()
    cur.execute(f"SELECT store, UPPER(TRIM(brand)) as brand, COUNT(DISTINCT item_number) as cnt FROM snapshots WHERE brand IS NOT NULL AND brand != '' {_lsf()} GROUP BY store, brand ORDER BY store, cnt DESC")
    rows = [{"store": r[0], "brand": r[1], "count": r[2]} for r in cur.fetchall()]
    conn.close()
    return JSONResponse(rows)


@app.get("/api/analytics/most_expensive_per_brand_store")
def api_most_expensive_per_brand_store():
    conn = db_connect(DB_PATH)
    cur  = conn.cursor()
    cur.execute(f"SELECT store, UPPER(TRIM(brand)) as brand, title, item_number, MAX(price) as price, link, image_url FROM snapshots WHERE price IS NOT NULL AND brand IS NOT NULL AND brand != '' {_lsf()} GROUP BY store, brand ORDER BY store, price DESC")
    rows = [{"store": r[0], "brand": r[1], "title": r[2], "item_number": r[3], "price": r[4], "link": r[5], "image": r[6]} for r in cur.fetchall()]
    conn.close()
    return JSONResponse(rows)


@app.get("/api/analytics/items_per_theme_store")
def api_items_per_theme_store():
    conn = db_connect(DB_PATH)
    cur  = conn.cursor()
    cur.execute(f"SELECT store, TRIM(theme) as theme, COUNT(DISTINCT item_number) as cnt FROM snapshots WHERE theme IS NOT NULL AND theme != '' {_lsf()} GROUP BY store, theme ORDER BY cnt DESC")
    rows = [{"store": r[0], "theme": r[1], "count": r[2]} for r in cur.fetchall()]
    conn.close()
    return JSONResponse(rows)


@app.get("/api/analytics/new_arrivals_per_store")
def api_new_arrivals_per_store():
    conn = db_connect(DB_PATH)
    cur  = conn.cursor()
    cur.execute("SELECT store, COUNT(DISTINCT item_number) as cnt FROM alerts WHERE alert_type='new_arrival' GROUP BY store ORDER BY cnt DESC")
    rows = [{"store": r[0], "count": r[1]} for r in cur.fetchall()]
    conn.close()
    return JSONResponse(rows)


@app.get("/api/analytics/discounts_per_brand_store")
def api_discounts_per_brand_store():
    conn = db_connect(DB_PATH)
    cur  = conn.cursor()
    cur.execute(f"SELECT store, UPPER(TRIM(brand)) as brand, COUNT(*) as cnt, AVG(ROUND((compare_at-price)/compare_at*100,1)) as avg_pct FROM snapshots WHERE compare_at IS NOT NULL AND compare_at > price AND brand IS NOT NULL AND brand != '' {_lsf()} GROUP BY store, brand ORDER BY cnt DESC")
    rows = [{"store": r[0], "brand": r[1], "count": r[2], "avg_discount_pct": round(r[3], 1) if r[3] else 0} for r in cur.fetchall()]
    conn.close()
    return JSONResponse(rows)


# ── Advanced analysis ──────────────────────────────────────────────────────────

@app.get("/advanced", response_class=HTMLResponse)
def advanced_page():
    f = os.path.join(APP_DIR, "Brickradar", "app", "advanced.html")
    return HTMLResponse(open(f, encoding="utf-8").read())


@app.get("/api/advanced/price_spread")
def api_price_spread():
    conn = db_connect(DB_PATH)
    cur  = conn.cursor()
    cur.execute(f"""
        SELECT item_number, title, brand, theme, category,
               COUNT(DISTINCT store) as store_count,
               MIN(price) as min_price, MAX(price) as max_price,
               ROUND(MAX(price)-MIN(price),2) as spread,
               ROUND((MAX(price)-MIN(price))/MIN(price)*100,1) as spread_pct,
               GROUP_CONCAT(store||':'||ROUND(price,2)||':'||availability,'|') as store_data
        FROM snapshots WHERE price IS NOT NULL AND price > 0 {_lsf()}
        GROUP BY item_number HAVING COUNT(DISTINCT store) >= 2 ORDER BY spread DESC
    """)
    rows = []
    for r in cur.fetchall():
        stores = {}
        for chunk in (r["store_data"] or "").split("|"):
            parts = chunk.split(":")
            if len(parts) >= 3:
                stores[parts[0]] = {"price": float(parts[1]), "availability": parts[2]}
        rows.append({"item_number": r["item_number"], "title": r["title"] or "", "brand": r["brand"] or "",
                     "theme": r["theme"] or "", "category": r["category"] or "",
                     "store_count": r["store_count"], "min_price": r["min_price"], "max_price": r["max_price"],
                     "spread": r["spread"], "spread_pct": r["spread_pct"], "stores": stores})
    conn.close()
    return JSONResponse(rows)


@app.get("/api/advanced/store_behavior")
def api_store_behavior():
    conn = db_connect(DB_PATH)
    cur  = conn.cursor()
    cur.execute(f"SELECT item_number FROM snapshots WHERE price IS NOT NULL AND price > 0 {_lsf()} GROUP BY item_number HAVING COUNT(DISTINCT store) >= 2")
    comparable_items = [r[0] for r in cur.fetchall()]
    if not comparable_items:
        conn.close()
        return JSONResponse([])
    ph = ",".join("?" * len(comparable_items))
    cur.execute(f"""
        SELECT s.store, COUNT(*) as items_carried, AVG(s.price) as avg_price,
               SUM(CASE WHEN s.price=m.min_price THEN 1 ELSE 0 END) as cheapest_count,
               SUM(CASE WHEN s.price=m.max_price THEN 1 ELSE 0 END) as priciest_count,
               AVG(ROUND((s.price-m.min_price)/NULLIF(m.min_price,0)*100,1)) as avg_premium_pct
        FROM snapshots s
        JOIN (SELECT item_number, MIN(price) as min_price, MAX(price) as max_price
              FROM snapshots WHERE price IS NOT NULL AND price > 0 AND item_number IN ({ph}) GROUP BY item_number) m
        ON s.item_number=m.item_number
        WHERE s.price IS NOT NULL AND s.price > 0 AND s.item_number IN ({ph})
        GROUP BY s.store ORDER BY avg_premium_pct ASC
    """, comparable_items + comparable_items)
    rows = []
    for r in cur.fetchall():
        total = r["items_carried"] or 1
        rows.append({"store": r["store"], "items_carried": r["items_carried"],
                     "avg_price": round(r["avg_price"] or 0, 2),
                     "cheapest_count": r["cheapest_count"] or 0,
                     "cheapest_pct": round((r["cheapest_count"] or 0) / total * 100, 1),
                     "priciest_count": r["priciest_count"] or 0,
                     "priciest_pct": round((r["priciest_count"] or 0) / total * 100, 1),
                     "avg_premium_pct": round(r["avg_premium_pct"] or 0, 1)})
    conn.close()
    return JSONResponse(rows)


@app.get("/api/advanced/deal_detector")
def api_deal_detector():
    conn = db_connect(DB_PATH)
    cur  = conn.cursor()
    cur.execute(f"""
        SELECT s.item_number, s.title, s.brand, s.theme, s.store,
               s.price, s.compare_at,
               ROUND((s.compare_at-s.price)/s.compare_at*100,1) as discount_pct,
               m.min_other_price, m.stores_full_price,
               ROUND(m.min_other_price-s.price,2) as saving_vs_others
        FROM snapshots s
        JOIN (SELECT item_number,
                     MIN(CASE WHEN (compare_at IS NULL OR compare_at<=price) THEN price END) as min_other_price,
                     COUNT(CASE WHEN (compare_at IS NULL OR compare_at<=price) THEN 1 END) as stores_full_price
              FROM snapshots WHERE price IS NOT NULL AND price > 0 {_lsf()} GROUP BY item_number) m
        ON s.item_number=m.item_number
        WHERE s.compare_at IS NOT NULL AND s.compare_at > s.price
          AND m.min_other_price IS NOT NULL AND m.min_other_price > s.price
          AND s.price IS NOT NULL {_lsf()}
        ORDER BY saving_vs_others DESC
    """)
    rows = [{"item_number": r["item_number"], "title": r["title"] or "", "brand": r["brand"] or "",
             "theme": r["theme"] or "", "store": r["store"], "price": r["price"],
             "compare_at": r["compare_at"], "discount_pct": r["discount_pct"],
             "min_other_price": r["min_other_price"], "stores_full_price": r["stores_full_price"],
             "saving_vs_others": r["saving_vs_others"]} for r in cur.fetchall()]
    conn.close()
    return JSONResponse(rows)


@app.post("/api/advanced/export")
async def api_advanced_export(request: Request):
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    body = await request.json()
    tab  = body.get("tab", "data")
    rows = body.get("rows", [])
    if not rows:
        return JSONResponse({"error": "No data"})
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = {"spread": "Price Spread", "behavior": "Store Behavior", "deals": "Deal Detector"}.get(tab, tab)
    headers  = list(rows[0].keys())
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h.replace("_", " ").title())
        cell.fill = PatternFill("solid", fgColor="1E293B")
        cell.font = Font(bold=True, color="94A3B8", size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for ri, row in enumerate(rows, 2):
        for ci, key in enumerate(headers, 1):
            val = row.get(key, "")
            if isinstance(val, str):
                try: val = float(val)
                except: pass
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = Font(size=9)
    ws.freeze_panes = "A2"
    for ci, key in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(ci)].width = 18 if "title" in key else 12
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    return StreamingResponse(iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=lego_{tab}_{ts}.xlsx"})


# ── RadarList ──────────────────────────────────────────────────────────────────

@app.get("/radarlist", response_class=HTMLResponse)
def radarlist_page():
    f = os.path.join(APP_DIR, "Brickradar", "app", "radarlist.html")
    return HTMLResponse(open(f, encoding="utf-8").read())


@app.get("/api/radarlist")
def api_radarlist_get():
    conn = db_connect(DB_PATH)
    cur  = conn.cursor()
    cur.execute("SELECT * FROM radarlist ORDER BY added_at DESC")
    items = [dict(r) for r in cur.fetchall()]
    lsf   = _lsf()
    for item in items:
        cur.execute(f"SELECT store, price, compare_at, availability, link, image_url FROM snapshots WHERE item_number=? {lsf} ORDER BY price ASC", (item["item_number"],))
        prices = [dict(r) for r in cur.fetchall()]
        item["current_prices"] = prices
        item["min_price"]  = min((p["price"] for p in prices if p["price"]), default=None)
        item["min_store"]  = next((p["store"] for p in prices if p["price"] == item["min_price"]), None)
        item["image_url"]  = next((p["image_url"] for p in prices if p.get("image_url")), None)
        if item["added_price"] and item["min_price"]:
            item["price_drop"]     = round(item["added_price"] - item["min_price"], 2)
            item["price_drop_pct"] = round((item["added_price"] - item["min_price"]) / item["added_price"] * 100, 1)
        else:
            item["price_drop"] = 0
            item["price_drop_pct"] = 0
    conn.close()
    return JSONResponse(items)


@app.post("/api/radarlist/add")
async def api_radarlist_add(request: Request):
    body        = await request.json()
    item_number = body.get("item_number", "").strip()
    if not item_number:
        return JSONResponse({"ok": False, "error": "item_number required"})
    conn = db_connect(DB_PATH)
    cur  = conn.cursor()
    try:
        cur.execute(f"SELECT title, brand, theme, price, store FROM snapshots WHERE item_number=? {_lsf()} ORDER BY price ASC LIMIT 1", (item_number,))
        row   = cur.fetchone()
        title = body.get("title") or (row["title"] if row else "")
        brand = body.get("brand") or (row["brand"] if row else "")
        theme = body.get("theme") or (row["theme"] if row else "")
        price = row["price"] if row else None
        store = row["store"] if row else None
        cur.execute("INSERT INTO radarlist(item_number,title,brand,theme,added_price,added_store) VALUES(?,?,?,?,?,?) ON CONFLICT(item_number) DO NOTHING",
                    (item_number, title, brand, theme, price, store))
        conn.commit()
        added = cur.rowcount > 0
    except Exception as e:
        conn.close()
        return JSONResponse({"ok": False, "error": str(e)})
    conn.close()
    return JSONResponse({"ok": True, "added": added})


@app.post("/api/radarlist/remove")
async def api_radarlist_remove(request: Request):
    body        = await request.json()
    item_number = body.get("item_number", "").strip()
    conn        = db_connect(DB_PATH)
    conn.execute("DELETE FROM radarlist WHERE item_number=?", (item_number,))
    conn.commit()
    conn.close()
    return JSONResponse({"ok": True})


@app.get("/api/radarlist/ids")
def api_radarlist_ids():
    conn = db_connect(DB_PATH)
    cur  = conn.cursor()
    cur.execute("SELECT item_number FROM radarlist")
    ids  = [r[0] for r in cur.fetchall()]
    conn.close()
    return JSONResponse(ids)


# ── AI routes ──────────────────────────────────────────────────────────────────

@app.get("/api/ai/test")
def api_ai_test():
    result = {
        "anthropic_key_set":  bool(ANTHROPIC_API_KEY),
        "groq_key_set":       bool(GROQ_API_KEY),
        "active_provider":    "anthropic" if ANTHROPIC_API_KEY else ("groq" if GROQ_API_KEY else "none"),
    }
    try:
        ctx = build_context(DB_PATH)
        result["context_length"] = len(ctx)
        result["context_ok"]     = True
    except Exception as e:
        import traceback
        result["context_ok"]    = False
        result["context_error"] = str(e)
        result["traceback"]     = traceback.format_exc()
    return JSONResponse(result)


@app.post("/api/ai/chat")
async def api_ai_chat(request: Request):
    if not ANTHROPIC_API_KEY and not GROQ_API_KEY:
        return JSONResponse({"ok": False, "error": "No AI API key configured."}, status_code=500)
    body     = await request.json()
    messages = body.get("messages", [])
    page     = body.get("page", "dashboard")
    if not messages:
        return JSONResponse({"ok": False, "error": "No messages"}, status_code=400)
    return StreamingResponse(
        stream_chat(DB_PATH, messages, page, ANTHROPIC_API_KEY, GROQ_API_KEY, MODULE["name"]),
        media_type="text/event-stream",
    )


@app.post("/api/ai/analyze_store")
async def api_ai_analyze_store(request: Request):
    if not ANTHROPIC_API_KEY and not GROQ_API_KEY:
        return JSONResponse({"ok": False, "error": "No AI API key configured."})
    body   = await request.json()
    result = await analyze_store(
        url=body.get("url", ""), platform=body.get("platform", ""),
        product_count=body.get("product_count", 0), samples=body.get("samples", []),
        anthropic_api_key=ANTHROPIC_API_KEY, groq_api_key=GROQ_API_KEY,
    )
    return JSONResponse(result)


# ── DB admin routes ────────────────────────────────────────────────────────────

ADMIN_PIN = os.getenv("ADMIN_PIN", "1234")

@app.get("/api/db/stats")
def db_stats():
    try:
        size_mb = round(os.path.getsize(DB_PATH) / 1024 / 1024, 2)
    except Exception:
        size_mb = 0
    conn = db_connect(DB_PATH)
    cur  = conn.cursor()
    try:
        cur.execute("SELECT COUNT(*) FROM snapshots")
        total = cur.fetchone()[0]
        cur.execute("SELECT store, COUNT(*) as rows, MAX(captured_at) as last_refresh FROM snapshots GROUP BY store ORDER BY rows DESC")
        stores = [{"store": r[0], "rows": r[1], "last_refresh": r[2]} for r in cur.fetchall()]
        return JSONResponse({"db_size_mb": size_mb, "total_snapshots": total, "stores": stores})
    except Exception as e:
        return JSONResponse({"error": str(e)})
    finally:
        conn.close()


@app.post("/api/db/query")
async def db_custom_query(request: Request):
    body      = await request.json()
    sql       = (body.get("sql") or "").strip()
    pin       = (body.get("pin") or "").strip()
    if not sql:
        return JSONResponse({"error": "No query provided"})
    sql_upper = sql.upper().lstrip()
    is_write  = any(sql_upper.startswith(k) for k in ("UPDATE", "DELETE", "INSERT", "DROP", "ALTER", "CREATE"))
    if is_write:
        if pin != ADMIN_PIN:
            return JSONResponse({"error": "❌ Admin PIN required for write operations", "pin_required": True})
    elif not (sql_upper.startswith("SELECT") or sql_upper.startswith("WITH")):
        return JSONResponse({"error": "Only SELECT queries are allowed without admin PIN"})
    conn = db_connect(DB_PATH)
    cur  = conn.cursor()
    try:
        cur.execute(sql)
        conn.commit()
        columns  = [d[0] for d in cur.description] if cur.description else []
        rows     = [list(r) for r in cur.fetchmany(200)] if cur.description else []
        affected = cur.rowcount if is_write else None
        return JSONResponse({"columns": columns, "rows": rows, "affected": affected})
    except Exception as e:
        return JSONResponse({"error": str(e)})
    finally:
        conn.close()


# ── Export ─────────────────────────────────────────────────────────────────────

@app.get("/api/export")
def api_export(request: Request):
    """Export filtered dashboard data as CSV, Excel, or JSON."""
    import csv
    from collections import defaultdict
    from core.utils import order_stores as _order_stores

    qp               = dict(request.query_params)
    fmt              = qp.get("fmt", "csv").lower()
    store_names      = get_all_store_names()
    raw_stores       = request.query_params.getlist("stores")
    stores           = _order_stores([s for s in raw_stores if s in store_names]) or store_names
    selected_category = qp.get("category", "All")
    selected_theme    = qp.get("theme",    "All")
    selected_brand    = qp.get("brand",    "All")
    search_item       = qp.get("search_item", "").strip()
    only_deals        = qp.get("only_deals",   "0") == "1"
    only_instock      = qp.get("only_instock", "0") == "1"

    conn = db_connect(DB_PATH)
    cur  = conn.cursor()
    ph   = ",".join("?" * len(stores))
    cur.execute(f"""
        SELECT s.store, s.item_number, s.title, s.brand, s.theme, s.category,
               s.price, s.compare_at, s.availability, s.link
        FROM snapshots s
        INNER JOIN (SELECT store, MAX(captured_at) as max_at FROM snapshots GROUP BY store) latest
        ON s.store=latest.store AND s.captured_at=latest.max_at
        WHERE s.store IN ({ph}) ORDER BY s.item_number, s.store
    """, stores)
    db_rows = cur.fetchall()
    conn.close()

    items = defaultdict(lambda: {"item_number": "", "brand": "", "title": "", "theme": "", "category": "", "stores": {}})
    for r in db_rows:
        key = r["item_number"] or r["title"] or ""
        it  = items[key]
        it["item_number"] = r["item_number"] or ""
        it["brand"]    = (r["brand"] or "").strip().upper()
        it["title"]    = r["title"] or ""
        it["theme"]    = r["theme"] or ""
        it["category"] = r["category"] or ""
        disc_pct = 0
        if r["compare_at"] and r["price"] and r["compare_at"] > r["price"]:
            disc_pct = round((r["compare_at"] - r["price"]) / r["compare_at"] * 100, 1)
        it["stores"][r["store"]] = {"price": r["price"], "availability": r["availability"] or "", "discount_pct": disc_pct, "link": r["link"] or ""}

    rows_out = []
    for it in items.values():
        if selected_category != "All" and it["category"] != selected_category: continue
        if selected_theme     != "All" and it["theme"]    != selected_theme:    continue
        if selected_brand     != "All" and it["brand"]    != selected_brand.upper(): continue
        if search_item and search_item not in it["item_number"]: continue
        if only_deals   and not any((v["discount_pct"] or 0) > 0 for v in it["stores"].values()): continue
        if only_instock and not any(v["availability"] in ("available", "in_stock") for v in it["stores"].values()): continue
        prices = [v["price"] for v in it["stores"].values() if v["price"] is not None]
        lp     = min(prices) if prices else None
        row    = {"item_number": it["item_number"], "brand": it["brand"], "title": it["title"],
                  "theme": it["theme"], "category": it["category"],
                  "lowest_price": f"{lp:.2f}" if lp is not None else ""}
        for s in stores:
            off = it["stores"].get(s, {})
            row[f"{s}_price"]        = f"{off['price']:.2f}" if off.get("price") is not None else ""
            row[f"{s}_avail"]        = off.get("availability", "")
            row[f"{s}_discount_pct"] = f"{off['discount_pct']:.0f}" if off.get("discount_pct") else ""
            row[f"{s}_link"]         = off.get("link", "")
        rows_out.append(row)

    if not rows_out:
        return JSONResponse({"error": "No data for current filters"})

    headers_row = list(rows_out[0].keys())
    ts          = datetime.now().strftime("%Y%m%d_%H%M")

    if fmt == "csv":
        buf = io.StringIO()
        w   = csv.DictWriter(buf, fieldnames=headers_row)
        w.writeheader(); w.writerows(rows_out); buf.seek(0)
        return StreamingResponse(iter([buf.getvalue().encode("utf-8")]),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f"attachment; filename=lego_tracker_{ts}.csv"})

    if fmt == "xlsx":
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "LEGO Tracker"
        for ci, h in enumerate(headers_row, 1):
            cell = ws.cell(row=1, column=ci, value=h.replace("_", " ").title())
            cell.fill = PatternFill("solid", fgColor="1E293B")
            cell.font = Font(bold=True, color="94A3B8", size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        deal_fill = PatternFill("solid", fgColor="1A3020")
        for ri, row in enumerate(rows_out, 2):
            for ci, key in enumerate(headers_row, 1):
                val = row[key]
                if key.endswith("_price") or key == "lowest_price":
                    try: val = float(val)
                    except: pass
                elif key.endswith("_discount_pct"):
                    try: val = int(val)
                    except: pass
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.font = Font(size=9)
                if key == "lowest_price" and val:
                    cell.font = Font(size=9, bold=True, color="F59E0B")
                if key.endswith("_discount_pct") and val:
                    cell.fill = deal_fill
        ws.row_dimensions[1].height = 28
        ws.freeze_panes = "A2"
        col_widths = {"item_number": 14, "brand": 10, "title": 42, "theme": 18, "category": 16, "lowest_price": 13}
        for ci, key in enumerate(headers_row, 1):
            ws.column_dimensions[get_column_letter(ci)].width = col_widths.get(key, 16 if "link" in key else 11)
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        return StreamingResponse(iter([buf.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=lego_tracker_{ts}.xlsx"})

    return JSONResponse(rows_out)


@app.get("/export/pdf", response_class=HTMLResponse)
def export_pdf_page(request: Request):
    """Print-ready HTML page — open and Ctrl+P to save as PDF."""
    from collections import defaultdict
    from core.utils import order_stores as _order_stores

    qp               = dict(request.query_params)
    store_names      = get_all_store_names()
    raw_stores       = request.query_params.getlist("stores")
    stores           = _order_stores([s for s in raw_stores if s in store_names]) or store_names
    selected_category = qp.get("category", "All")
    selected_theme    = qp.get("theme",    "All")
    selected_brand    = qp.get("brand",    "All")
    search_item       = qp.get("search_item", "").strip()
    only_deals        = qp.get("only_deals",   "0") == "1"
    only_instock      = qp.get("only_instock", "0") == "1"

    conn = db_connect(DB_PATH)
    cur  = conn.cursor()
    ph   = ",".join("?" * len(stores))
    cur.execute(f"""
        SELECT s.store, s.item_number, s.title, s.brand, s.theme, s.category,
               s.price, s.compare_at, s.availability, s.link
        FROM snapshots s
        INNER JOIN (SELECT store, MAX(captured_at) as max_at FROM snapshots GROUP BY store) latest
        ON s.store=latest.store AND s.captured_at=latest.max_at
        WHERE s.store IN ({ph}) ORDER BY s.item_number, s.store
    """, stores)
    db_rows = cur.fetchall()
    conn.close()

    items = defaultdict(lambda: {"item_number": "", "brand": "", "title": "", "theme": "", "category": "", "stores": {}})
    for r in db_rows:
        key = r["item_number"] or r["title"] or ""
        it  = items[key]
        it["item_number"] = r["item_number"] or ""
        it["brand"]    = (r["brand"] or "").strip().upper()
        it["title"]    = r["title"] or ""
        it["theme"]    = r["theme"] or ""
        it["category"] = r["category"] or ""
        disc_pct = 0
        if r["compare_at"] and r["price"] and r["compare_at"] > r["price"]:
            disc_pct = round((r["compare_at"] - r["price"]) / r["compare_at"] * 100, 1)
        it["stores"][r["store"]] = {"price": r["price"], "discount_pct": disc_pct, "availability": r["availability"] or ""}

    rows_out = []
    for it in items.values():
        if selected_category != "All" and it["category"] != selected_category: continue
        if selected_theme     != "All" and it["theme"]    != selected_theme:    continue
        if selected_brand     != "All" and it["brand"]    != selected_brand.upper(): continue
        if search_item and search_item not in it["item_number"]: continue
        if only_deals   and not any((v["discount_pct"] or 0) > 0 for v in it["stores"].values()): continue
        if only_instock and not any(v["availability"] in ("available", "in_stock") for v in it["stores"].values()): continue
        prices = [v["price"] for v in it["stores"].values() if v["price"] is not None]
        it["lowest_price"] = min(prices) if prices else None
        rows_out.append(it)

    def store_cells(it):
        cells = ""
        for s in stores:
            off = it["stores"].get(s)
            if off and off["price"] is not None:
                disc   = f' <span style="color:#b45309">-{off["discount_pct"]:.0f}%</span>' if off["discount_pct"] else ""
                cells += f'<td>${off["price"]:.2f}{disc}</td>'
            else:
                cells += "<td>—</td>"
        return cells

    store_headers = "".join(f"<th>{s}</th>" for s in stores)
    ts            = datetime.now().strftime("%Y-%m-%d %H:%M")
    rows_html     = ""
    for it in rows_out:
        lp = f'<b>${it["lowest_price"]:.2f}</b>' if it["lowest_price"] else "—"
        rows_html += f"""<tr>
          <td>{it["item_number"]}</td><td>{it["brand"]}</td><td>{it["title"]}</td>
          <td>{it["theme"]}</td><td>{it["category"]}</td>
          {store_cells(it)}<td>{lp}</td></tr>"""

    filters_desc = " | ".join(filter(None, [
        f"Stores: {', '.join(stores)}" if stores != store_names else "",
        f"Category: {selected_category}" if selected_category != "All" else "",
        f"Theme: {selected_theme}" if selected_theme != "All" else "",
        f"Brand: {selected_brand}" if selected_brand != "All" else "",
        f"Search: {search_item}" if search_item else "",
        "Deals only" if only_deals else "",
        "In stock only" if only_instock else "",
    ])) or "All products"

    return f"""<!doctype html><html><head><meta charset="UTF-8">
<title>LEGO Tracker Export — {ts}</title>
<style>
  body{{font-family:Arial,sans-serif;font-size:8pt;color:#111;margin:0;padding:1cm}}
  h1{{font-size:13pt;margin:0 0 2px}}.meta{{font-size:7.5pt;color:#666;margin-bottom:8px}}
  table{{width:100%;border-collapse:collapse}}
  th{{background:#1e293b;color:#fff;padding:4px 6px;text-align:left;font-size:7.5pt}}
  td{{padding:3px 6px;border-bottom:1px solid #e5e7eb;vertical-align:top}}
  tr:nth-child(even) td{{background:#f9fafb}}
  @page{{margin:1.5cm;size:A4 landscape}}
  @media screen{{body{{max-width:1400px;margin:auto;padding:20px}}}}
  .print-btn{{margin-bottom:12px;padding:8px 18px;background:#4f46e5;color:#fff;border:none;border-radius:6px;cursor:pointer;font-size:10pt}}
  @media print{{.print-btn{{display:none}}}}
</style></head><body>
<button class="print-btn" onclick="window.print()">&#128196; Print / Save as PDF</button>
<h1>LEGO Tracker — Product Export</h1>
<div class="meta">Generated: {ts} &nbsp;|&nbsp; {len(rows_out)} products &nbsp;|&nbsp; {filters_desc}</div>
<table><thead><tr><th>Item #</th><th>Brand</th><th>Title</th><th>Theme</th><th>Category</th>{store_headers}<th>Lowest</th></tr></thead>
<tbody>{rows_html}</tbody></table>
</body></html>"""


@app.post("/api/ai/discover-stores")
async def api_discover_stores(request: Request):
    if not ANTHROPIC_API_KEY and not GROQ_API_KEY:
        return JSONResponse({"ok": False, "error": "No AI API key configured."}, status_code=500)
    body   = await request.json()
    region = body.get("region", "Lebanon").strip()
    return StreamingResponse(
        stream_discover_stores(region, list(HARDCODED_STORE_URLS.keys()), ANTHROPIC_API_KEY, GROQ_API_KEY),
        media_type="text/event-stream",
    )
